import json
import secrets
import os
import calendar
import sys
import time
import textwrap
import hashlib
import sqlite3
from datetime import datetime

# --- INTERFAZ GRÁFICA ---
import tkinter as tk
from tkinter import ttk 
from tkinter import messagebox, filedialog, colorchooser, CENTER, LEFT, RIGHT, TOP, BOTTOM, BOTH, X, Y, VERTICAL, HORIZONTAL, W, E, END, NW, NO

import ttkbootstrap as tb 
from ttkbootstrap.constants import *
from ttkbootstrap.widgets import ToolTip, DateEntry

# --- IMÁGENES ---
from PIL import Image, ImageTk

# --- EXCEL ---
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side as ExcelSide

# --- GRÁFICAS Y PDF (MATPLOTLIB) ---
import matplotlib


matplotlib.use("TkAgg") # Configura el backend para interfaz gráfica

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import numpy as np


import matplotlib.image as mpimg       # Para leer el logo en el PDF
from matplotlib.patches import Rectangle # Para dibujar los cuadros azules en el PDF



# --- RUTAS DEL SISTEMA ---
# Detectar carpeta segura para guardar la Base de Datos y Configuración
appdata = os.getenv('APPDATA')
carpeta_configuracion = os.path.join(appdata, "SistemaInventarioUNINDETEC_SQL")

if not os.path.exists(carpeta_configuracion):
    try:
        os.makedirs(carpeta_configuracion)
    except:
        carpeta_configuracion = os.path.join(os.getenv('TEMP'), "SistemaInventarioUNINDETEC_SQL")
        os.makedirs(carpeta_configuracion, exist_ok=True)

RUTA_DB = os.path.join(carpeta_configuracion, "inventario_master.db")
ARCHIVO_UI_CONFIG = os.path.join(carpeta_configuracion, "config_ui.json")

class GestorBaseDatos:
    def __init__(self, ruta_db):
        """
        Inicializa el gestor con la ruta específica (Local o NAS).
        Si la carpeta no existe, intenta crearla.
        """
        self.ruta_db = ruta_db
        
        # Asegurar que el directorio exista
        directorio = os.path.dirname(ruta_db)
        if directorio and not os.path.exists(directorio):
            try:
                os.makedirs(directorio)
            except Exception as e:
                print(f"Error creando directorio DB: {e}")

        # Ejecutar creación/actualización de tablas al iniciar
        self.crear_tablas()

    def conectar(self):
        conn = sqlite3.connect(self.ruta_db, timeout=10)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")   # Evita bloqueos en escrituras simultáneas
        conn.execute("PRAGMA foreign_keys=ON")
        return conn

    def crear_tablas(self):
        """Crea las tablas y aplica migraciones"""
        sql_script = """
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            rol TEXT DEFAULT 'OPERADOR',
            nombre_completo TEXT DEFAULT 'Usuario del Sistema',
            email TEXT DEFAULT '@unindetec.edu.mx',
            foto_path TEXT DEFAULT '',
            permisos TEXT DEFAULT '{}'
        );
        CREATE TABLE IF NOT EXISTS inventario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            partida TEXT,
            material TEXT,
            factura TEXT DEFAULT 'S/F',
            stock REAL DEFAULT 0,
            ultimo_movimiento TEXT,
            UNIQUE(partida, material)
        );
        CREATE TABLE IF NOT EXISTS historial (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_hora TEXT,
            tipo TEXT,
            partida TEXT,
            material TEXT,
            cantidad REAL,
            factura TEXT,
            destino TEXT,
            responsable TEXT,
            entrego TEXT,
            usuario_sistema TEXT DEFAULT 'SISTEMA'
        );
        CREATE TABLE IF NOT EXISTS catalogos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT,
            valor TEXT
        );
        CREATE TABLE IF NOT EXISTS partidas_desc (
            codigo TEXT PRIMARY KEY,
            descripcion TEXT
        );
        CREATE TABLE IF NOT EXISTS config_sistema (
            clave TEXT PRIMARY KEY,
            valor TEXT
        );
        """
        try:
            with self.conectar() as conn:
                conn.executescript(sql_script)
                # Migraciones previas
                try: conn.execute("ALTER TABLE inventario ADD COLUMN factura TEXT DEFAULT 'S/F'")
                except: pass
                try: conn.execute("ALTER TABLE historial ADD COLUMN factura TEXT DEFAULT ''")
                except: pass
                # NUEVA MIGRACIÓN: columna usuario_sistema
                try: conn.execute("ALTER TABLE historial ADD COLUMN usuario_sistema TEXT DEFAULT 'SISTEMA'")
                except: pass
                # Lógica Admin
                conn.execute("UPDATE usuarios SET rol='ADMIN' WHERE usuario='ADMIN'")
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM usuarios")
                if cursor.fetchone()[0] == 0:
                    pass_hash = hashlib.sha256("admin".encode()).hexdigest()
                    p_full = '{"crear":1, "entrada":1, "salida":1, "editar":1, "eliminar":1, "catalogos":1, "historico":1, "ajustes":1}'
                    conn.execute("INSERT INTO usuarios (usuario, password, rol, permisos) VALUES (?, ?, ?, ?)",
                                 ("ADMIN", pass_hash, "ADMIN", p_full))
        except Exception as e:
            print(f"Error tablas: {e}")


    def validar_login(self, usuario, password):
        try:
            with self.conectar() as conn:
                cursor = conn.cursor()
            # Búsqueda EXACTA — respeta mayúsculas y minúsculas
                cursor.execute(
                    "SELECT * FROM usuarios WHERE usuario = ?",
                    (usuario.strip(),)
                )
                user = cursor.fetchone()

                if not user:
                    return None

                stored_hash = user['password']

                if ":" in stored_hash:
                    salt, hash_guardado = stored_hash.split(":", 1)
                    hash_intento = hashlib.pbkdf2_hmac(
                        'sha256',
                        password.encode(),
                        salt.encode(),
                        100000
                    ).hex()
                    valido = (hash_intento == hash_guardado)
                else:
                    hash_viejo = hashlib.sha256(password.encode()).hexdigest()
                    valido = (hash_viejo == stored_hash)

                    if valido:
                        salt = secrets.token_hex(16)
                        hash_nuevo = hashlib.pbkdf2_hmac(
                            'sha256', password.encode(), salt.encode(), 100000
                        ).hex()
                        nuevo_stored = f"{salt}:{hash_nuevo}"
                        conn.execute(
                            "UPDATE usuarios SET password = ? WHERE usuario = ?",
                            (nuevo_stored, usuario.strip())
                        )
                        conn.commit()
                        print(f"✅ Contraseña de {usuario} migrada al nuevo formato")

                return dict(user) if valido else None

        except Exception as e:
            print(f"Error en login: {e}")
            return None
        
    def get_config(self, clave):
        """Recupera configuraciones (Ruta de Logo, Título App, etc.)"""
        try:
            with self.conectar() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT valor FROM config_sistema WHERE clave = ?", (clave,))
                res = cursor.fetchone()
                return res['valor'] if res else None
        except:
            return None

    def set_config(self, clave, valor):
        """Guarda o actualiza una configuración"""
        try:
            with self.conectar() as conn:
                conn.execute("REPLACE INTO config_sistema (clave, valor) VALUES (?, ?)", (clave, valor))
                conn.commit()
        except Exception as e:
            print(f"Error guardando config {clave}: {e}")

    def consultar(self, query, params=()):
        """Helper para consultas SELECT"""
        try:
            with self.conectar() as conn:
                cursor = conn.cursor()
                cursor.execute(query, params)
                return cursor.fetchall()
        except Exception as e:
            print(f"Error SQL (SELECT): {e}")
            return []

    def ejecutar(self, query, params=()):
        """Helper para INSERT, UPDATE, DELETE"""
        try:
            with self.conectar() as conn:
                conn.execute(query, params)
                conn.commit()
        except Exception as e:
            print(f"Error SQL (EXEC): {e}")
            raise e # Lanzamos el error para que la interfaz muestre el mensaje
class GestorTemas:
    TEMAS_PREDEFINIDOS = {
        "Azul Profesional": {
            "color_primario":   "#1F4E79",
            "color_secundario": "#4472C4",
            "color_acento":     "#00B0F0",
            "color_fondo":      "#FFFFFF",
            "color_texto":      "#000000",
            "tema_bootstrap":   "flatly"
        },
        "Verde Moderno": {
            "color_primario":   "#2E7D32",
            "color_secundario": "#66BB6A",
            "color_acento":     "#4CAF50",
            "color_fondo":      "#FAFAFA",
            "color_texto":      "#212121",
            "tema_bootstrap":   "minty"
        },
        "Oscuro": {
            "color_primario":   "#212121",
            "color_secundario": "#424242",
            "color_acento":     "#FFC107",
            "color_fondo":      "#303030",
            "color_texto":      "#FFFFFF",
            "tema_bootstrap":   "darkly"
        },
        "Rojo Corporativo": {
            "color_primario":   "#C00000",
            "color_secundario": "#E74C3C",
            "color_acento":     "#FF6B6B",
            "color_fondo":      "#FFFFFF",
            "color_texto":      "#2C3E50",
            "tema_bootstrap":   "journal"
        },
        "Morado Creativo": {
            "color_primario":   "#6A1B9A",
            "color_secundario": "#9C27B0",
            "color_acento":     "#BA68C8",
            "color_fondo":      "#F5F5F5",
            "color_texto":      "#212121",
            "tema_bootstrap":   "pulse"
        },
        "Naranja Energético": {
            "color_primario":   "#E65100",
            "color_secundario": "#FF6F00",
            "color_acento":     "#FFB300",
            "color_fondo":      "#FFFFFF",
            "color_texto":      "#212121",
            "tema_bootstrap":   "sandstone"
        }
    }

    @staticmethod
    def get_tema_actual(db_manager):
        return {
            "color_primario":   db_manager.get_config("TEMA_COLOR_PRIMARIO")   or "#1F4E79",
            "color_secundario": db_manager.get_config("TEMA_COLOR_SECUNDARIO") or "#4472C4",
            "color_acento":     db_manager.get_config("TEMA_COLOR_ACENTO")     or "#00B0F0",
            "color_fondo":      db_manager.get_config("TEMA_COLOR_FONDO")      or "#FFFFFF",
            "color_texto":      db_manager.get_config("TEMA_COLOR_TEXTO")      or "#000000",
            "tema_bootstrap":   db_manager.get_config("TEMA_BOOTSTRAP")        or "flatly"
        }

    @staticmethod
    def guardar_tema(db_manager, tema):
        db_manager.set_config("TEMA_COLOR_PRIMARIO",   tema["color_primario"])
        db_manager.set_config("TEMA_COLOR_SECUNDARIO", tema["color_secundario"])
        db_manager.set_config("TEMA_COLOR_ACENTO",     tema["color_acento"])
        db_manager.set_config("TEMA_COLOR_FONDO",      tema["color_fondo"])
        db_manager.set_config("TEMA_COLOR_TEXTO",      tema["color_texto"])
        db_manager.set_config("TEMA_BOOTSTRAP",        tema["tema_bootstrap"])

    @staticmethod
    def _luminancia(hex_color):
        """Calcula si el color es claro u oscuro"""
        try:
            h = hex_color.lstrip("#")
            r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
            return (0.299*r + 0.587*g + 0.114*b) / 255
        except:
            return 1.0

    @staticmethod
    def aplicar_estilos_completos(root, tema):
        """
        Aplica el tema a TODO el sistema:
        1. Cambia el tema bootstrap base
        2. Sobreescribe colores en ttk.Style para cada widget
        """
        c_prim  = tema["color_primario"]
        c_sec   = tema["color_secundario"]
        c_acent = tema["color_acento"]
        c_fondo = tema["color_fondo"]
        c_texto = tema["color_texto"]
        t_boot  = tema["tema_bootstrap"]

        # Texto legible sobre fondo primario
        txt_prim = "#FFFFFF" if GestorTemas._luminancia(c_prim) < 0.5 else "#000000"
        # Texto legible sobre fondo general
        txt_fnd  = c_texto

        try:
            # 1. CAMBIAR TEMA BOOTSTRAP (maneja la base de toda la UI)
            root.style.theme_use(t_boot)
        except Exception as e:
            print(f"Error cambiando tema bootstrap: {e}")

        style = ttk.Style()

        # 2. FONDO DE LA VENTANA PRINCIPAL
        try:
            root.configure(bg=c_fondo)
        except:
            pass

        # 3. ESTILOS GLOBALES
        style.configure(".",
                         background=c_fondo,
                         foreground=txt_fnd,
                         bordercolor=c_sec,
                         darkcolor=c_prim,
                         lightcolor=c_sec,
                         troughcolor=c_fondo,
                         selectbackground=c_prim,
                         selectforeground=txt_prim,
                         insertcolor=c_texto,
                         font=("Segoe UI", 10))

        # 4. FRAMES Y CONTENEDORES
        style.configure("TFrame",      background=c_fondo)
        style.configure("TLabelframe", background=c_fondo, bordercolor=c_sec)
        style.configure("TLabelframe.Label",
                         background=c_fondo,
                         foreground=c_prim,
                         font=("Segoe UI", 10, "bold"))

        # 5. ETIQUETAS
        style.configure("TLabel", background=c_fondo, foreground=txt_fnd)

        # 6. BOTONES
        style.configure("TButton",
                         background=c_prim,
                         foreground=txt_prim,
                         borderwidth=0,
                         focusthickness=2,
                         padding=6,
                         font=("Segoe UI", 10, "bold"))
        style.map("TButton",
                  background=[("active",   c_sec),
                               ("pressed",  c_prim),
                               ("disabled", "#CCCCCC")],
                  foreground=[("active",   txt_prim),
                               ("disabled", "#888888")])

        # 7. ENTRADAS DE TEXTO
        style.configure("TEntry",
                         fieldbackground="#FFFFFF" if GestorTemas._luminancia(c_fondo) > 0.5 else "#404040",
                         foreground=txt_fnd,
                         bordercolor=c_sec,
                         insertcolor=c_texto)
        style.map("TEntry",
                  bordercolor=[("focus", c_acent)],
                  lightcolor=[("focus", c_acent)],
                  darkcolor=[("focus",  c_acent)])

        # 8. COMBOBOX
        style.configure("TCombobox",
                         fieldbackground="#FFFFFF" if GestorTemas._luminancia(c_fondo) > 0.5 else "#404040",
                         background=c_prim,
                         foreground=txt_fnd,
                         selectbackground=c_prim,
                         selectforeground=txt_prim,
                         arrowcolor=c_prim)
        style.map("TCombobox",
                  fieldbackground=[("readonly", c_fondo)],
                  bordercolor=[("focus", c_acent)])

        # 9. TREEVIEW (TABLAS)
        style.configure("Treeview",
                         background="#FFFFFF" if GestorTemas._luminancia(c_fondo) > 0.5 else "#3a3a3a",
                         foreground=txt_fnd,
                         fieldbackground="#FFFFFF" if GestorTemas._luminancia(c_fondo) > 0.5 else "#3a3a3a",
                         rowheight=30)
        style.configure("Treeview.Heading",
                         background=c_prim,
                         foreground=txt_prim,
                         font=("Segoe UI", 10, "bold"),
                         relief="flat")
        style.map("Treeview",
                  background=[("selected", c_acent)],
                  foreground=[("selected", texto_sobre_color(c_acent))])
        style.map("Treeview.Heading",
                  background=[("active", c_sec)])

        # 10. NOTEBOOK (PESTAÑAS)
        style.configure("TNotebook",
                         background=c_fondo,
                         bordercolor=c_sec)
        style.configure("TNotebook.Tab",
                         background=c_fondo,
                         foreground=txt_fnd,
                         padding=[12, 6],
                         font=("Segoe UI", 10, "bold"))
        style.map("TNotebook.Tab",
                  background=[("selected", c_prim),
                               ("active",   c_sec)],
                  foreground=[("selected", txt_prim),
                               ("active",   txt_prim)])

        # 11. SCROLLBAR
        style.configure("TScrollbar",
                         background=c_sec,
                         troughcolor=c_fondo,
                         arrowcolor=txt_fnd,
                         bordercolor=c_fondo)
        style.map("TScrollbar",
                  background=[("active", c_prim)])

        # 12. SEPARADORES
        style.configure("TSeparator", background=c_sec)

        # 13. PROGRESSBAR
        style.configure("TProgressbar",
                         background=c_acent,
                         troughcolor=c_fondo,
                         bordercolor=c_fondo)

        # 14. CHECKBUTTON
        style.configure("TCheckbutton",
                         background=c_fondo,
                         foreground=txt_fnd)
        style.map("TCheckbutton",
                  background=[("active", c_fondo)],
                  indicatorcolor=[("selected", c_prim)])

        print(f"✅ Tema '{t_boot}' aplicado correctamente.")
    
def texto_sobre_color(hex_color):
    """Devuelve blanco o negro según la luminosidad del color de fondo"""
    try:
        h = hex_color.lstrip("#")
        r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
        luminancia = (0.299*r + 0.587*g + 0.114*b) / 255
        return "#FFFFFF" if luminancia < 0.5 else "#000000"
    except:
        return "#000000"
    
class LoginWindow(tk.Toplevel):
    def __init__(self, parent, db_manager, on_success):
        super().__init__(parent)
        self.title("Acceso al Sistema")
        self.withdraw()

        self.db       = db_manager
        self.callback = on_success
        self.protocol("WM_DELETE_WINDOW", self.cancelar_login)
        self.resizable(False, False)

        # ── Datos de configuración ────────────────────────────────────
        self.titulo   = self.db.get_config("TITULO_APP")    or "SISTEMA"
        self.subtitulo = self.db.get_config("SUBTITULO_APP") or "CONTROL DE INVENTARIO"
        self.logo_path = self.db.get_config("LOGO_APP")

        if (not self.logo_path or
                not os.path.exists(self.logo_path)):
            self.logo_path = ("logo.png"
                               if os.path.exists("logo.png")
                               else None)

        # ── Colores del tema ──────────────────────────────────────────
        tema        = GestorTemas.get_tema_actual(self.db)
        self.C_PRIM = tema.get("color_primario",   "#1F4E79")
        self.C_SEC  = tema.get("color_secundario", "#4472C4")
        self.C_ACNT = tema.get("color_acento",     "#00B0F0")

        # Oscurecer primario para degradado
        def oscurecer(hex_c, factor=0.6):
            h = hex_c.lstrip("#")
            r = int(int(h[0:2], 16) * factor)
            g = int(int(h[2:4], 16) * factor)
            b = int(int(h[4:6], 16) * factor)
            return f"#{r:02x}{g:02x}{b:02x}"

        self.C_DARK = oscurecer(self.C_PRIM, 0.55)

        # ── Tamaño fijo ───────────────────────────────────────────────
        WIN_W     = 900
        WIN_H     = 540
        PANEL_IZQ = 420

        self.configure(bg=self.C_PRIM)

        # ════════════════════════════════════════════
        # PANEL IZQUIERDO
        # ════════════════════════════════════════════
        self.canvas_izq = tk.Canvas(
            self, width=PANEL_IZQ, height=WIN_H,
            highlightthickness=0, bd=0)
        self.canvas_izq.place(x=0, y=0)

        self._dibujar_panel_izquierdo(PANEL_IZQ, WIN_H)

        # ════════════════════════════════════════════
        # PANEL DERECHO
        # ════════════════════════════════════════════
        fr_der = tk.Frame(
            self, bg="#FFFFFF",
            width=WIN_W - PANEL_IZQ, height=WIN_H)
        fr_der.place(x=PANEL_IZQ, y=0)
        fr_der.pack_propagate(False) # ── Centrar y mostrar

        # Contenedor centrado verticalmente
        fr_form = tk.Frame(fr_der, bg="#FFFFFF")
        fr_form.place(relx=0.5, rely=0.5, anchor="center")

        # ── Icono de candado ──────────────────────────────────────────
        lbl_icono = tk.Label(
            fr_form, text="🔐",
            font=("Segoe UI Emoji", 42),
            bg="#FFFFFF")
        lbl_icono.pack(pady=(0, 8))

        # ── Título del formulario ─────────────────────────────────────
        tk.Label(
            fr_form,
            text="INICIAR SESIÓN",
            font=("Segoe UI", 18, "bold"),
            fg=self.C_PRIM, bg="#FFFFFF"
        ).pack(pady=(0, 4))

        tk.Label(
            fr_form,
            text="Ingresa tus credenciales para continuar",
            font=("Segoe UI", 9),
            fg="#888888", bg="#FFFFFF"
        ).pack(pady=(0, 28))

        # ── Campo Usuario ─────────────────────────────────────────────
        fr_u = tk.Frame(fr_form, bg="#FFFFFF")
        fr_u.pack(fill=X, pady=(0, 12))

        tk.Label(
            fr_u, text="USUARIO",
            font=("Segoe UI", 8, "bold"),
            fg="#555555", bg="#FFFFFF"
        ).pack(anchor=W)

        fr_e_u = tk.Frame(
            fr_u, bg="#F0F4F8",
            highlightbackground=self.C_SEC,
            highlightthickness=1)
        fr_e_u.pack(fill=X, ipady=6)

        tk.Label(
            fr_e_u, text=" 👤 ",
            font=("Segoe UI Emoji", 12),
            bg="#F0F4F8", fg=self.C_PRIM
        ).pack(side=LEFT)

        self.entry_user = tk.Entry(
            fr_e_u,
            font=("Segoe UI", 12),
            bg="#F0F4F8", fg="#222222",
            relief="flat", bd=0,
            insertbackground=self.C_PRIM,
            width=22)
        self.entry_user.pack(
            side=LEFT, fill=X, expand=True, padx=(0, 8))

        # Hover sobre campo usuario
        def on_enter_u(e):
            fr_e_u.configure(
                highlightbackground=self.C_PRIM,
                highlightthickness=2)
        def on_leave_u(e):
            fr_e_u.configure(
                highlightbackground=self.C_SEC,
                highlightthickness=1)
        fr_e_u.bind("<Enter>", on_enter_u)
        fr_e_u.bind("<Leave>", on_leave_u)
        self.entry_user.bind("<Enter>", on_enter_u)
        self.entry_user.bind("<Leave>", on_leave_u)

        # ── Campo Contraseña ──────────────────────────────────────────
        fr_p = tk.Frame(fr_form, bg="#FFFFFF")
        fr_p.pack(fill=X, pady=(0, 8))

        tk.Label(
            fr_p, text="CONTRASEÑA",
            font=("Segoe UI", 8, "bold"),
            fg="#555555", bg="#FFFFFF"
        ).pack(anchor=W)

        fr_e_p = tk.Frame(
            fr_p, bg="#F0F4F8",
            highlightbackground=self.C_SEC,
            highlightthickness=1)
        fr_e_p.pack(fill=X, ipady=6)

        tk.Label(
            fr_e_p, text=" 🔑 ",
            font=("Segoe UI Emoji", 12),
            bg="#F0F4F8", fg=self.C_PRIM
        ).pack(side=LEFT)

        self.entry_pass = tk.Entry(
            fr_e_p,
            font=("Segoe UI", 12),
            bg="#F0F4F8", fg="#222222",
            relief="flat", bd=0,
            show="●",
            insertbackground=self.C_PRIM,
            width=18)
        self.entry_pass.pack(
            side=LEFT, fill=X, expand=True)

        # Ojo para ver contraseña
        self._ver_pass = False
        self.btn_ojo = tk.Label(
            fr_e_p, text=" 👁 ",
            font=("Segoe UI Emoji", 12),
            bg="#F0F4F8", fg="#AAAAAA",
            cursor="hand2")
        self.btn_ojo.pack(side=RIGHT, padx=4)
        self.btn_ojo.bind("<Button-1>",
                           lambda e: self._toggle_pass())

        def on_enter_p(e):
            fr_e_p.configure(
                highlightbackground=self.C_PRIM,
                highlightthickness=2)
        def on_leave_p(e):
            fr_e_p.configure(
                highlightbackground=self.C_SEC,
                highlightthickness=1)
        fr_e_p.bind("<Enter>", on_enter_p)
        fr_e_p.bind("<Leave>", on_leave_p)
        self.entry_pass.bind("<Enter>", on_enter_p)
        self.entry_pass.bind("<Leave>", on_leave_p)

        # ── Nota mayúsculas ───────────────────────────────────────────
        tk.Label(
            fr_form,
            text="",
            font=("Segoe UI", 8),
            fg="#AAAAAA", bg="#FFFFFF"
        ).pack(pady=(0, 20))

        # ── Botón Ingresar ────────────────────────────────────────────
        self.btn_login = tk.Button(
            fr_form,
            text="  INGRESAR AL SISTEMA  ",
            font=("Segoe UI", 11, "bold"),
            bg=self.C_PRIM, fg="#FFFFFF",
            activebackground=self.C_SEC,
            activeforeground="#FFFFFF",
            relief="flat", bd=0,
            cursor="hand2",
            pady=10,
            command=self.entrar)
        self.btn_login.pack(fill=X, pady=(0, 16))

        # Hover botón
        self.btn_login.bind(
            "<Enter>",
            lambda e: self.btn_login.configure(
                bg=self.C_SEC))
        self.btn_login.bind(
            "<Leave>",
            lambda e: self.btn_login.configure(
                bg=self.C_PRIM))

        # ── Mensaje de error ──────────────────────────────────────────
        self.lbl_msg = tk.Label(
            fr_form, text="",
            font=("Segoe UI", 9, "bold"),
            fg="#E74C3C", bg="#FFFFFF")
        self.lbl_msg.pack()

        # ── Bindings de teclado ───────────────────────────────────────
        self.entry_user.bind("<Return>",
            lambda e: self.entry_pass.focus_set())
        self.entry_pass.bind("<Return>",
            lambda e: self.entrar())
        self.entry_user.bind("<Tab>",
            lambda e: (self.entry_pass.focus_set(),
                       "break")[1])

        # ── Centrar y mostrar ─────────────────────────────────────────
        self.update_idletasks()
        ws = self.winfo_screenwidth()
        hs = self.winfo_screenheight()
        x  = (ws // 2) - (WIN_W // 2)
        y  = (hs // 2) - (WIN_H // 2)
        self.geometry(f"{WIN_W}x{WIN_H}+{x}+{y}")

        self.deiconify()
        self.entry_user.focus_set()

        # ── Animación de entrada ──────────────────────────────────────
        self._animar_entrada()

    # ── Métodos auxiliares ────────────────────────────────────────────
    def _dibujar_panel_izquierdo(self, w, h):
        """
        Panel izquierdo AUTO-ADAPTABLE con layout PROPORCIONAL FIJO.
        Logo grande (120px) conservando su forma/proporción original.
        """
        c = self.canvas_izq

        # ══════════════════════════════════════════════════════════════
        # A. FONDO DEGRADADO
        # ══════════════════════════════════════════════════════════════
        pasos = 60
        for i in range(pasos):
            t  = i / pasos
            r0 = int(int(self.C_DARK.lstrip("#")[0:2], 16) * (1 - t)
                     + int(self.C_PRIM.lstrip("#")[0:2], 16) * t)
            g0 = int(int(self.C_DARK.lstrip("#")[2:4], 16) * (1 - t)
                     + int(self.C_PRIM.lstrip("#")[2:4], 16) * t)
            b0 = int(int(self.C_DARK.lstrip("#")[4:6], 16) * (1 - t)
                     + int(self.C_PRIM.lstrip("#")[4:6], 16) * t)
            color = f"#{r0:02x}{g0:02x}{b0:02x}"
            y0 = int(h * i / pasos)
            y1 = int(h * (i + 1) / pasos)
            c.create_rectangle(0, y0, w, y1, fill=color, outline="")

        # ══════════════════════════════════════════════════════════════
        # B. CÍRCULOS DECORATIVOS + LÍNEAS
        # ══════════════════════════════════════════════════════════════
        def circulo(cx, cy, r, color):
            c.create_oval(cx - r, cy - r, cx + r, cy + r,
                          fill=color, outline="")

        circulo(w - 40,   -40,  120, self.C_SEC)
        circulo(-30,    h + 30, 100, self.C_DARK)
        circulo(w // 2, h // 2, 200, self.C_DARK)
        circulo(30,       80,    50, self.C_ACNT)
        circulo(w - 60, h - 80,  70, self.C_ACNT)

        for i in range(0, w + h, 45):
            c.create_line(i, 0, 0, i,
                          fill="#FFFFFF", width=1, stipple="gray12")

        # ══════════════════════════════════════════════════════════════
        # C. TAMAÑO DE FUENTE ADAPTABLE SEGÚN LONGITUD DEL TÍTULO
        # ══════════════════════════════════════════════════════════════
        titulo_texto = (self.titulo or "SISTEMA").upper()
        subtit_texto =  self.subtitulo or ""
        n = len(titulo_texto)

        if   n <= 12:   fs = 22
        elif n <= 20:   fs = 19
        elif n <= 32:   fs = 16
        elif n <= 46:   fs = 14
        else:           fs = 12

        # ══════════════════════════════════════════════════════════════
        # D. POSICIONES Y PROPORCIONALES A h
        # ══════════════════════════════════════════════════════════════
        logo_max    = 200           # tamaño MÁXIMO del lado mayor del logo
        logo_cy     = int(h * 0.20) # centro Y del logo
        titulo_cy   = int(h * 0.52) # centro del bloque de título
        linea_y     = int(h * 0.64) # línea decorativa
        subtit_cy   = int(h * 0.73) # subtítulo
        pie_y       = h - 22        # pie

        # ══════════════════════════════════════════════════════════════
        # E. DIBUJAR ELEMENTOS
        # ══════════════════════════════════════════════════════════════

        # ── Logo (PROPORCIÓN ORIGINAL PRESERVADA) ─────────────────────
        # thumbnail() escala la imagen para que quepa en (logo_max x logo_max)
        # manteniendo el aspect ratio original — NUNCA aplasta ni estira.
        if self.logo_path:
            try:
                img = Image.open(self.logo_path)

                # thumbnail modifica en-place y respeta proporciones
                img_copia = img.copy()
                img_copia.thumbnail((logo_max, logo_max), Image.LANCZOS)

                self._img_logo = ImageTk.PhotoImage(img_copia)
                c.create_image(w // 2, logo_cy,
                               image=self._img_logo, anchor="center")
            except Exception:
                c.create_text(w // 2, logo_cy,
                              text="⚓",
                              font=("Segoe UI Emoji", 54),
                              fill="#FFFFFF", anchor="center")
        else:
            c.create_text(w // 2, logo_cy,
                          text="⚓",
                          font=("Segoe UI Emoji", 54),
                          fill="#FFFFFF", anchor="center")

        # ── Título principal ──────────────────────────────────────────
        c.create_text(
            w // 2, titulo_cy,
            text=titulo_texto,
            font=("Arial Black", fs),
            fill="#FFFFFF",
            anchor="center",
            width=w - 50,
            justify="center")

        # ── Línea decorativa de acento ────────────────────────────────
        largo = min(80, int(w * 0.20))
        c.create_rectangle(
            w // 2 - largo, linea_y,
            w // 2 + largo, linea_y + 4,
            fill=self.C_ACNT, outline="")

        # ── Subtítulo ─────────────────────────────────────────────────
        c.create_text(
            w // 2, subtit_cy,
            text=subtit_texto,
            font=("Segoe UI", 11),
            fill="#DDEEFF",
            anchor="center",
            width=w - 60,
            justify="center")

        # ── Pie de página / versión ───────────────────────────────────
        c.create_text(
            w // 2, pie_y,
            text="Sistema de Gestión de Inventario  •  v2.0",
            font=("Segoe UI", 8),
            fill="#AACCEE",
            anchor="center")
    def _toggle_pass(self):
        """Muestra u oculta la contraseña al pulsar el ojo"""
        self._ver_pass = not self._ver_pass
        if self._ver_pass:
            self.entry_pass.config(show="")
            self.btn_ojo.config(fg=self.C_PRIM)
        else:
            self.entry_pass.config(show="●")
            self.btn_ojo.config(fg="#AAAAAA")
            
    def _animar_entrada(self):
        """Efecto de aparición suave (fade-in simulado con alpha)"""
        try:
            self.attributes("-alpha", 0.0)
            def fade(alpha=0.0):
                if alpha < 1.0:
                    alpha = min(alpha + 0.07, 1.0)
                    self.attributes("-alpha", alpha)
                    self.after(18, lambda: fade(alpha))
            fade()
        except:
            pass

    def cancelar_login(self):
        if messagebox.askyesno(
                "Salir", "¿Deseas salir del sistema?"):
            sys.exit()

    def entrar(self):
        u = self.entry_user.get().strip()
        p = self.entry_pass.get().strip()

        if not u or not p:
            self.lbl_msg.configure(
                text="⚠  Ingresa usuario y contraseña.",
                fg="#E67E22")
            return

        # Feedback visual en el botón
        self.btn_login.configure(
            text="  Verificando...  ",
            bg="#888888", state="disabled")
        self.update()

        user_data = self.db.validar_login(u, p)

        if user_data:
            self.lbl_msg.configure(
                text="✅  Acceso correcto, cargando...",
                fg="#27AE60")
            self.btn_login.configure(
                text="  ✅  ACCESO CONCEDIDO  ",
                bg="#27AE60")
            self.after(
                600,
                lambda: [self.destroy(),
                          self.callback(user_data)])
        else:
            self.btn_login.configure(
                text="  INGRESAR AL SISTEMA  ",
                bg=self.C_PRIM, state="normal")
            self.lbl_msg.configure(
                text="❌  Usuario o contraseña incorrectos.",
                fg="#E74C3C")
            # Efecto shake en la ventana
            self._shake()
            self.entry_pass.delete(0, END)
            self.entry_pass.focus_set()

    def _shake(self):
        """Animación de sacudida cuando la contraseña es incorrecta"""
        x0 = self.winfo_x()
        y0 = self.winfo_y()
        movs = [8, -8, 6, -6, 4, -4, 2, -2, 0]

        def paso(i=0):
            if i < len(movs):
                self.geometry(
                    f"+{x0 + movs[i]}+{y0}")
                self.after(30, lambda: paso(i + 1))
            else:
                self.geometry(f"+{x0}+{y0}")

        paso()
            
class SistemaInventario:
    def __init__(self, root, db_manager, usuario_actual):
        self.root = root
        self.db = db_manager
        self.usuario = usuario_actual
        # Cargar tema actual
        self.tema_actual = GestorTemas.get_tema_actual(self.db)
        # Aplicar tema a toda la ventana principal
        GestorTemas.aplicar_estilos_completos(self.root, self.tema_actual)
        
        # LEER TITULO DE LA BASE DE DATOS
        titulo_ventana = self.db.get_config("TITULO_APP")
        if not titulo_ventana: 
            titulo_ventana = "Sistema de Inventario"
        
        self.root.title(f"{titulo_ventana} - [Usuario: {self.usuario['usuario']}]")
        self.animacion_actual = None
        
        # --- CAMBIO: INICIAR MAXIMIZADO ---
        try:
            self.root.state('zoomed') # Para Windows
        except:
            self.root.attributes('-zoomed', True) # Para Linux/Otros
            
        # Si por alguna razón falla el zoomed, usamos un tamaño base grande
        if self.root.state() != 'zoomed':
            self.root.geometry("1280x850")
        
        # Estilos visuales
        style = ttk.Style()
        style.configure("Treeview", rowheight=30)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        
        # CREAR INTERFAZ
        self.notebook = ttk.Notebook(self.root, bootstyle="primary")
        
        # Inicializar variable para el logo
        self.tk_logo = None
        
        # Llamamos al header
        self.setup_header()
        
        # Empaquetamos el notebook después del header
        self.notebook.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
       # Crear pestañas base
        self.tab_inv     = ttk.Frame(self.notebook, padding=10)
        self.tab_audit   = ttk.Frame(self.notebook, padding=10)
        self.tab_consumo = ttk.Frame(self.notebook, padding=10)
        self.tab_hist    = ttk.Frame(self.notebook, padding=10)

        rol_actual = self.usuario.get('rol', 'OPERADOR')

        if rol_actual == 'SOLO LECTURA':
            # Solo lectura: únicamente la pestaña de inventario (solo vista)
            self.notebook.add(self.tab_inv, text="📦 INVENTARIO")
            self.setup_tab_inventario()

        else:
            # Cualquier otro rol: acceso completo según permisos
            self.notebook.add(self.tab_inv,     text="📦 INVENTARIO")
            self.notebook.add(self.tab_audit,   text="📑 DATOS")
            self.notebook.add(self.tab_consumo, text="📈 CONSUMO")
            self.notebook.add(self.tab_hist,    text="🕒 HISTORIAL")

            self.setup_tab_inventario()
            self.setup_tab_auditoria()
            self.setup_tab_consumo()
            self.setup_tab_historial()

        # Carga inicial de datos
        self.actualizar_combos()
        self.cargar_tabla_inventario()

        # Solo cargar historial si la pestaña existe
        if rol_actual != 'SOLO LECTURA':
            self.cargar_tabla_historial()

        self.datos_kardex_procesados = []

   
    def _actualizar_filtros_visibles(self):
        """Muestra u oculta grupos de filtros según el tipo de reporte"""
        tipo = self.var_tipo_reporte.get()

        # Ocultar todo primero
        self.fr_grupo_fechas.pack_forget()
        self.fr_grupo_partida.pack_forget()

        if tipo == "PERIODO":
            # Solo rango de fechas
            self.fr_grupo_fechas.pack(side=LEFT, fill=X)

        elif tipo == "PERIODO_PARTIDA":
            # Fechas + partida
            self.fr_grupo_fechas.pack(side=LEFT, fill=X)
            self.fr_grupo_partida.pack(side=LEFT, fill=X, padx=(20, 0))

        elif tipo == "GENERAL_PARTIDA":
            # Solo partida (todo el historial)
            self.fr_grupo_partida.pack(side=LEFT, fill=X)

    def limpiar_filtros(self):
        """Limpia buscador y filtro de partida, recarga todo el inventario."""
        try:
            # Limpiar el Entry de búsqueda
            if hasattr(self, 'cb_busqueda_material'):
                self.cb_busqueda_material.delete(0, 'end')

            # Resetear filtro de partida a TODAS
            if hasattr(self, 'cb_filtro_partida'):
                try:
                    self.cb_filtro_partida.current(0)
                except:
                    pass

            # Recargar inventario completo
            self.cargar_tabla_inventario()

        except Exception as e:
            print(f"Advertencia al limpiar filtros: {e}")


    def cargar_config_ui(self):
        # Valores por defecto
        self.ui_titulo = "UNINDETEC"
        self.ui_subtitulo = "CONTROL DE INVENTARIO"
        self.ui_tema = "primary"
        self.ui_logo = ""
        
        if os.path.exists(ARCHIVO_UI_CONFIG):
            try:
                with open(ARCHIVO_UI_CONFIG, 'r') as f:
                    data = json.load(f)
                    self.ui_titulo = data.get("titulo", self.ui_titulo)
                    self.ui_subtitulo = data.get("subtitulo", self.ui_subtitulo)
                    self.ui_tema = data.get("tema", self.ui_tema)
                    self.ui_logo = data.get("logo", "")
            except: pass

    # --- EN LA CLASE SistemaInventario ---
    def setup_header(self):
        # 1. Limpiar encabezado anterior
        for widget in self.root.pack_slaves():
            if isinstance(widget, ttk.Frame) and widget != self.notebook:
                widget.destroy()

        # 2. Frame Principal del Header
        fr = ttk.Frame(self.root, padding=(15, 10))
        try:
            fr.pack(side=TOP, fill=X, before=self.notebook)
        except:
            fr.pack(side=TOP, fill=X)

        # Color de fondo del tema (para componer la transparencia del PNG)
        bg_color = self.tema_actual.get("color_fondo", "#FFFFFF")

        def cargar_imagen_sin_fondo(ruta, max_px):
            """
            Carga un PNG (con o sin alpha) y lo compone sobre el color
            de fondo del tema. Así la transparencia se ve natural.
            Respeta la proporción original (thumbnail).
            """
            img = Image.open(ruta).copy()
            img.thumbnail((max_px, max_px), Image.LANCZOS)

            # Convertir a RGBA para manejar cualquier modo (RGB, P, RGBA…)
            img_rgba = img.convert("RGBA")

            # Crear fondo del color exacto del tema
            fondo = Image.new("RGBA", img_rgba.size, bg_color)

            # Pegar la imagen usando su canal alfa como máscara
            fondo.paste(img_rgba, mask=img_rgba.split()[3])

            return fondo.convert("RGB")

        # ══════════════════════════════════════════════════════════════
        # SECCIÓN IZQUIERDA: LOGO + TÍTULO
        # ══════════════════════════════════════════════════════════════
        fr_left = ttk.Frame(fr)
        fr_left.pack(side=LEFT, fill=Y, anchor=W)

        # ── Logo ──────────────────────────────────────────────────────
        logo_path = self.db.get_config("LOGO_APP")
        if not logo_path or not os.path.exists(logo_path):
            logo_path = "logo.png" if os.path.exists("logo.png") else None

        if logo_path:
            try:
                img_final = cargar_imagen_sin_fondo(logo_path, 150)
                self.tk_logo = ImageTk.PhotoImage(img_final, master=fr)
                ttk.Label(fr_left, image=self.tk_logo).pack(
                    side=LEFT, padx=(0, 15))
            except:
                ttk.Label(fr_left, text="⚓",
                          font=("Arial", 36)).pack(side=LEFT, padx=12)
        else:
            ttk.Label(fr_left, text="⚓",
                      font=("Arial", 36)).pack(side=LEFT, padx=12)

        # ── Textos: Título + Subtítulo ────────────────────────────────
        txt_fr = ttk.Frame(fr_left)
        txt_fr.pack(side=LEFT, fill=Y, anchor=CENTER)

        t_principal = self.db.get_config("TITULO_APP")    or "SOFTWARE DE USO LIBRE"
        t_sub       = self.db.get_config("SUBTITULO_APP") or "CONTROL DE PARTIDAS"

        # Fuente adaptable según longitud
        n = len(t_principal)
        if   n <= 12:   fs_titulo = 28
        elif n <= 20:   fs_titulo = 24
        elif n <= 30:   fs_titulo = 21
        elif n <= 40:   fs_titulo = 18
        elif n <= 52:   fs_titulo = 15
        else:           fs_titulo = 13

        ttk.Label(
            txt_fr,
            text=t_principal,
            font=("Arial Black", fs_titulo),
            bootstyle="primary",
            wraplength=700
        ).pack(anchor=W)

        ttk.Label(
            txt_fr,
            text=t_sub,
            font=("Segoe UI", 13, "bold"),
            bootstyle="secondary"
        ).pack(anchor=W)

        # ══════════════════════════════════════════════════════════════
        # SECCIÓN DERECHA: FICHA DE USUARIO
        # ══════════════════════════════════════════════════════════════
        fr_right = ttk.Frame(fr)
        fr_right.pack(side=RIGHT, fill=Y, anchor=E)

        nombre_user    = self.usuario.get('nombre_completo',
                                          self.usuario['usuario']).upper()
        email_user     = self.usuario.get('email', 'Usuario del Sistema')
        foto_user_path = self.usuario.get('foto_path', '')
        rol_user       = self.usuario.get('rol', 'OPERADOR')

        # Textos del usuario
        fr_textos_user = ttk.Frame(fr_right)
        fr_textos_user.pack(side=LEFT, padx=(0, 15), anchor=E)

        ttk.Label(fr_textos_user,
                  text="¡BIENVENIDO(A)!",
                  font=("Segoe UI", 9, "bold"),
                  bootstyle="success",
                  anchor=E).pack(anchor=E)

        ttk.Label(fr_textos_user,
                  text=nombre_user,
                  font=("Segoe UI", 16, "bold"),
                  bootstyle="primary",
                  anchor=E).pack(anchor=E)

        ttk.Label(fr_textos_user,
                  text=f"{rol_user} | {email_user}",
                  font=("Segoe UI", 9),
                  bootstyle="secondary",
                  anchor=E).pack(anchor=E)

        # Foto de perfil (también sin fondo blanco)
        fr_foto_marco = ttk.Frame(fr_right, padding=2, bootstyle="secondary")
        fr_foto_marco.pack(side=LEFT, padx=(0, 12))

        if foto_user_path and os.path.exists(foto_user_path):
            try:
                img_foto = cargar_imagen_sin_fondo(foto_user_path, 70)
                self.tk_foto_user = ImageTk.PhotoImage(img_foto, master=fr)
                ttk.Label(fr_foto_marco,
                          image=self.tk_foto_user).pack()
            except:
                ttk.Label(fr_foto_marco,
                          text="👤",
                          font=("Segoe UI Emoji", 40)).pack()
        else:
            icono_def = "👨‍✈️" if rol_user == 'ADMIN' else "👤"
            ttk.Label(fr_foto_marco,
                      text=icono_def,
                      font=("Segoe UI Emoji", 40)).pack()

        # Separador + Botón configuración
        if self.usuario.get('rol', '') != 'SOLO LECTURA':
            ttk.Separator(fr_right, orient=VERTICAL).pack(
                side=LEFT, fill=Y, padx=8
            )
            btn_conf = ttk.Button(
                fr_right, text="⚙️",
                bootstyle="link",
                command=self.abrir_menu_admin
            )
            btn_conf.pack(side=LEFT, padx=5)
            try:
                for child in btn_conf.winfo_children():
                    child.configure(font=("Segoe UI Emoji", 28))
            except:
                pass

    def setup_autocomplete(self, combo, lista_completa):
        """
        Autocomplete:
        - Filtra en tiempo real mientras escribes
        - Abre automáticamente al hacer clic con la lista completa
        - NO se reabre después de seleccionar
        """
        self._seleccionando = False

        def filtrar(event=None):
            if event and event.keysym in ('Up', 'Down', 'Return', 'Tab',
                                           'Left', 'Right', 'Escape'):
                return
            texto = combo.get().strip().upper()
            if texto == "":
                combo['values'] = lista_completa
            else:
                filtrados = [x for x in lista_completa
                             if texto in str(x).upper()]
                combo['values'] = filtrados

        def al_seleccionar(event=None):
            """Marca que acabamos de seleccionar para no reabrir"""
            self._seleccionando = True
            combo['values'] = lista_completa
            # Resetear la bandera después de un momento
            combo.after(300, lambda: setattr(self, '_seleccionando', False))

        def abrir_al_click(event=None):
            """Abre el dropdown al hacer clic, solo si no acabamos de seleccionar"""
            if self._seleccionando:
                return
            combo['values'] = lista_completa
            combo.after(100, lambda: combo.event_generate('<Down>'))

        combo.bind('<KeyRelease>',         filtrar)
        combo.bind('<Button-1>',           abrir_al_click)
        combo.bind('<<ComboboxSelected>>', al_seleccionar)

        # Carga inicial
        combo['values'] = lista_completa

    def setup_tab_inventario(self):
        rol_actual = self.usuario.get('rol', 'OPERADOR')
        es_solo_lectura = (rol_actual == 'SOLO LECTURA')

        # ── PANEL IZQUIERDO: solo visible si NO es solo lectura ───────
        if not es_solo_lectura:
            self.p_izq = ttk.Frame(self.tab_inv, width=570)
            self.p_izq.pack(side=LEFT, fill=Y, padx=(0, 10))
            self.p_izq.pack_propagate(False)

            # === Indicador de selección ===
            fr_info = ttk.LabelFrame(
                self.p_izq,
                text=" 📦 Material Seleccionado ",
                padding=10, bootstyle="secondary"
            )
            fr_info.pack(fill=X, pady=(0, 10))

            self.lbl_seleccionado = ttk.Label(
                fr_info,
                text="Ninguno (Selecciona en la tabla ➡)",
                font=("Segoe UI", 11, "bold"),
                foreground="#E67E22", wraplength=380
            )
            self.lbl_seleccionado.pack(anchor=CENTER)

            self.lbl_stock_actual = ttk.Label(
                fr_info, text="Stock: --",
                font=("Segoe UI", 10)
            )
            self.lbl_stock_actual.pack(anchor=CENTER)

            # === Pestañas de acción ===
            self.nb_acciones = ttk.Notebook(self.p_izq, bootstyle="primary")
            self.nb_acciones.pack(fill=BOTH, expand=True)

            # ── PESTAÑA 1: ENTRADAS ───────────────────────────────────
            self.tab_entradas = ttk.Frame(self.nb_acciones, padding=15)
            self.nb_acciones.add(self.tab_entradas, text="⬇️ ENTRADAS")

            ttk.Label(
                self.tab_entradas,
                text="ENTRADA DE NUEVO MATERIAL",
                font=("Segoe UI", 12, "bold"),
                foreground="#27ae60"
            ).pack(pady=(0, 15))

            ttk.Label(self.tab_entradas, text="Cantidad a Ingresar:").pack(anchor=W)
            self.ent_cant_ent = ttk.Entry(
                self.tab_entradas,
                font=("Segoe UI", 14, "bold"),
                justify=CENTER
            )
            self.ent_cant_ent.pack(fill=X, pady=5)

            ttk.Label(
                self.tab_entradas,
                text="Factura / Referencia:"
            ).pack(anchor=W, pady=(10, 0))
            self.ent_factura_ent = ttk.Entry(self.tab_entradas)
            self.ent_factura_ent.pack(fill=X, pady=2)

            estado_ent = "normal" if self.tiene_permiso('entrada') else "disabled"
            btn_ent = ttk.Button(
                self.tab_entradas,
                text="✅ REGISTRAR ENTRADA",
                bootstyle="success", state=estado_ent,
                command=lambda: self.procesar_movimiento("ENTRADA")
            )
            btn_ent.pack(fill=X, pady=30, ipady=5)
            if estado_ent == "disabled":
                ToolTip(btn_ent, text="No tienes permiso para registrar entradas")

            # ── PESTAÑA 2: SALIDAS ────────────────────────────────────
            self.tab_salidas = ttk.Frame(self.nb_acciones, padding=10)
            self.nb_acciones.add(self.tab_salidas, text="⬆️ SALIDAS")

            ttk.Label(
                self.tab_salidas,
                text="Registrar Salida de Materiales",
                font=("Segoe UI", 12, "bold"),
                foreground="#c0392b"
            ).pack(pady=(0, 8))

            fr_datos = ttk.LabelFrame(
                self.tab_salidas,
                text=" Datos del Vale ",
                padding=8, bootstyle="danger"
            )
            fr_datos.pack(fill=X, pady=(0, 8))

            ttk.Label(fr_datos, text="Destino / Área:").grid(row=0, column=0, sticky=W, pady=2)
            self.cb_area_sal = ttk.Combobox(fr_datos)
            self.cb_area_sal.grid(row=0, column=1, sticky=EW, padx=5, pady=2)

            ttk.Label(fr_datos, text="Solicita (Nombre):").grid(row=1, column=0, sticky=W, pady=2)
            self.ent_resp_sal = ttk.Entry(fr_datos)
            self.ent_resp_sal.grid(row=1, column=1, sticky=EW, padx=5, pady=2)

            ttk.Label(fr_datos, text="Autoriza / Entrega:").grid(row=2, column=0, sticky=W, pady=2)
            self.cb_jefe_sal = ttk.Combobox(fr_datos)
            self.cb_jefe_sal.grid(row=2, column=1, sticky=EW, padx=5, pady=2)
            fr_datos.columnconfigure(1, weight=1)

            fr_agregar = ttk.LabelFrame(
                self.tab_salidas,
                text=" ➕ Agregar Material al Vale ",
                padding=8, bootstyle="warning"
            )
            fr_agregar.pack(fill=X, pady=(0, 6))

            ttk.Label(
                fr_agregar,
                text="💡 Selecciona un material en la tabla y escribe la cantidad:",
                font=("Segoe UI", 8), foreground="gray"
            ).pack(anchor=W)

            fr_cant_add = ttk.Frame(fr_agregar)
            fr_cant_add.pack(fill=X, pady=5)

            ttk.Label(
                fr_cant_add,
                text="Cantidad:", font=("Segoe UI", 10, "bold")
            ).pack(side=LEFT)
            self.ent_cant_sal = ttk.Entry(
                fr_cant_add, width=8,
                font=("Segoe UI", 13, "bold"), justify=CENTER
            )
            self.ent_cant_sal.pack(side=LEFT, padx=8)

            estado_sal = "normal" if self.tiene_permiso('salida') else "disabled"
            ttk.Button(
                fr_cant_add,
                text="➕ Agregar",
                bootstyle="warning", state=estado_sal,
                command=self.agregar_al_carrito
            ).pack(side=LEFT)

            fr_carrito = ttk.LabelFrame(
                self.tab_salidas,
                text=" 🛒 Materiales en este Vale ",
                padding=5, bootstyle="info"
            )
            fr_carrito.pack(fill=BOTH, expand=True, pady=(0, 6))

            cols_c = ("MATERIAL", "CANT", "STOCK")
            self.tree_carrito = ttk.Treeview(
                fr_carrito, columns=cols_c,
                show="headings", height=5, bootstyle="warning"
            )
            self.tree_carrito.heading("MATERIAL", text="Material")
            self.tree_carrito.column("MATERIAL", width=220)
            self.tree_carrito.heading("CANT", text="Cantidad")
            self.tree_carrito.column("CANT", width=65, anchor=CENTER)
            self.tree_carrito.heading("STOCK", text="Stock Actual")
            self.tree_carrito.column("STOCK", width=80, anchor=CENTER)

            sc_c = ttk.Scrollbar(
                fr_carrito, orient=VERTICAL,
                command=self.tree_carrito.yview
            )
            self.tree_carrito.configure(yscrollcommand=sc_c.set)
            self.tree_carrito.pack(side=LEFT, fill=BOTH, expand=True)
            sc_c.pack(side=RIGHT, fill=Y)

            ttk.Button(
                self.tab_salidas,
                text="🗑️ Quitar seleccionado del carrito",
                bootstyle="secondary-outline",
                command=self.quitar_del_carrito
            ).pack(fill=X, pady=(0, 4))

            btn_sal = ttk.Button(
                self.tab_salidas,
                text="🔥 REGISTRAR VALE DE SALIDA",
                bootstyle="danger", state=estado_sal,
                command=self.procesar_salida_multiple
            )
            btn_sal.pack(fill=X, ipady=6)
            if estado_sal == "disabled":
                ToolTip(btn_sal, text="No tienes permiso para registrar salidas")

            self._carrito = []

            # ── PESTAÑA 3: CREAR NUEVO MATERIAL ──────────────────────
            self.tab_nuevo = ttk.Frame(self.nb_acciones, padding=15)
            self.nb_acciones.add(self.tab_nuevo, text="➕ CREAR NUEVO MATERIAL")

            ttk.Label(
                self.tab_nuevo,
                text="Crear Producto",
                font=("Segoe UI", 12, "bold"),
                foreground="#2980b9"
            ).pack(pady=(0, 15))

            ttk.Label(self.tab_nuevo, text="Partida:").pack(anchor=W)
            self.cb_partida = ttk.Combobox(self.tab_nuevo, state="readonly")
            self.cb_partida.pack(fill=X, pady=2)

            ttk.Label(
                self.tab_nuevo,
                text="Descripción:"
            ).pack(anchor=W, pady=(10, 0))
            self.txt_desc = tk.Text(
                self.tab_nuevo, height=4,
                font=("Segoe UI", 10), wrap="word"
            )
            self.txt_desc.pack(fill=X, pady=2, padx=1)

            # ── Tab en el Text salta al siguiente campo ───────────────
            def tab_en_descripcion(event):
                self.ent_stock_inicial.focus_set()
                return "break"   # "break" evita que inserte el tabulador

            self.txt_desc.bind("<Tab>", tab_en_descripcion)

            ttk.Label(
                self.tab_nuevo,
                text="Cantidad Inicial (Stock):"
            ).pack(anchor=W, pady=(10, 0))
            self.ent_stock_inicial = ttk.Entry(
                self.tab_nuevo, justify=CENTER,
                font=("Segoe UI", 10, "bold")
            )
            # SIN valor por defecto — queda en blanco
            self.ent_stock_inicial.pack(fill=X, pady=2)

            ttk.Label(
                self.tab_nuevo,
                text="Factura Inicial:"
            ).pack(anchor=W, pady=(10, 0))
            self.txt_factura_alta = ttk.Entry(self.tab_nuevo)
            self.txt_factura_alta.pack(fill=X, pady=2)

            estado_crear = "normal" if self.tiene_permiso('crear') else "disabled"
            btn_crear = ttk.Button(
                self.tab_nuevo,
                text="💾 GUARDAR",
                bootstyle="info", state=estado_crear,
                command=self.agregar_material
            )
            btn_crear.pack(fill=X, pady=30, ipady=5)
            if estado_crear == "disabled":
                ToolTip(btn_crear, text="No tienes permiso para crear materiales")

        else:
            # SOLO LECTURA: cartel informativo en lugar del panel de acciones
            self._carrito = []
            fr_aviso = ttk.LabelFrame(
                self.tab_inv,
                text=" 🔒 Modo Solo Lectura ",
                padding=20, bootstyle="warning"
            )
            fr_aviso.pack(side=LEFT, fill=Y, padx=(0, 10))

            ttk.Label(
                fr_aviso,
                text="👁️",
                font=("Segoe UI Emoji", 42)
            ).pack(pady=(20, 10))

            ttk.Label(
                fr_aviso,
                text="SOLO LECTURA",
                font=("Segoe UI", 13, "bold"),
                foreground="#E67E22"
            ).pack()

            ttk.Label(
                fr_aviso,
                text="Solo puedes consultar\nel inventario actual.",
                font=("Segoe UI", 10),
                foreground="gray",
                justify=CENTER
            ).pack(pady=(8, 0))

            # Etiquetas de selección (necesarias para on_tree_select)
            self.lbl_seleccionado = ttk.Label(
                fr_aviso, text="",
                font=("Segoe UI", 9),
                foreground="#E67E22",
                wraplength=160
            )
            self.lbl_seleccionado.pack(pady=(20, 0))

            self.lbl_stock_actual = ttk.Label(
                fr_aviso, text="",
                font=("Segoe UI", 10, "bold"),
                foreground="#27ae60"
            )
            self.lbl_stock_actual.pack(pady=(4, 0))

        # ════════════════════════════════════════════════════════════
        # PANEL DERECHO: TABLA DE INVENTARIO (igual para todos)
        # ════════════════════════════════════════════════════════════
        p_der = ttk.Frame(self.tab_inv)
        p_der.pack(side=RIGHT, fill=BOTH, expand=True)

        fr_top = ttk.Frame(p_der, padding=(0, 5))
        fr_top.pack(fill=X)

        ttk.Label(
            fr_top, text="🔍 Buscar:",
            font=("Segoe UI", 9, "bold")
        ).pack(side=LEFT)

        self.cb_busqueda_material = ttk.Entry(fr_top, width=30)
        self.cb_busqueda_material.pack(side=LEFT, padx=5)

        def buscar_inventario_realtime(event=None):
            if event and event.keysym in ('Tab', 'Escape', 'Return'):
                return
            self.cargar_tabla_inventario()

        # Un solo binding limpio — Entry no tiene dropdown, no necesita más
        self.cb_busqueda_material.bind('<KeyRelease>', buscar_inventario_realtime)
      

        # Solo filtra mientras escribe — NO abre dropdown al hacer clic
        
        

       
       

        ttk.Label(
            fr_top, text="📂 Filtrar Partida:",
            font=("Segoe UI", 9, "bold")
        ).pack(side=LEFT, padx=(15, 5))

        self.cb_filtro_partida = ttk.Combobox(
            fr_top, state="readonly", width=15
        )
        self.cb_filtro_partida.pack(side=LEFT, padx=5)
        self.cb_filtro_partida.bind(
            "<<ComboboxSelected>>", self.cargar_tabla_inventario
        )

        ttk.Button(
            fr_top, text="🔄 Actualizar",
            bootstyle="link",
            command=self.limpiar_filtros
        ).pack(side=LEFT)

        # Tabla principal de inventario
        cols = ("ID", "PARTIDA", "MATERIAL", "STOCK")
        self.tree_inv = ttk.Treeview(
            p_der, columns=cols,
            show="headings", bootstyle="info"
        )
        self.tree_inv.heading("ID",       text="ID")
        self.tree_inv.column( "ID",       width=40,  anchor=CENTER)
        self.tree_inv.heading("PARTIDA",  text="PARTIDA")
        self.tree_inv.column( "PARTIDA",  width=80,  anchor=CENTER)
        self.tree_inv.heading("MATERIAL", text="DESCRIPCIÓN")
        self.tree_inv.column( "MATERIAL", width=400)
        self.tree_inv.heading("STOCK",    text="STOCK")
        self.tree_inv.column( "STOCK",    width=80,  anchor=CENTER)

        sc = ttk.Scrollbar(p_der, orient=VERTICAL, command=self.tree_inv.yview)
        self.tree_inv.configure(yscrollcommand=sc.set)
        self.tree_inv.pack(side=LEFT, fill=BOTH, expand=True)
        sc.pack(side=RIGHT, fill=Y)

        self.tree_inv.tag_configure(
            "BAJO", background="#ffcccc", foreground="#8a1f1f"
        )
        self.tree_inv.bind("<<TreeviewSelect>>", self.on_tree_select)

        # Menú contextual: SOLO si no es modo solo lectura
        if not es_solo_lectura:
            self.menu_inv = tk.Menu(self.root, tearoff=0)
            if self.tiene_permiso('editar'):
                self.menu_inv.add_command(
                    label="✏️ Corregir/Editar Material",
                    command=self.editar_material_seleccionado
                )
            if self.tiene_permiso('eliminar') or self.usuario.get('rol') == 'ADMIN':
                self.menu_inv.add_separator()
                self.menu_inv.add_command(
                    label="🗑️ Eliminar Material (Solo Admin)",
                    command=self.eliminar_material_seleccionado
                )

            def mostrar_menu_inv(event):
                item = self.tree_inv.identify_row(event.y)
                if item:
                    self.tree_inv.selection_set(item)
                    estado_borrar = (
                        "normal"
                        if self.usuario.get('rol') == 'ADMIN'
                        else "disabled"
                    )
                    try:
                        self.menu_inv.entryconfig(
                            "🗑️ Eliminar Material (Solo Admin)",
                            state=estado_borrar
                        )
                    except:
                        pass
                    self.menu_inv.post(event.x_root, event.y_root)

            if (self.tiene_permiso('editar') or
                    self.tiene_permiso('eliminar') or
                    self.usuario.get('rol') == 'ADMIN'):
                self.tree_inv.bind("<Button-3>", mostrar_menu_inv)


    # ══════════════════════════════════════════════════════════════════
    # NUEVOS MÉTODOS DEL CARRITO — AGRÉGALOS EN LA CLASE
    # ══════════════════════════════════════════════════════════════════

    def agregar_al_carrito(self):
        """Agrega el material seleccionado en la tabla al carrito."""
        sel = self.tree_inv.selection()
        if not sel:
            messagebox.showwarning(
                "Atención", "Selecciona un material de la tabla derecha.")
            return

        item   = self.tree_inv.item(sel[0])
        vals   = item['values']
        id_mat, partida, nombre_mat, stock_actual = (
            vals[0], vals[1], vals[2], float(vals[3]))

        # Validar cantidad
        try:
            cantidad_raw = float(self.ent_cant_sal.get())
            if cantidad_raw <= 0:
                raise ValueError
            cantidad = int(cantidad_raw) if cantidad_raw == int(cantidad_raw) else cantidad_raw
        except:
            messagebox.showerror("Error", "Ingresa una cantidad válida.")
            return

        # Verificar que no exceda el stock
        # (descontando lo que ya hay en el carrito para este material)
        ya_en_carrito = sum(
            x['cantidad'] for x in self._carrito
            if x['id'] == id_mat)

        if cantidad + ya_en_carrito > stock_actual:
            messagebox.showerror(
                "Stock insuficiente",
                f"Stock disponible: {stock_actual - ya_en_carrito}")
            return

        # Verificar si ya está en el carrito → sumar cantidad
        for item_c in self._carrito:
            if item_c['id'] == id_mat:
                item_c['cantidad'] += cantidad
                self._refrescar_carrito()
                self.ent_cant_sal.delete(0, END)
                return

        # Nuevo item en el carrito
        self._carrito.append({
            'id':       id_mat,
            'partida':  partida,
            'material': nombre_mat,
            'cantidad': cantidad,
            'stock':    stock_actual
        })
        self._refrescar_carrito()
        self.ent_cant_sal.delete(0, END)

    def quitar_del_carrito(self):
        """Elimina el renglón seleccionado del carrito."""
        sel = self.tree_carrito.selection()
        if not sel:
            messagebox.showwarning("Atención", "Selecciona un item del carrito.")
            return
        idx = self.tree_carrito.index(sel[0])
        self._carrito.pop(idx)
        self._refrescar_carrito()

    def _refrescar_carrito(self):
        """Actualiza la tabla visual del carrito."""
        for i in self.tree_carrito.get_children():
            self.tree_carrito.delete(i)
        for item in self._carrito:
            cant_fmt  = int(item['cantidad']) if item['cantidad'] == int(item['cantidad']) else item['cantidad']
            stock_fmt = int(item['stock'])    if item['stock']    == int(item['stock'])    else item['stock']
            self.tree_carrito.insert(
                "", END,
                values=(item['material'], cant_fmt, stock_fmt)
            )

    def procesar_salida_multiple(self):
        """
        Procesa TODOS los materiales del carrito en UNA sola transacción
        y genera UN solo vale PDF con todos los renglones.
        """
        if not self._carrito:
            messagebox.showwarning(
                "Carrito vacío",
                "Agrega al menos un material al carrito antes de registrar."
            )
            return

        destino     = self.cb_area_sal.get().strip().upper()  or "S/N"
        responsable = self.ent_resp_sal.get().strip().upper() or "S/N"
        entrego     = self.cb_jefe_sal.get().strip().upper()  or "S/N"

        if destino == "S/N" or responsable == "S/N":
            messagebox.showwarning(
                "Faltan datos",
                "Completa el Área y el nombre de quien solicita."
            )
            return

        resumen = "\n".join(
            f"  • {x['material'][:40]}  →  {x['cantidad']} pzas"
            for x in self._carrito
        )
        if not messagebox.askyesno(
                "Confirmar Vale de Salida",
                f"Se registrarán {len(self._carrito)} material(es):\n\n"
                f"{resumen}\n\n"
                f"Área: {destino}\nSolicita: {responsable}"):
            return

        fecha      = datetime.now().strftime("%d/%m/%Y")
        fecha_full = datetime.now().strftime("%d/%m/%Y %H:%M")
        folio      = self.generar_folio()
        usuario_act = self.usuario.get('usuario', 'SISTEMA')

        # ── Transacción atómica ────────────────────────────────────────
        try:
            with self.db.conectar() as conn:
                conn.execute("BEGIN")
                try:
                    for item in self._carrito:
                        nuevo_stock = item['stock'] - item['cantidad']

                        # Actualizar stock
                        conn.execute(
                            "UPDATE inventario SET stock=?, ultimo_movimiento=? WHERE id=?",
                            (nuevo_stock, fecha, item['id'])
                        )

                        # Historial individual por material con usuario
                        conn.execute(
                            "INSERT INTO historial "
                            "(fecha_hora, tipo, partida, material, cantidad, "
                            "destino, responsable, entrego, factura, usuario_sistema) "
                            "VALUES (?,?,?,?,?,?,?,?,?,?)",
                            (fecha_full, "SALIDA",
                             item['partida'], item['material'],
                             item['cantidad'],
                             destino, responsable, entrego, folio,
                             usuario_act)
                        )

                    conn.commit()

                except Exception as e:
                    conn.rollback()
                    messagebox.showerror(
                        "Error crítico",
                        f"Vale cancelado. BD no modificada.\n\n{e}"
                    )
                    return

            # ── Generar PDF ────────────────────────────────────────────
            self.generar_pdf_vale_multiple(
                self._carrito, destino, responsable, entrego, folio
            )

            messagebox.showinfo(
                "✅ Éxito",
                f"Vale {folio} registrado con {len(self._carrito)} material(es)."
            )

            # Limpiar carrito y campos
            self._carrito = []
            self._refrescar_carrito()
            self.ent_resp_sal.delete(0, END)

            self.cargar_tabla_inventario()
            self.cargar_tabla_historial()

        except Exception as e:
            messagebox.showerror("Error de conexión", f"{e}")



    def on_tree_select(self, event):
        """Actualiza el letrero naranja de la izquierda al seleccionar un producto"""
        sel = self.tree_inv.selection()
        if sel:
            item = self.tree_inv.item(sel[0])
            vals = item['values']
            # vals[2] es Nombre, vals[3] es Stock
            self.lbl_seleccionado.config(text=vals[2]) 
            self.lbl_stock_actual.config(text=f"Stock Actual: {vals[3]}")
            
            # Efecto visual: Rojo si es poco stock
            try:
                stock = float(vals[3])
                if stock <= 2:
                    self.lbl_seleccionado.config(foreground="red")
                else:
                    self.lbl_seleccionado.config(foreground="#E67E22")
            except: pass
        else:
            self.lbl_seleccionado.config(text="Ninguno (Selecciona en la tabla ➡)", foreground="gray")
            self.lbl_stock_actual.config(text="Stock: --")

        def mostrar_menu_inv(event):
            item = self.tree_inv.identify_row(event.y)
            if item:
                self.tree_inv.selection_set(item)
                estado_borrar = "normal" if self.usuario.get('rol') == 'ADMIN' else "disabled"
                self.menu_inv.entryconfig("🗑️ Eliminar Material (Solo Admin)", state=estado_borrar)
                self.menu_inv.post(event.x_root, event.y_root)
        self.tree_inv.bind("<Button-3>", mostrar_menu_inv)


    # ------------------------------------------------------------------
    #  LÓGICA NUEVA PARA PROCESAR MOVIMIENTOS SEPARADOS
    # ------------------------------------------------------------------
    def procesar_movimiento(self, tipo):
        sel = self.tree_inv.selection()
        if not sel:
            messagebox.showwarning("Atención", "Selecciona un material de la tabla derecha.")
            return

        item   = self.tree_inv.item(sel[0])
        valores = item['values']
        id_mat, partida, nombre_mat, stock_actual = (
            valores[0], valores[1], valores[2], float(valores[3])
        )

        cantidad = 0; factura = "S/F"; destino = "S/N"; responsable = "S/N"; entrego = "S/N"

        try:
            if tipo == "SALIDA":
                cantidad    = float(self.ent_cant_sal.get())
                destino     = self.cb_area_sal.get().strip().upper()  or "S/N"
                responsable = self.ent_resp_sal.get().strip().upper() or "S/N"
                entrego     = self.cb_jefe_sal.get().strip().upper()  or "S/N"
                if cantidad > stock_actual:
                    messagebox.showerror("Error", f"Stock insuficiente ({stock_actual}).")
                    return
            elif tipo == "ENTRADA":
                cantidad = float(self.ent_cant_ent.get())
                factura  = self.ent_factura_ent.get().strip().upper() or "S/F"

            if cantidad <= 0:
                raise ValueError
        except:
            messagebox.showerror("Error", "Cantidad inválida.")
            return

        nuevo_stock = stock_actual + cantidad if tipo == "ENTRADA" else stock_actual - cantidad
        fecha       = datetime.now().strftime("%d/%m/%Y")
        fecha_full  = datetime.now().strftime("%d/%m/%Y %H:%M")
        usuario_act = self.usuario.get('usuario', 'SISTEMA')

        # ── Transacción atómica ────────────────────────────────────────
        try:
            with self.db.conectar() as conn:
                conn.execute("BEGIN")
                try:
                    # 1. Actualizar stock
                    if tipo == "ENTRADA" and factura != "S/F":
                        conn.execute(
                            "UPDATE inventario SET stock=?, ultimo_movimiento=?, factura=? WHERE id=?",
                            (nuevo_stock, fecha, factura, id_mat)
                        )
                    else:
                        conn.execute(
                            "UPDATE inventario SET stock=?, ultimo_movimiento=? WHERE id=?",
                            (nuevo_stock, fecha, id_mat)
                        )

                    # 2. Guardar historial con usuario
                    conn.execute(
                        "INSERT INTO historial "
                        "(fecha_hora, tipo, partida, material, cantidad, "
                        "destino, responsable, entrego, factura, usuario_sistema) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (fecha_full, tipo, partida, nombre_mat, cantidad,
                         destino, responsable, entrego, factura, usuario_act)
                    )
                    conn.commit()

                except Exception as e:
                    conn.rollback()
                    messagebox.showerror(
                        "Error crítico",
                        f"Movimiento cancelado. La BD no fue modificada.\n\nDetalle: {e}"
                    )
                    return

            # ── PDF de salida ──────────────────────────────────────────
            if tipo == "SALIDA":
                folio = self.generar_folio()
                self.generar_pdf_vale(nombre_mat, cantidad, destino, responsable, entrego, folio)

            messagebox.showinfo("Éxito", f"Movimiento registrado. Nuevo stock: {nuevo_stock}")

            if tipo == "SALIDA":
                self.ent_cant_sal.delete(0, END)
            else:
                self.ent_cant_ent.delete(0, END)

            self.cargar_tabla_inventario()
            self.cargar_tabla_historial()

            try:
                self.tree_inv.selection_set(sel[0])
                self.on_tree_select(None)
            except:
                pass

        except Exception as e:
            messagebox.showerror(
                "Error de conexión",
                f"No se pudo conectar a la base de datos.\n\nDetalle: {e}"
            )

    def cargar_tabla_inventario(self, event=None):
        for i in self.tree_inv.get_children():
            self.tree_inv.delete(i)

        partida_sel    = self.cb_filtro_partida.get()
        texto_busqueda = self.cb_busqueda_material.get().strip().upper()

        sql    = "SELECT * FROM inventario WHERE 1=1"
        params = []

        if partida_sel and partida_sel != "TODAS":
            sql += " AND partida = ?"
            params.append(partida_sel)

        if texto_busqueda:
            sql += " AND material LIKE ?"
            params.append(f"%{texto_busqueda}%")

        sql += " ORDER BY id DESC"

        filas = self.db.consultar(sql, tuple(params))
        for f in filas:
            tag = "BAJO" if f['stock'] <= 2 else ""
            # Siempre entero si no tiene decimales reales
            stock_val = f['stock']
            try:
                stock_fmt = int(stock_val) if float(stock_val) == int(float(stock_val)) else round(float(stock_val), 2)
            except:
                stock_fmt = stock_val
            self.tree_inv.insert(
                "", END,
                values=(f['id'], f['partida'], f['material'], stock_fmt),
                tags=(tag,)
            )

    def agregar_material(self):
        partida = self.cb_partida.get()
        mat     = self.txt_desc.get("1.0", "end-1c").strip().upper()
        fact    = self.txt_factura_alta.get().strip().upper() or "S/F"

        try:
            stock_ini = float(self.ent_stock_inicial.get())
            if stock_ini < 0:
                raise ValueError
        except:
            stock_ini = 0.0

        if not partida or not mat:
            messagebox.showwarning("Faltan datos", "Indica Partida y Descripción")
            return

        try:
            # 1. Insertar en INVENTARIO con stock inicial
            self.db.ejecutar(
                "INSERT INTO inventario (partida, material, factura, stock, ultimo_movimiento) "
                "VALUES (?, ?, ?, ?, 'ALTA')",
                (partida, mat, fact, stock_ini)
            )

            # 2. Si hay stock inicial, registrar en HISTORIAL
            if stock_ini > 0:
                fecha_full  = datetime.now().strftime("%d/%m/%Y %H:%M")
                usuario_act = self.usuario.get('usuario', 'SISTEMA')

                self.db.ejecutar(
                    "INSERT INTO historial "
                    "(fecha_hora, tipo, partida, material, cantidad, "
                    "destino, responsable, entrego, factura, usuario_sistema) "
                    "VALUES (?, 'ALTA INICIAL', ?, ?, ?, 'ALMACEN', ?, 'SISTEMA', ?, ?)",
                    (fecha_full, partida, mat, stock_ini,
                     usuario_act, fact, usuario_act)
                )

            messagebox.showinfo("Éxito", f"Material creado correctamente.\nStock inicial: {stock_ini}")

            # Limpiar campos
            self.txt_desc.delete("1.0", END)
            self.txt_factura_alta.delete(0, END)
            self.ent_stock_inicial.delete(0, END)
            self.cb_partida.set("")
            self.cb_partida.focus_set()

            self.cargar_tabla_inventario()
            self.actualizar_combos()

        except sqlite3.IntegrityError:
            messagebox.showerror("Duplicado", "Este material ya existe en esa partida")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")


    def setup_tab_historial(self):
        cols = ("ID", "FECHA", "TIPO", "PARTIDA", "MATERIAL", "CANT", "DESTINO", "RESP", "ENTREGO", "USUARIO")
        self.tree_hist = ttk.Treeview(
            self.tab_hist, columns=cols, show="headings", bootstyle="primary"
        )

        # Columnas cortas/numéricas → CENTER | Columnas de texto → W (izquierda)
        config_cols = {
            "ID":       (45,   CENTER),
            "FECHA":    (130,  CENTER),
            "TIPO":     (90,   CENTER),
            "PARTIDA":  (65,   CENTER),
            "MATERIAL": (260,  W),
            "CANT":     (55,   CENTER),
            "DESTINO":  (110,  W),
            "RESP":     (120,  W),
            "ENTREGO":  (130,  W),
            "USUARIO":  (110,  CENTER),
        }

        for col, (ancho, alineacion) in config_cols.items():
            self.tree_hist.heading(col, text=col, anchor=alineacion)
            self.tree_hist.column(col, width=ancho, anchor=alineacion)

        sc = ttk.Scrollbar(self.tab_hist, orient=VERTICAL, command=self.tree_hist.yview)
        self.tree_hist.configure(yscrollcommand=sc.set)
        self.tree_hist.pack(side=LEFT, fill=BOTH, expand=True)
        sc.pack(side=RIGHT, fill=Y)

        self.tree_hist.tag_configure("ENTRADA",   foreground="green")
        self.tree_hist.tag_configure("SALIDA",    foreground="blue")
        self.tree_hist.tag_configure("SISTEMA",   foreground="gray")
        self.tree_hist.tag_configure("ELIMINADO", foreground="red")

        # ── Menú contextual (solo ADMIN) ──────────────────────────────
        self.menu_hist = tk.Menu(self.root, tearoff=0)
        self.menu_hist.add_command(
            label="⚠️ Revertir y Eliminar Registro (Admin)",
            command=self.revertir_historial_admin
        )

        def mostrar_menu_hist(event):
            if self.usuario.get('rol') == 'ADMIN':
                item = self.tree_hist.identify_row(event.y)
                if item:
                    self.tree_hist.selection_set(item)
                    self.menu_hist.post(event.x_root, event.y_root)

        self.tree_hist.bind("<Button-3>", mostrar_menu_hist)

        # ── Auto-refresco cada 15 segundos ────────────────────────────
        self.root.after(15000, self._autorefresh_historial)

    def cargar_tabla_historial(self):
        try:
            for i in self.tree_hist.get_children():
                self.tree_hist.delete(i)

            filas = self.db.consultar(
                "SELECT id, fecha_hora, tipo, partida, material, cantidad, "
                "destino, responsable, entrego, usuario_sistema "
                "FROM historial ORDER BY id DESC LIMIT 500"
            )
            for f in filas:
                tipo_f = f['tipo'] or ""
                color  = ("ENTRADA" if "ENTRADA" in tipo_f
                          else ("SALIDA" if "SALIDA" in tipo_f
                                else "SISTEMA"))
                usuario_mov = f['usuario_sistema'] or "SISTEMA"

                # Cantidad siempre entera si no tiene decimales reales
                cant_val = f['cantidad']
                try:
                    cant_fmt = int(cant_val) if float(cant_val) == int(float(cant_val)) else round(float(cant_val), 2)
                except:
                    cant_fmt = cant_val

                self.tree_hist.insert(
                    "", END,
                    values=(
                        f['id'], f['fecha_hora'], f['tipo'],
                        f['partida'], f['material'], cant_fmt,
                        f['destino'], f['responsable'], f['entrego'],
                        usuario_mov
                    ),
                    tags=(color,)
                )
        except Exception as e:
            print(f"Error cargando historial: {e}")

    def _autorefresh_historial(self):
        """Refresca el historial automáticamente cada 15 segundos"""
        try:
            if hasattr(self, 'tree_hist') and self.tree_hist.winfo_exists():
                self.cargar_tabla_historial()
                self.root.after(15000, self._autorefresh_historial)
        except Exception as e:
            print(f"Auto-refresh historial detenido: {e}")

    # --- PESTAÑA AUDITORIA (KARDEX) ---
    def setup_tab_auditoria(self):
        # 1. BARRA DE FILTROS SUPERIOR
        fr_top = ttk.Frame(self.tab_audit, padding=10)
        fr_top.pack(fill=X)

        ttk.Label(fr_top, text="Mes:").pack(side=LEFT)
        self.cb_mes_k = ttk.Combobox(
            fr_top,
            values=[str(i) for i in range(1, 13)],
            width=3, state="readonly")
        self.cb_mes_k.current(datetime.now().month - 1)
        self.cb_mes_k.pack(side=LEFT, padx=5)

        ttk.Label(fr_top, text="Año:").pack(side=LEFT)
        self.ent_anio_k = ttk.Entry(fr_top, width=6)
        self.ent_anio_k.insert(0, str(datetime.now().year))
        self.ent_anio_k.pack(side=LEFT, padx=5)

        ttk.Label(fr_top, text="Filtrar Partida:").pack(side=LEFT, padx=(15, 5))
        self.cb_partida_k = ttk.Combobox(fr_top, state="readonly", width=15)
        self.cb_partida_k.pack(side=LEFT)

        # Sin event_generate — la flecha nativa abre el dropdown sin bugs
        self.cb_partida_k.bind('<<ComboboxSelected>>',
            lambda e: self.cb_partida_k.selection_clear())

        ttk.Button(
            fr_top,
            text="🔍 Generar Vista Previa",
            bootstyle="primary",
            command=self.generar_vista_anexo_c
        ).pack(side=LEFT, padx=15)

        ttk.Button(
            fr_top,
            text="💾 Exportar Excel (Anexo C)",
            bootstyle="success",
            command=self.exportar_excel_anexo_c
        ).pack(side=LEFT)

        # ── NUEVO BOTÓN: Reporte de movimientos por fecha ─────────────
        ttk.Separator(fr_top, orient=VERTICAL).pack(
            side=LEFT, fill=Y, padx=15)

        ttk.Button(
            fr_top,
            text="📋 Reporte por Rango de Fechas",
            bootstyle="warning",
            command=self.abrir_reporte_movimientos
        ).pack(side=LEFT)

        # 2. TABLA TIPO EXCEL (Treeview complejo con scroll doble)
        fr_tabla = ttk.Frame(self.tab_audit)
        fr_tabla.pack(fill=BOTH, expand=True, pady=5)

        sc_y = ttk.Scrollbar(fr_tabla, orient=VERTICAL)
        sc_x = ttk.Scrollbar(fr_tabla, orient=HORIZONTAL)

        dias = [str(d) for d in range(1, 32)]
        cols = (["NP", "UNIDAD", "DESC", "FACTURA", "EX_ANT", "RECIBIDOS"]
                + dias
                + ["TOTAL_SAL", "EX_ACT"])

        self.tree_kardex = ttk.Treeview(
            fr_tabla, columns=cols, show="headings",
            yscrollcommand=sc_y.set,
            xscrollcommand=sc_x.set,
            selectmode="browse")

        sc_y.config(command=self.tree_kardex.yview)
        sc_y.pack(side=RIGHT, fill=Y)
        sc_x.config(command=self.tree_kardex.xview)
        sc_x.pack(side=BOTTOM, fill=X)
        self.tree_kardex.pack(side=LEFT, fill=BOTH, expand=True)

        # Encabezados
        self.tree_kardex.heading("NP",        text="N.P.")
        self.tree_kardex.column( "NP",        width=35,  stretch=NO)
        self.tree_kardex.heading("UNIDAD",    text="UNIDAD")
        self.tree_kardex.column( "UNIDAD",    width=40,  stretch=NO)
        self.tree_kardex.heading("DESC",      text="DESCRIPCIÓN")
        self.tree_kardex.column( "DESC",      width=200, minwidth=150)
        self.tree_kardex.heading("FACTURA",   text="FACTURA")
        self.tree_kardex.column( "FACTURA",   width=80)
        self.tree_kardex.heading("EX_ANT",    text="EXISTENCIA ANTERIOR")
        self.tree_kardex.column( "EX_ANT",    width=50,  anchor=CENTER)
        self.tree_kardex.heading("RECIBIDOS", text="ENTRADA")
        self.tree_kardex.column( "RECIBIDOS", width=50,  anchor=CENTER)

        for d in dias:
            self.tree_kardex.heading(d, text=d)
            self.tree_kardex.column( d, width=25, stretch=NO, anchor=CENTER)

        self.tree_kardex.heading("TOTAL_SAL", text="TOTAL SALIDA")
        self.tree_kardex.column( "TOTAL_SAL", width=50, anchor=CENTER)
        self.tree_kardex.heading("EX_ACT",    text="ACT.")
        self.tree_kardex.column( "EX_ACT",    width=50, anchor=CENTER)

    def generar_kardex(self):
        mat = self.cb_kardex_mat.get()
        if not mat: return
        
        # Limpiar tabla visual
        for i in self.tree_kardex.get_children(): self.tree_kardex.delete(i)
        self.datos_kardex_procesados = [] # Limpiar datos para Excel
        
        # Consultar DB
        movs = self.db.consultar("SELECT * FROM historial WHERE material = ? ORDER BY id ASC", (mat,))
        
        saldo = 0
        for m in movs:
            tipo = m['tipo']
            cant = m['cantidad']
            
            # Calcular saldo
            if "ENTRADA" in tipo or "HISTORICO (+)" in tipo:
                saldo += cant
            elif "SALIDA" in tipo or "HISTORICO (-)" in tipo:
                saldo -= cant
                
            # 1. Insertar en la tabla VISUAL
            self.tree_kardex.insert("", END, values=(m['fecha_hora'], tipo, cant, saldo))
            
            # 2. Guardar datos PROCESADOS para Excel (Diccionario limpio)
            self.datos_kardex_procesados.append({
                "FECHA": m['fecha_hora'],
                "MOVIMIENTO": tipo,
                "CANTIDAD": cant,
                "SALDO": saldo,
                "DESTINO": m['destino'],
                "RESPONSABLE": m['responsable']
            })

    def exportar_excel_kardex(self):
        # Verifica si hay datos procesados (calculados en generar_kardex)
        if not hasattr(self, 'datos_kardex_procesados') or not self.datos_kardex_procesados:
            messagebox.showwarning("Alerta", "Primero genera el Kardex en pantalla.")
            return
        
        ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if ruta:
            try:
                # Usamos la lista procesada que TIENE EL SALDO
                df = pd.DataFrame(self.datos_kardex_procesados)
                
                # Reordenar columnas para que se vea bien
                cols_ordenadas = ["FECHA", "MOVIMIENTO", "CANTIDAD", "SALDO", "DESTINO", "RESPONSABLE"]
                # Asegurarnos de que solo usamos columnas que existen (por si acaso)
                cols_final = [c for c in cols_ordenadas if c in df.columns]
                df = df[cols_final]
                
                df.to_excel(ruta, index=False)
                messagebox.showinfo("Exportado", "Archivo Excel generado correctamente.")
                os.startfile(ruta) # Abrir automáticamente
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo exportar: {e}")
    # --- MENÚ ADMIN ---
    # --- EN LA CLASE SistemaInventario ---
    def abrir_menu_admin(self):
        top = tk.Toplevel(self.root)
        top.title("Administración del Sistema")
        self.centrar_ventana_emergente(top, 420, 630)

        fr = ttk.Frame(top, padding=20)
        fr.pack(fill=BOTH, expand=True)

    # Encabezado del menú con info de rol
        rol_actual = self.usuario.get('rol', 'OPERADOR')
        es_admin_rol = (rol_actual == 'ADMIN')

        ttk.Label(fr, text="Menú de Configuración",
                font=("Segoe UI", 14, "bold"), justify=CENTER).pack(pady=(0, 5))

        if not es_admin_rol:
            ttk.Label(fr, text="Modo Restringido: Según tus permisos.",
                    bootstyle="warning", font=("Segoe UI", 9)).pack(pady=(0, 15))
        else:
            ttk.Label(fr, text="Modo Administrador: Acceso total.",
                     bootstyle="success", font=("Segoe UI", 9)).pack(pady=(0, 15))

    # --- BOTONES ---

    # 1. Gestión de Usuarios (SOLO ADMIN ROL)
        estado_users = "normal" if es_admin_rol else "disabled"
        ttk.Button(fr, text="👥  Gestionar Usuarios y Permisos", bootstyle="primary",
                state=estado_users,
                command=self.abrir_gestion_usuarios).pack(fill=X, pady=8, ipady=8)

        ttk.Separator(fr).pack(fill=X, pady=10)

    # 2. Temas
        ttk.Button(fr, text="🎨  Personalizar Temas y Colores", bootstyle="info",
                command=self.abrir_editor_temas).pack(fill=X, pady=8, ipady=8)

    # 3. Catálogos
        estado_cat = "normal" if self.tiene_permiso('catalogos') else "disabled"
        btn_cat = ttk.Button(fr, text="📋  Gestión de Catálogos (Partidas)",
                            bootstyle="warning", state=estado_cat,
                            command=self.abrir_gestor_catalogos)
        btn_cat.pack(fill=X, pady=8, ipady=8)
        if estado_cat == "disabled":
            ToolTip(btn_cat, text="No tienes permiso para editar catálogos")

    # 4. Histórico
        estado_hist = "normal" if self.tiene_permiso('historico') else "disabled"
        btn_hist = ttk.Button(fr, text="📅  Modificar Histórico (Pasado)",
                            bootstyle="danger", state=estado_hist,
                            command=self.abrir_registro_pasado)
        btn_hist.pack(fill=X, pady=8, ipady=8)
        if estado_hist == "disabled":
            ToolTip(btn_hist, text="No tienes permiso para modificar el historial")

    # 5. Ajustes Visuales
        estado_conf = "normal" if self.tiene_permiso('ajustes') else "disabled"
        btn_ajustes = ttk.Button(fr, text="⚙️  Ajustes del Sistema (Logos)",
                                bootstyle="secondary", state=estado_conf,
                                command=self.abrir_ajustes_visuales)
        btn_ajustes.pack(fill=X, pady=8, ipady=8)
        if estado_conf == "disabled":
            ToolTip(btn_ajustes, text="No tienes permiso para cambiar logos")

        ttk.Separator(fr).pack(fill=X, pady=10)

    # 6. ─── CERRAR SESIÓN ───
        def cerrar_sesion():
            top.destroy()
            if messagebox.askyesno(
                "Cerrar Sesión",
                f"¿Deseas cerrar la sesión del usuario '{self.usuario['usuario']}'?\n\n"
                "Volverás a la pantalla de inicio de sesión."
            ):
            # Limpiar ventana principal
                for widget in self.root.winfo_children():
                    widget.destroy()

            # Ocultar ventana mientras reconstruye
                self.root.withdraw()

            # Volver al login
                def volver_login():
                    LoginWindow(self.root, self.db, lambda user_data: iniciar_app_principal(user_data))

                def iniciar_app_principal(usuario_data):
                    for widget in self.root.winfo_children():
                        widget.destroy()
                    self.root.withdraw()
                    SistemaInventario(self.root, self.db, usuario_data)
                    try:
                        self.root.state('zoomed')
                    except:
                        self.root.attributes('-zoomed', True)
                    self.root.update_idletasks()
                    self.root.deiconify()

                volver_login()

        ttk.Button(fr, text="🔒  Cerrar Sesión", bootstyle="danger",
                command=cerrar_sesion).pack(fill=X, pady=8, ipady=8)

        ttk.Separator(fr).pack(fill=X, pady=5)

        ttk.Button(fr, text="Cerrar Menú", bootstyle="outline",
                command=top.destroy).pack(side=BOTTOM, fill=X, pady=(10, 0))

    def abrir_gestor_catalogos(self):
        top = tk.Toplevel(self.root)
        top.title("Gestión de Catálogos")
        top.geometry("700x600")
        top.minsize(600, 500)

        main_frame = ttk.Frame(top)
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        tabs = ttk.Notebook(main_frame)
        tabs.pack(fill=BOTH, expand=True)

        # ══════════════════════════════════════════════
        # PESTAÑA PARTIDAS
        # ══════════════════════════════════════════════
        fr_part = ttk.Frame(tabs, padding=15)
        tabs.add(fr_part, text="📂 Partidas (Códigos)")

        ttk.Label(
            fr_part,
            text="Catálogo de Partidas Presupuestales:",
            font=("Segoe UI", 10, "bold")
        ).pack(anchor=W)

        ttk.Label(
            fr_part,
            text="* Selecciona una fila para editar su descripción",
            font=("Segoe UI", 8), foreground="gray"
        ).pack(anchor=W, pady=(0, 5))

        fr_tree_container = ttk.Frame(fr_part)
        fr_tree_container.pack(fill=BOTH, expand=True, pady=5)

        cols_p = ("CODIGO", "DESCRIPCION")
        tree_part = ttk.Treeview(
            fr_tree_container, columns=cols_p, show="headings")
        tree_part.heading("CODIGO", text="Código")
        tree_part.column( "CODIGO", width=100,
                          anchor=CENTER, stretch=False)
        tree_part.heading("DESCRIPCION", text="Descripción")
        tree_part.column( "DESCRIPCION", width=400, stretch=True)

        sc_p = ttk.Scrollbar(
            fr_tree_container, orient=VERTICAL,
            command=tree_part.yview)
        tree_part.configure(yscrollcommand=sc_p.set)
        tree_part.pack(side=LEFT, fill=BOTH, expand=True)
        sc_p.pack(side=RIGHT, fill=Y)

        def cargar_partidas():
            sel_id   = tree_part.selection()
            prev_cod = (tree_part.item(sel_id[0])['values'][0]
                        if sel_id else None)
            for i in tree_part.get_children():
                tree_part.delete(i)
            filas = self.db.consultar(
                "SELECT valor FROM catalogos "
                "WHERE tipo='PARTIDA' ORDER BY valor ASC")
            for f in filas:
                cod      = f['valor']
                desc_res = self.db.consultar(
                    "SELECT descripcion FROM partidas_desc "
                    "WHERE codigo=?", (cod,))
                nombre = (desc_res[0]['descripcion']
                          if desc_res else "(Sin descripción)")
                item = tree_part.insert(
                    "", END, values=(cod, nombre))
                if prev_cod and str(cod) == str(prev_cod):
                    tree_part.selection_set(item)
                    tree_part.see(item)

        cargar_partidas()

        # Área de edición
        fr_add_p = ttk.LabelFrame(
            fr_part, text=" Editar / Agregar ",
            padding=10, bootstyle="info")
        fr_add_p.pack(fill=X, pady=10)
        fr_add_p.columnconfigure(1, weight=1)

        ttk.Label(fr_add_p, text="Código:").grid(
            row=0, column=0, padx=5, sticky=W)
        e_cod = ttk.Entry(
            fr_add_p, width=15,
            font=("Segoe UI", 10, "bold"))
        e_cod.grid(row=1, column=0, padx=5, sticky=EW)

        ttk.Label(
            fr_add_p,
            text="Descripción (Nombre Largo):"
        ).grid(row=0, column=1, padx=5, sticky=W)
        e_desc = ttk.Entry(fr_add_p, width=40)
        e_desc.grid(row=1, column=1, padx=5, sticky=EW)

        # ✅ Tab entre campos de partida
        e_cod.bind("<Tab>",
            lambda e: (e_desc.focus_set(), "break")[1])
        e_desc.bind("<Shift-Tab>",
            lambda e: (e_cod.focus_set(), "break")[1])

        def al_seleccionar_partida(event):
            sel = tree_part.selection()
            if not sel:
                return
            vals = tree_part.item(sel[0])['values']
            e_cod.delete(0, END)
            e_cod.insert(0, vals[0])
            e_desc.delete(0, END)
            e_desc.insert(0, vals[1])

        tree_part.bind("<<TreeviewSelect>>",
                        al_seleccionar_partida)

        def guardar_partida():
            c = e_cod.get().strip().upper()
            d = e_desc.get().strip().upper()
            if not c:
                messagebox.showwarning(
                    "Error",
                    "El código es obligatorio",
                    parent=top)
                top.lift()
                return
            existe = self.db.consultar(
                "SELECT * FROM catalogos "
                "WHERE tipo='PARTIDA' AND valor=?", (c,))
            if not existe:
                self.db.ejecutar(
                    "INSERT INTO catalogos (tipo, valor) "
                    "VALUES ('PARTIDA', ?)", (c,))
            self.db.ejecutar(
                "REPLACE INTO partidas_desc "
                "(codigo, descripcion) VALUES (?, ?)", (c, d))
            e_cod.delete(0, END)
            e_desc.delete(0, END)
            cargar_partidas()
            self.actualizar_combos()
            messagebox.showinfo(
                "Guardado",
                f"Partida {c} guardada/actualizada.",
                parent=top)
            top.lift()
            e_cod.focus_set()

        # ✅ Enter guarda en ambos campos
        e_cod.bind( "<Return>", lambda e: guardar_partida())
        e_desc.bind("<Return>", lambda e: guardar_partida())

        btn_frame = ttk.Frame(fr_add_p)
        btn_frame.grid(row=1, column=2, padx=10)
        ttk.Button(
            btn_frame,
            text="💾 Guardar / Actualizar",
            bootstyle="success",
            command=guardar_partida
        ).pack(fill=X)

        def eliminar_partida():
            sel = tree_part.selection()
            if not sel:
                messagebox.showwarning(
                    "Atención",
                    "Selecciona una partida para eliminar.",
                    parent=top)
                return
            cod = tree_part.item(sel[0])['values'][0]
            if messagebox.askyesno(
                    "Confirmar",
                    f"¿Borrar partida {cod}?",
                    parent=top):
                self.db.ejecutar(
                    "DELETE FROM catalogos "
                    "WHERE tipo='PARTIDA' AND valor=?", (cod,))
                self.db.ejecutar(
                    "DELETE FROM partidas_desc "
                    "WHERE codigo=?", (cod,))
                cargar_partidas()
                self.actualizar_combos()
                e_cod.delete(0, END)
                e_desc.delete(0, END)
                top.lift()

        ttk.Button(
            fr_part,
            text="🗑️ Eliminar Seleccionada",
            bootstyle="danger",
            command=eliminar_partida
        ).pack(fill=X, pady=(0, 5))

        # ══════════════════════════════════════════════
        # PESTAÑAS SIMPLES (Áreas y Jefes)
        # ══════════════════════════════════════════════
        def crear_tab_lista_simple(tipo_cat, titulo):
            fr = ttk.Frame(tabs, padding=15)
            tabs.add(fr, text=titulo)

            fr_list_cont = ttk.Frame(fr)
            fr_list_cont.pack(fill=BOTH, expand=True, pady=5)

            lst = tk.Listbox(fr_list_cont, height=10)
            lst.pack(side=LEFT, fill=BOTH, expand=True)

            sb = ttk.Scrollbar(
                fr_list_cont, orient=VERTICAL,
                command=lst.yview)
            sb.pack(side=RIGHT, fill=Y)
            lst.config(yscrollcommand=sb.set)

            def cargar():
                lst.delete(0, END)
                fs = self.db.consultar(
                    "SELECT valor FROM catalogos "
                    "WHERE tipo=? ORDER BY valor ASC",
                    (tipo_cat,))
                for x in fs:
                    lst.insert(END, x['valor'])

            cargar()

            fr_controls = ttk.Frame(fr)
            fr_controls.pack(fill=X, pady=5)

            e_val = ttk.Entry(fr_controls)
            e_val.pack(
                side=LEFT, fill=X,
                expand=True, padx=(0, 5))

            def add():
                v = e_val.get().strip().upper()
                if v:
                    self.db.ejecutar(
                        "INSERT INTO catalogos "
                        "(tipo, valor) VALUES (?, ?)",
                        (tipo_cat, v))
                    e_val.delete(0, END)
                    cargar()
                    self.actualizar_combos()
                    e_val.focus_set()

            def delete():
                s = lst.curselection()
                if s:
                    v = lst.get(s[0])
                    if messagebox.askyesno(
                            "Confirmar",
                            f"¿Eliminar '{v}'?",
                            parent=top):
                        self.db.ejecutar(
                            "DELETE FROM catalogos "
                            "WHERE tipo=? AND valor=?",
                            (tipo_cat, v))
                        cargar()
                        self.actualizar_combos()
                        top.lift()

            # ✅ Enter agrega el elemento
            e_val.bind("<Return>", lambda e: add())

            ttk.Button(
                fr_controls,
                text="➕ Agregar",
                bootstyle="success",
                command=add
            ).pack(side=LEFT)

            ttk.Button(
                fr,
                text="🗑️ Eliminar Seleccionado",
                bootstyle="danger",
                command=delete
            ).pack(fill=X)

        crear_tab_lista_simple("AREA", "🏢 Áreas")
        crear_tab_lista_simple("JEFE", "👤 Jefes")

        # ══════════════════════════════════════════════
        # PESTAÑA CÓNSTAME
        # ══════════════════════════════════════════════
        fr_const = ttk.Frame(tabs, padding=15)
        tabs.add(fr_const, text="✍️ Cónstame")

        ttk.Label(
            fr_const,
            text="Nombre / Firma de Autoridad (Cónstame):",
            font=("Segoe UI", 10, "bold")
        ).pack(anchor=W)

        ttk.Label(
            fr_const,
            text="Este nombre aparece en la parte inferior "
                 "de los Vales de Salida PDF.",
            font=("Segoe UI", 8), foreground="gray"
        ).pack(anchor=W, pady=(0, 10))

        e_const = ttk.Entry(fr_const, font=("Segoe UI", 11))
        e_const.pack(fill=X, pady=5)

        # Cargar valor actual
        res_firma = self.db.consultar(
            "SELECT valor FROM catalogos WHERE tipo='FIRMA'")
        if res_firma:
            e_const.insert(0, res_firma[0]['valor'])

        def guardar_constame():
            nueva_firma = e_const.get().strip().upper()
            if nueva_firma:
                self.db.ejecutar(
                    "DELETE FROM catalogos WHERE tipo='FIRMA'")
                self.db.ejecutar(
                    "INSERT INTO catalogos (tipo, valor) "
                    "VALUES ('FIRMA', ?)", (nueva_firma,))
                messagebox.showinfo(
                    "Guardado",
                    "Firma 'Cónstame' actualizada correctamente.",
                    parent=top)
                top.lift()
                e_const.focus_set()
            else:
                messagebox.showwarning(
                    "Atención",
                    "El campo no puede estar vacío.",
                    parent=top)
                top.lift()

        # ✅ Enter guarda la firma
        e_const.bind("<Return>", lambda e: guardar_constame())

        ttk.Button(
            fr_const,
            text="💾 Guardar Firma",
            bootstyle="success",
            command=guardar_constame
        ).pack(pady=15, fill=X)


    def actualizar_combos(self):
        # 1. PARTIDAS
        rows    = self.db.consultar(
            "SELECT valor FROM catalogos WHERE tipo='PARTIDA' ORDER BY valor ASC"
        )
        lista_p = [r['valor'] for r in rows]

        if hasattr(self, 'cb_partida'):
            self.cb_partida['values'] = lista_p
            self.setup_autocomplete(self.cb_partida, lista_p)

        if hasattr(self, 'cb_filtro_partida'):
            lista_p_todas = ["TODAS"] + lista_p
            self.cb_filtro_partida['values'] = lista_p_todas
            if not self.cb_filtro_partida.get():
                self.cb_filtro_partida.current(0)

        # FIX BUG 3: condición correcta — cb_partida_k es independiente de cb_partida_consumo
        if hasattr(self, 'cb_partida_k'):
            lista_pk = ["TODAS"] + lista_p
            self.cb_partida_k['values'] = lista_pk
            if not self.cb_partida_k.get():
                self.cb_partida_k.current(0)
            # readonly — la flecha nativa lo abre, sin event_generate
            self.cb_partida_k.bind('<<ComboboxSelected>>',
                lambda e: self.cb_partida_k.selection_clear())

        # 2. ÁREAS
        rows    = self.db.consultar(
            "SELECT valor FROM catalogos WHERE tipo='AREA' ORDER BY valor ASC"
        )
        lista_a = [r['valor'] for r in rows]
        if hasattr(self, 'cb_area_sal'):
            self.setup_autocomplete(self.cb_area_sal, lista_a)

        # 3. JEFES
        rows    = self.db.consultar(
            "SELECT valor FROM catalogos WHERE tipo='JEFE' ORDER BY valor ASC"
        )
        lista_j = [r['valor'] for r in rows]
        if hasattr(self, 'cb_jefe_sal'):
            self.setup_autocomplete(self.cb_jefe_sal, lista_j)

        # 4. Buscador es Entry — no necesita values
        # (se filtra en tiempo real con cargar_tabla_inventario)

        # FIX BUG 4: cb_partida_consumo es readonly, sin event_generate
        if hasattr(self, 'cb_partida_consumo'):
            self.cb_partida_consumo['values'] = lista_p
            if not self.cb_partida_consumo.get() and lista_p:
                self.cb_partida_consumo.current(0)
            self.cb_partida_consumo.bind('<<ComboboxSelected>>',
                lambda e: self.cb_partida_consumo.selection_clear())
            
    def abrir_registro_pasado(self):
        """Registro Histórico Manual CON DISEÑO RESPONSIVO"""
        top = tb.Toplevel(self.root)
        top.title("Registro Histórico Manual")
        top.geometry("500x700")
        top.minsize(450, 600)

        ttk.Label(
            top,
            text="⚠️ CUIDADO: Esto afecta el stock actual.",
            bootstyle="warning"
        ).pack(pady=10)

        fr = ttk.Frame(top, padding=20)
        fr.pack(fill=BOTH, expand=True)

        # 1. FECHA
        ttk.Label(fr, text="Fecha del Movimiento:").pack(anchor=W)
        e_fecha = tb.DateEntry(fr, bootstyle="info", dateformat="%d/%m/%Y")
        e_fecha.pack(fill=X, pady=(0, 10))

        # 2. TIPO
        ttk.Label(fr, text="Tipo de Movimiento:").pack(anchor=W)
        c_tipo = ttk.Combobox(
            fr,
            values=["HISTORICO (+) Entrada/Saldo Inicial", "HISTORICO (-) Salida/Ajuste"],
            state="readonly"
        )
        c_tipo.current(0)
        c_tipo.pack(fill=X, pady=(0, 10))

        # 3. PARTIDA
        ttk.Label(fr, text="Partida (Obligatorio):").pack(anchor=W)
        vals_partidas = []
        if hasattr(self, 'cb_partida'):
            vals_partidas = self.cb_partida['values']

        c_partida_hist = ttk.Combobox(fr, values=vals_partidas, state="readonly")
        c_partida_hist.pack(fill=X, pady=(0, 10))
        self.setup_autocomplete(c_partida_hist, list(vals_partidas))

        # 4. MATERIAL
       
        ttk.Label(fr, text="Material Exacto (Busca):").pack(anchor=W)
        rows_mat = self.db.consultar(
            "SELECT DISTINCT material FROM inventario ORDER BY material ASC"
        )
        vals_mat = [r['material'] for r in rows_mat]

        c_mat = ttk.Combobox(fr, values=vals_mat)
        c_mat.pack(fill=X, pady=(0, 10))
        self.setup_autocomplete(c_mat, vals_mat)

        def al_elegir_material(event):
            mat_name = c_mat.get()
            res = self.db.consultar(
                "SELECT partida FROM inventario WHERE material=?", (mat_name,)
            )
            if res:
                c_partida_hist.set(res[0]['partida'])

        c_mat.bind("<<ComboboxSelected>>", al_elegir_material)

        # 5. CANTIDAD
        ttk.Label(fr, text="Cantidad:").pack(anchor=W)
        e_cant = ttk.Entry(fr)
        e_cant.pack(fill=X, pady=(0, 10))

        # 6. FACTURA
        ttk.Label(fr, text="Factura / Documento (Opcional):").pack(anchor=W)
        e_factura_hist = ttk.Entry(fr)
        e_factura_hist.pack(fill=X, pady=(0, 10))

        # 7. OBSERVACIÓN
        ttk.Label(fr, text="Observación / Responsable:").pack(anchor=W)
        e_obs = ttk.Entry(fr)
        e_obs.pack(fill=X, pady=(0, 20))

        # ── Lógica de guardado ─────────────────────────────────────────
        def guardar_historico():
            mat    = c_mat.get().strip().upper()
            part   = c_partida_hist.get().strip()
            tipo_sel = c_tipo.get()
            fecha  = e_fecha.entry.get()
            obs    = e_obs.get().strip().upper() or "AJUSTE MANUAL"
            fact   = e_factura_hist.get().strip().upper() or "S/F"
            usuario_act = self.usuario.get('usuario', 'SISTEMA')

            if not mat or not fecha or not part:
                messagebox.showwarning(
                    "Faltan datos",
                    "Material, Partida y Fecha son obligatorios"
                )
                return

            try:
                cant = float(e_cant.get())
                if cant <= 0:
                    raise ValueError
            except:
                messagebox.showerror("Error", "Cantidad inválida")
                return

            existe = self.db.consultar(
                "SELECT id FROM inventario WHERE material = ? AND partida = ?",
                (mat, part)
            )

            if "(+)" in tipo_sel:
                tipo_db = "HISTORICO (+)"
                if existe:
                    self.db.ejecutar(
                        "UPDATE inventario SET stock = stock + ? WHERE material = ? AND partida = ?",
                        (cant, mat, part)
                    )
                else:
                    if messagebox.askyesno(
                            "Nuevo Material",
                            "Este material no existe en esa partida. ¿Crearlo con este stock inicial?"):
                        self.db.ejecutar(
                            "INSERT INTO inventario (partida, material, stock, ultimo_movimiento, factura) "
                            "VALUES (?, ?, ?, ?, ?)",
                            (part, mat, cant, fecha, fact)
                        )
                    else:
                        return
            else:
                tipo_db = "HISTORICO (-)"
                if existe:
                    self.db.ejecutar(
                        "UPDATE inventario SET stock = stock - ? WHERE material = ? AND partida = ?",
                        (cant, mat, part)
                    )
                else:
                    messagebox.showerror(
                        "Error",
                        "No puedes restar stock de un material que no existe."
                    )
                    return

            try:
                self.db.ejecutar(
                    "INSERT INTO historial "
                    "(fecha_hora, tipo, partida, material, cantidad, "
                    "responsable, entrego, factura, usuario_sistema) "
                    "VALUES (?, ?, ?, ?, ?, ?, 'AJUSTE HISTORICO', ?, ?)",
                    (fecha, tipo_db, part, mat, cant,
                     obs, fact, usuario_act)
                )

                messagebox.showinfo("Éxito", "Registro histórico aplicado correctamente.")
                self.cargar_tabla_inventario()
                self.cargar_tabla_historial()
                top.destroy()

            except Exception as e:
                messagebox.showerror("Error DB", f"{e}")

        ttk.Button(
            fr,
            text="💾 APLICAR MOVIMIENTO",
            bootstyle="success",
            command=guardar_historico
        ).pack(fill=X, side=BOTTOM, pady=10, ipady=5)

    def generar_folio(self):
        rows = self.db.consultar("SELECT COUNT(*) as total FROM historial WHERE tipo='SALIDA'")
        consecutivo = rows[0]['total'] + 1
        return f"{consecutivo:03d}-{datetime.now().year}"

    # --- REEMPLAZA ESTA FUNCIÓN COMPLETA ---
    # --- REEMPLAZA ESTA FUNCIÓN COMPLETA ---
    # --- REEMPLAZA ESTA FUNCIÓN COMPLETA PARA RECUPERAR EL FORMATO DE TABLA ---
    # --- REEMPLAZA ESTA FUNCIÓN COMPLETA PARA TENER EL FORMATO TABLA EXACTO ---
    # --- REEMPLAZA ESTA FUNCIÓN COMPLETA (AJUSTE DE FIRMAS) ---
    # --- REEMPLAZA ESTA FUNCIÓN COMPLETA (CORRECCIÓN DE ANCHO DE TEXTO) ---
    # --- REEMPLAZA ESTA FUNCIÓN COMPLETA (AJUSTE FINAL DE POSICIÓN DEL LOGO) ---
    def generar_pdf_vale_multiple(self, carrito, area, resp, jefe, folio):
        """
        Vale de Salida — Tamaño Carta (8.5 × 11 in).
        Regla de renglones:
          • Mínimo 4 renglones siempre visibles.
          • Si hay más de 4 materiales → exactamente len(carrito) renglones.
          • Encabezado ordenado: logo izq + título centro + subtítulo.
          • Firmas siempre fijas al pie de la hoja.
        """
        try:
            from matplotlib.patches import Rectangle
            from matplotlib.offsetbox import OffsetImage, AnnotationBbox
            from PIL import Image as PilImage

            # ── Ruta de salida ─────────────────────────────────────────
            escritorio = os.path.join(os.environ['USERPROFILE'], 'Desktop')
            ruta_pdf   = os.path.join(
                escritorio,
                f"VALE_{folio.replace('/', '-')}.pdf"
            )

            # ── Datos de configuración ─────────────────────────────────
            empresa_titulo  = self.db.get_config("TITULO_APP")    or "NOMBRE EMPRESA"
            depto_subtitulo = self.db.get_config("SUBTITULO_APP") or "DEPARTAMENTO"
            res_firma       = self.db.consultar(
                "SELECT valor FROM catalogos WHERE tipo='FIRMA'"
            )
            firma_constame = res_firma[0]['valor'] if res_firma else "AUTORIDAD"

            # ── Tamaño de fuente adaptable al largo del título ─────────
            n = len(empresa_titulo)
            if   n <= 12:  fs_titulo = 22
            elif n <= 20:  fs_titulo = 19
            elif n <= 30:  fs_titulo = 16
            elif n <= 42:  fs_titulo = 13
            elif n <= 55:  fs_titulo = 11
            else:          fs_titulo = 9

            # ── Figura tamaño carta SIN recorte automático ─────────────
            plt.switch_backend('Agg')
            fig = plt.figure(figsize=(8.5, 11))
            ax  = fig.add_subplot(111)
            ax.set_xlim(0, 8.5)
            ax.set_ylim(0, 11)
            ax.axis('off')

            # Márgenes exactos en fracción de la figura
            fig.subplots_adjust(
                left   = 0.04,
                right  = 0.96,
                top    = 0.97,
                bottom = 0.03
            )

            AZUL = "#1F4E79"
            ROJO = "#C00000"

            # ── Columnas X ────────────────────────────────────────────
            X_INI  = 0.40
            X_COL1 = 1.65
            X_COL2 = 2.90
            X_FIN  = 8.10
            X_MID  = (X_INI + X_FIN) / 2   # = 4.25

            # ══════════════════════════════════════════════════════════
            # ENCABEZADO
            # ══════════════════════════════════════════════════════════

            # 1. Barra azul decorativa superior
            ax.add_patch(Rectangle(
                (X_INI, 10.70), X_FIN - X_INI, 0.10,
                facecolor=AZUL, zorder=2
            ))

            # 2. Logo — izquierda, alineado con el bloque de títulos
            logo_path = (
                self.db.get_config("LOGO_PDF") or
                self.db.get_config("LOGO_APP")
            )
            logo_cargado = False
            if logo_path and os.path.exists(logo_path):
                try:
                    img_pil  = PilImage.open(logo_path).copy()
                    img_pil.thumbnail((140, 140), PilImage.LANCZOS)
                    img_rgba = img_pil.convert("RGBA")
                    fondo    = PilImage.new("RGBA", img_rgba.size, "#FFFFFF")
                    fondo.paste(img_rgba, mask=img_rgba.split()[3])
                    img_np   = np.array(fondo.convert("RGB"))
                    zoom     = 55 / max(img_np.shape[0], img_np.shape[1])
                    imagebox = OffsetImage(img_np, zoom=zoom)
                    imagebox.image.axes = ax
                    ab = AnnotationBbox(
                        imagebox,
                        (X_INI + 0.50, 10.32),
                        frameon=False,
                        box_alignment=(0.5, 0.5),
                        zorder=3
                    )
                    ax.add_artist(ab)
                    logo_cargado = True
                except Exception as e:
                    print(f"Error logo PDF: {e}")

            # Si no hay logo, mostrar ícono de texto como fallback
            if not logo_cargado:
                ax.text(
                    X_INI + 0.50, 10.32,
                    "⬜",
                    fontsize=28, ha='center', va='center',
                    color='#CCCCCC'
                )

            # 3. Título principal — centrado en el área derecha al logo
            # El logo ocupa hasta aprox X_INI + 1.05
            # El título se centra en el espacio restante
            X_TITULO_CENTER = (X_INI + 1.10 + X_FIN) / 2   # ≈ 4.75

            ax.text(
                X_TITULO_CENTER, 10.44,
                empresa_titulo,
                fontsize=fs_titulo, fontweight='bold',
                color=AZUL, ha='center', va='center'
            )

            # 4. Subtítulo — mismo centro que el título
            ax.text(
                X_TITULO_CENTER, 10.16,
                depto_subtitulo.upper(),
                fontsize=9, fontweight='bold',
                color='#666666', ha='center', va='center'
            )

            # 5. Línea divisoria bajo el bloque de títulos
            ax.plot(
                [X_INI, X_FIN], [9.88, 9.88],
                color=AZUL, linewidth=1.4
            )

            # 6. Folio (izquierda)
            ax.text(
                X_INI, 9.66,
                "FOLIO: ",
                fontsize=10, fontweight='bold',
                color=ROJO, va='center'
            )
            ax.text(
                X_INI + 0.72, 9.66,
                folio,
                fontsize=10, fontweight='bold',
                color=ROJO, va='center'
            )

            # 7. Fecha (derecha)
            ax.text(
                X_FIN, 9.66,
                f"FECHA:  {datetime.now().strftime('%d-%b-%Y').upper()}",
                fontsize=9.5, fontweight='bold',
                color=AZUL, ha='right', va='center'
            )

            # 8. Caja "VALE DE SALIDA" centrada en la hoja
            CAJA_W = 3.20
            CAJA_H = 0.34
            CAJA_X = X_MID - CAJA_W / 2
            CAJA_Y = 9.22
            ax.add_patch(Rectangle(
                (CAJA_X, CAJA_Y), CAJA_W, CAJA_H,
                facecolor=AZUL, zorder=2
            ))
            ax.text(
                X_MID, CAJA_Y + CAJA_H / 2,
                "V A L E   D E   S A L I D A",
                color='white', fontweight='bold', fontsize=11,
                ha='center', va='center', zorder=3
            )

            # 9. Línea divisoria
            ax.plot(
                [X_INI, X_FIN], [9.08, 9.08],
                color=AZUL, linewidth=1.0
            )

            # 10. Área solicitante
            ax.text(
                X_INI, 8.86,
                "ÁREA SOLICITANTE:",
                fontweight='bold', color=AZUL,
                fontsize=9.5, va='center'
            )
            ax.text(
                X_INI + 1.92, 8.86,
                area.upper(),
                fontsize=9.5, va='center', color='#222222'
            )

            # 11. Línea gris bajo área solicitante
            ax.plot(
                [X_INI, X_FIN], [8.68, 8.68],
                color='#AAAAAA', linewidth=0.6
            )

            # ══════════════════════════════════════════════════════════
            # ENCABEZADO DE TABLA (fondo azul)
            # ══════════════════════════════════════════════════════════
            H_HEADER    = 0.40
            Y_TABLA_TOP = 8.52
            Y_HDR_BOT   = Y_TABLA_TOP - H_HEADER   # = 8.12

            ax.add_patch(Rectangle(
                (X_INI, Y_HDR_BOT), X_FIN - X_INI, H_HEADER,
                facecolor=AZUL, zorder=2
            ))

            y_hdr_c = Y_HDR_BOT + H_HEADER / 2

            ax.text(
                (X_INI + X_COL1) / 2, y_hdr_c,
                "CANT.",
                color='white', fontweight='bold',
                ha='center', va='center', fontsize=9, zorder=3
            )
            ax.text(
                (X_COL1 + X_COL2) / 2, y_hdr_c,
                "UNIDAD",
                color='white', fontweight='bold',
                ha='center', va='center', fontsize=9, zorder=3
            )
            ax.text(
                X_COL2 + 0.15, y_hdr_c,
                "DESCRIPCIÓN DEL MATERIAL",
                color='white', fontweight='bold',
                ha='left', va='center', fontsize=9, zorder=3
            )

            # ══════════════════════════════════════════════════════════
            # CÁLCULO DE RENGLONES
            #   • Mínimo 4 renglones siempre.
            #   • Más materiales → exactamente ese número de renglones.
            #   • Altura fija 0.48 in por renglón.
            # ══════════════════════════════════════════════════════════
            H_ROW      = 0.48
            num_filas  = max(len(carrito), 4)
            Y_TOP_ROWS = Y_HDR_BOT
            Y_BOT_ROWS = Y_TOP_ROWS - H_ROW * num_filas

            # ── Filas alternas con fondo suave ────────────────────────
            for i in range(num_filas):
                if i % 2 == 1:
                    y_rect = Y_TOP_ROWS - H_ROW * (i + 1)
                    ax.add_patch(Rectangle(
                        (X_INI, y_rect), X_FIN - X_INI, H_ROW,
                        facecolor='#F4F7FB', zorder=0
                    ))

            # ── Líneas horizontales ───────────────────────────────────
            for i in range(num_filas + 1):
                y_lin    = Y_TOP_ROWS - H_ROW * i
                es_borde = (i == 0 or i == num_filas)
                ax.plot(
                    [X_INI, X_FIN], [y_lin, y_lin],
                    color     = AZUL    if es_borde else '#CCCCCC',
                    linewidth = 0.85    if es_borde else 0.30
                )

            # ── Líneas verticales — bordes exteriores ─────────────────
            ax.plot([X_INI, X_INI], [Y_BOT_ROWS, Y_TOP_ROWS],
                    color=AZUL, linewidth=0.85)
            ax.plot([X_FIN, X_FIN], [Y_BOT_ROWS, Y_TOP_ROWS],
                    color=AZUL, linewidth=0.85)

            # ── Líneas verticales — divisores interiores ──────────────
            ax.plot([X_COL1, X_COL1], [Y_BOT_ROWS, Y_TOP_ROWS],
                    color='#CCCCCC', linewidth=0.40)
            ax.plot([X_COL2, X_COL2], [Y_BOT_ROWS, Y_TOP_ROWS],
                    color='#CCCCCC', linewidth=0.40)

            # ══════════════════════════════════════════════════════════
            # DATOS DEL CARRITO
            # ══════════════════════════════════════════════════════════
            for i, item in enumerate(carrito):
                y_centro = Y_TOP_ROWS - H_ROW * i - H_ROW / 2

                ax.text(
                    (X_INI + X_COL1) / 2, y_centro,
                    str(item['cantidad']),
                    ha='center', va='center',
                    fontsize=9.5, fontweight='bold'
                )
                ax.text(
                    (X_COL1 + X_COL2) / 2, y_centro,
                    "PZA",
                    ha='center', va='center', fontsize=9
                )
                ax.text(
                    X_COL2 + 0.15, y_centro,
                    textwrap.fill(item['material'], 56),
                    ha='left', va='center',
                    fontsize=8.5, linespacing=1.30
                )

            # ══════════════════════════════════════════════════════════
            # FIRMAS — FIJAS AL PIE
            # ══════════════════════════════════════════════════════════
           # ── Firmas ────────────────────────────────────────────────
            # ── Firmas ────────────────────────────────────────────────
            BARRA_AZUL_Y  = 0.18
            Y_LINEA_FIRMA = 1.45   # Más espacio hacia arriba para textos largos
            FIRMA_ANCHO   = 1.10

            def dibujar_firma(xc, titulo, nombre, chars_linea=22, max_lineas=5):
                """
                chars_linea: cuántos caracteres por línea
                max_lineas:  máximo de renglones antes de recortar
                """
                # Línea de firma
                ax.plot(
                    [xc - FIRMA_ANCHO, xc + FIRMA_ANCHO],
                    [Y_LINEA_FIRMA, Y_LINEA_FIRMA],
                    color='#333333', linewidth=0.9
                )

                # Título
                ax.text(
                    xc, Y_LINEA_FIRMA - 0.13,
                    titulo,
                    ha='center', va='top',
                    fontweight='bold', fontsize=8.5, color=AZUL
                )

                # Nombre wrapeado
                lineas     = textwrap.wrap(nombre, chars_linea)
                lineas     = lineas[:max_lineas]
                nombre_fmt = "\n".join(lineas)

                # Fontsize dinámico: más texto → letra más chica
                n = len(lineas)
                if   n <= 2: fs = 8.0
                elif n <= 3: fs = 7.5
                elif n <= 4: fs = 7.0
                else:        fs = 6.5

                ax.text(
                    xc, Y_LINEA_FIRMA - 0.28,
                    nombre_fmt,
                    ha='center', va='top',
                    fontsize=fs,
                    color='#333333',
                    linespacing=1.20
                )

            # ENTREGÓ y RECIBIÓ — normal
            dibujar_firma(1.95, "ENTREGÓ",            jefe,           chars_linea=22, max_lineas=5)
            dibujar_firma(4.25, "RECIBIÓ / SOLICITÓ", resp,           chars_linea=22, max_lineas=5)
            # CONSTAME — más caracteres por línea y más renglones
            dibujar_firma(6.55, "CONSTAME:",           firma_constame, chars_linea=26, max_lineas=6)

            # Barra azul inferior fija al fondo
            ax.add_patch(Rectangle(
                (X_INI, BARRA_AZUL_Y), X_FIN - X_INI, 0.10,
                facecolor=AZUL, zorder=2
            ))

            # ── Guardar PDF SIN bbox_inches='tight' ───────────────────
            # (tight recorta la figura y desplaza coordenadas)
            fig.savefig(ruta_pdf, dpi=300)
            plt.close(fig)

            intentos = 0
            while not os.path.exists(ruta_pdf) and intentos < 20:
                time.sleep(0.1)
                intentos += 1

            if os.path.exists(ruta_pdf):
                os.startfile(ruta_pdf)

        except Exception as e:
            messagebox.showerror("Error PDF", f"{e}")
            plt.close()
    
    def abrir_ajustes_visuales(self):
        # 1. Verificar Permisos
        if self.usuario.get('rol') != 'ADMIN' and not self.tiene_permiso('ajustes'):
            messagebox.showerror("Acceso Denegado", "No tienes permiso para modificar la configuración del sistema.")
            return

        top = tk.Toplevel(self.root)
        top.title("Configuración Visual y de Reportes")
        # Ventana un poco más grande para que quepa la nueva sección
        self.centrar_ventana_emergente(top, 900, 750) 

        # Contenedor con Scroll
        canvas = tk.Canvas(top, highlightthickness=0)
        scrollbar = ttk.Scrollbar(top, orient=VERTICAL, command=canvas.yview)
        fr = ttk.Frame(canvas, padding=30) 

        scrollable_window = canvas.create_window((0, 0), window=fr, anchor="nw")

        def configure_scroll_region(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(scrollable_window, width=canvas.winfo_width())
        
        canvas.bind("<Configure>", configure_scroll_region)
        fr.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)

        # --- CONTENIDO ---
        ttk.Label(fr, text="Personalización del Sistema", font=("Segoe UI", 18, "bold"), bootstyle="primary").pack(pady=(0, 20), anchor=CENTER)

        # ========================================================
        # SECCIÓN 1: LOGOTIPOS (DISTRIBUCIÓN HORIZONTAL)
        # ========================================================
        fr_imgs = ttk.LabelFrame(fr, text=" 🖼️ Logotipos del Sistema ", padding=15, bootstyle="info")
        fr_imgs.pack(fill=X, pady=10)

        f_izq = ttk.Frame(fr_imgs); f_izq.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 10))
        f_der = ttk.Frame(fr_imgs); f_der.pack(side=LEFT, fill=BOTH, expand=True, padx=(10, 0))

        ttk.Label(f_izq, text="Logo Interfaz (Pantalla):", font=("Segoe UI", 9, "bold")).pack(anchor=W)
        cont_l1 = ttk.Frame(f_izq)
        cont_l1.pack(fill=X, pady=5)
        self.e_logo_app = ttk.Entry(cont_l1)
        self.e_logo_app.pack(side=LEFT, fill=X, expand=True, padx=(0,5))
        def b_app():
            r = filedialog.askopenfilename(parent=top, filetypes=[("Imágenes", "*.png;*.jpg;*.ico")])
            if r: self.e_logo_app.delete(0, END); self.e_logo_app.insert(0, r); top.lift()
        ttk.Button(cont_l1, text="📂 Buscar", command=b_app).pack(side=LEFT)

        ttk.Label(f_der, text="Logo Reportes (PDF/Excel):", font=("Segoe UI", 9, "bold")).pack(anchor=W)
        cont_l2 = ttk.Frame(f_der)
        cont_l2.pack(fill=X, pady=5)
        self.e_logo_pdf = ttk.Entry(cont_l2)
        self.e_logo_pdf.pack(side=LEFT, fill=X, expand=True, padx=(0,5))
        def b_pdf():
            r = filedialog.askopenfilename(parent=top, filetypes=[("Imágenes", "*.png;*.jpg;*.jpeg")])
            if r: self.e_logo_pdf.delete(0, END); self.e_logo_pdf.insert(0, r); top.lift()
        ttk.Button(cont_l2, text="📂 Buscar", command=b_pdf).pack(side=LEFT)

        # ========================================================
        # SECCIÓN 2: TÍTULOS (DISTRIBUCIÓN HORIZONTAL)
        # ========================================================
        fr_txt = ttk.LabelFrame(fr, text=" 🏷️ Títulos de la Ventana ", padding=15, bootstyle="secondary")
        fr_txt.pack(fill=X, pady=10)

        f_t1 = ttk.Frame(fr_txt); f_t1.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 10))
        f_t2 = ttk.Frame(fr_txt); f_t2.pack(side=LEFT, fill=BOTH, expand=True, padx=(10, 0))

        ttk.Label(f_t1, text="Título Principal (Barra Superior):", font=("Segoe UI", 9, "bold")).pack(anchor=W)
        self.e_titulo = ttk.Entry(f_t1); self.e_titulo.pack(fill=X, pady=5)

        ttk.Label(f_t2, text="Subtítulo (Descripción):", font=("Segoe UI", 9, "bold")).pack(anchor=W)
        self.e_subtitulo = ttk.Entry(f_t2); self.e_subtitulo.pack(fill=X, pady=5)

        # ========================================================
        # SECCIÓN 3: ENCABEZADOS DE REPORTES (ANCHO COMPLETO)
        # ========================================================
        fr_rep = ttk.LabelFrame(fr, text=" 📄 Membrete / Encabezados (Excel y PDF) ", padding=15, bootstyle="warning")
        fr_rep.pack(fill=X, pady=10)
        
        fr_rep.columnconfigure(1, weight=1)

        ttk.Label(fr_rep, text="Línea 1 (Institución):").grid(row=0, column=0, sticky=W, pady=5)
        self.e_h1 = ttk.Entry(fr_rep); self.e_h1.grid(row=0, column=1, sticky=EW, padx=10, pady=5)
        
        ttk.Label(fr_rep, text="Línea 2 (Sub-Institución):").grid(row=1, column=0, sticky=W, pady=5)
        self.e_h2 = ttk.Entry(fr_rep); self.e_h2.grid(row=1, column=1, sticky=EW, padx=10, pady=5)
        
        ttk.Label(fr_rep, text="Línea 3 (Dirección General):").grid(row=2, column=0, sticky=W, pady=5)
        self.e_h3 = ttk.Entry(fr_rep); self.e_h3.grid(row=2, column=1, sticky=EW, padx=10, pady=5)
        
        ttk.Label(fr_rep, text="Línea 4 (Unidad/Depto):").grid(row=3, column=0, sticky=W, pady=5)
        self.e_h4 = ttk.Entry(fr_rep); self.e_h4.grid(row=3, column=1, sticky=EW, padx=10, pady=5)

        # ========================================================
        # SECCIÓN 4: INFORMACIÓN DE LA BASE DE DATOS (SOLO ADMIN)
        # ========================================================
        if self.usuario.get('rol') == 'ADMIN':
            fr_db = ttk.LabelFrame(fr, text=" 🗄️ Ruta de la Base de Datos (Modo Admin) ", padding=15, bootstyle="danger")
            fr_db.pack(fill=X, pady=10)

            ttk.Label(fr_db, text="El sistema está conectado actualmente al siguiente archivo:", font=("Segoe UI", 9)).pack(anchor=W, pady=(0,5))
            
            e_ruta_db = ttk.Entry(fr_db, font=("Consolas", 10, "bold"))
            e_ruta_db.pack(fill=X, pady=5)
            # Insertamos la ruta real que está usando el gestor
            e_ruta_db.insert(0, self.db.ruta_db) 
            e_ruta_db.configure(state="readonly") # Para que puedan copiarlo pero no borrarlo

            def abrir_carpeta_db():
                directorio = os.path.dirname(self.db.ruta_db)
                if not directorio: directorio = os.getcwd()
                if os.path.exists(directorio):
                    os.startfile(directorio)
                else:
                    messagebox.showwarning("Aviso", "La carpeta no se puede abrir directamente.")

            ttk.Button(fr_db, text="📂 Abrir ubicación del archivo", bootstyle="outline-danger", command=abrir_carpeta_db).pack(anchor=E, pady=(5,0))

        # --- CARGA DE DATOS ---
        self.e_logo_app.insert(0, self.db.get_config("LOGO_APP") or "")
        self.e_logo_pdf.insert(0, self.db.get_config("LOGO_PDF") or "")
        self.e_titulo.insert(0, self.db.get_config("TITULO_APP") or "SISTEMA INVENTARIO")
        self.e_subtitulo.insert(0, self.db.get_config("SUBTITULO_APP") or "CONTROL DE STOCK")
        
        self.e_h1.insert(0, self.db.get_config("HEADER_L1") or "SECRETARÍA DE MARINA")
        self.e_h2.insert(0, self.db.get_config("HEADER_L2") or "SUBSECRETARÍA DE MARINA")
        self.e_h3.insert(0, self.db.get_config("HEADER_L3") or "DIRECCIÓN GENERAL DE INDUSTRIA NAVAL")
        self.e_h4.insert(0, self.db.get_config("HEADER_L4") or "UNIDAD DE INVESTIGACIÓN Y DESARROLLO TECNOLÓGICO")

        # --- GUARDAR ---
        def guardar_cambios():
            self.db.set_config("LOGO_APP", self.e_logo_app.get().strip())
            self.db.set_config("LOGO_PDF", self.e_logo_pdf.get().strip())
            self.db.set_config("TITULO_APP", self.e_titulo.get().strip())
            self.db.set_config("SUBTITULO_APP", self.e_subtitulo.get().strip())
            
            self.db.set_config("HEADER_L1", self.e_h1.get().strip())
            self.db.set_config("HEADER_L2", self.e_h2.get().strip())
            self.db.set_config("HEADER_L3", self.e_h3.get().strip())
            self.db.set_config("HEADER_L4", self.e_h4.get().strip())

            if messagebox.askyesno("Reiniciar", "Configuración guardada.\n¿Reiniciar sistema ahora para ver cambios?"):
                import sys, subprocess
                top.destroy(); self.root.destroy()
                script = f'"{sys.argv[0]}"' if " " in sys.argv[0] else sys.argv[0]
                subprocess.Popen(f"{sys.executable} {script}", shell=True)
                sys.exit()
            else:
                top.destroy()

        ttk.Button(fr, text="💾 GUARDAR TODA LA CONFIGURACIÓN", bootstyle="success", command=guardar_cambios).pack(fill=X, pady=20)

        

        # --- GUARDAR ---
        
    
    def abrir_editor_temas(self):
        top = tk.Toplevel(self.root)
        top.title("🎨 Personalización de Temas")
        self.centrar_ventana_emergente(top, 780, 720)

        tabs = ttk.Notebook(top)
        tabs.pack(fill=BOTH, expand=True, padx=10, pady=10)

    # ─────────────────────────────────────────────
    # PESTAÑA 1: TEMAS PREDEFINIDOS
    # ─────────────────────────────────────────────
        tab_predef = ttk.Frame(tabs, padding=20)
        tabs.add(tab_predef, text="🎨 Temas Predefinidos")

        ttk.Label(tab_predef, text="Selecciona un tema y haz clic en Aplicar:",
                font=("Segoe UI", 11, "bold")).pack(anchor=W, pady=(0, 15))

    # Canvas con scroll
        canvas_temas = tk.Canvas(tab_predef, highlightthickness=0)
        scroll_temas = ttk.Scrollbar(tab_predef, orient=VERTICAL,
                                    command=canvas_temas.yview)
        frame_temas = ttk.Frame(canvas_temas)

        frame_temas.bind("<Configure>",
                        lambda e: canvas_temas.configure(
                             scrollregion=canvas_temas.bbox("all")))
        canvas_temas.create_window((0, 0), window=frame_temas, anchor=NW)
        canvas_temas.configure(yscrollcommand=scroll_temas.set)
        canvas_temas.pack(side=LEFT, fill=BOTH, expand=True)
        scroll_temas.pack(side=RIGHT, fill=Y)

        # Scroll con rueda
        canvas_temas.bind("<Enter>",
            lambda e: canvas_temas.bind_all("<MouseWheel>",
                lambda ev: canvas_temas.yview_scroll(int(-1*(ev.delta/120)), "units")))
        canvas_temas.bind("<Leave>",
            lambda e: canvas_temas.unbind_all("<MouseWheel>"))

        def seleccionar_tema_predef(nombre_tema):
            tema = GestorTemas.TEMAS_PREDEFINIDOS[nombre_tema]
            GestorTemas.guardar_tema(self.db, tema)
            top.destroy()
            self.solicitar_reinicio()

    # Generar tarjetas de temas
        for nombre, tema_data in GestorTemas.TEMAS_PREDEFINIDOS.items():
            fr_tema = tk.Frame(frame_temas, relief="solid", borderwidth=1,
                            bg=tema_data["color_fondo"])
            fr_tema.pack(fill=X, pady=4, padx=5, ipady=8)

            # ✅ PREVIEW DE COLORES USANDO CANVAS (siempre visible)
            fr_preview = tk.Frame(fr_tema, bg=tema_data["color_fondo"])
            fr_preview.pack(side=LEFT, padx=12)

            canvas_prev = tk.Canvas(fr_preview, width=90, height=24,
                                    highlightthickness=0,
                                    bg=tema_data["color_fondo"])
            canvas_prev.pack()

        # Dibujar 3 rectángulos de colores
            colores_prev = [
                tema_data["color_primario"],
                tema_data["color_secundario"],
                tema_data["color_acento"]
            ]
            for idx_c, color in enumerate(colores_prev):
                x0 = idx_c * 32
                canvas_prev.create_rectangle(x0+2, 2, x0+28, 22,
                                          fill=color, outline="#888888", width=1)

        # Nombre del tema
            tk.Label(fr_tema,
                    text=nombre,
                    font=("Segoe UI", 11, "bold"),
                    bg=tema_data["color_fondo"],
                    fg=tema_data["color_texto"]).pack(side=LEFT, padx=5)

        # Botón Aplicar
            tk.Button(fr_tema,
                    text="✓ Aplicar",
                    bg=tema_data["color_primario"],
                    fg=texto_sobre_color(tema_data["color_primario"]),
                    relief="flat", padx=12, pady=4,
                    font=("Segoe UI", 9, "bold"),
                    cursor="hand2",
                    command=lambda n=nombre: seleccionar_tema_predef(n)
                    ).pack(side=RIGHT, padx=12)

        frame_temas.update_idletasks()
        canvas_temas.configure(scrollregion=canvas_temas.bbox("all"))

    # ─────────────────────────────────────────────
    # PESTAÑA 2: TEMA PERSONALIZADO
    # ─────────────────────────────────────────────
        tab_custom = ttk.Frame(tabs, padding=20)
        tabs.add(tab_custom, text="🖌️ Personalizado")

        ttk.Label(tab_custom, text="Configura tus propios colores:",
                 font=("Segoe UI", 11, "bold")).pack(anchor=W, pady=(0, 15))

        colores_personalizados = {
            "color_primario":   tk.StringVar(value=self.tema_actual["color_primario"]),
            "color_secundario": tk.StringVar(value=self.tema_actual["color_secundario"]),
            "color_acento":     tk.StringVar(value=self.tema_actual["color_acento"]),
            "color_fondo":      tk.StringVar(value=self.tema_actual["color_fondo"]),
            "color_texto":      tk.StringVar(value=self.tema_actual["color_texto"])
        }

        opciones_color = [
            ("🔵 Color Primario (Botones/Títulos)",  "color_primario"),
            ("🔷 Color Secundario (Hover/Detalles)", "color_secundario"),
            ("✨ Color Acento (Resaltado)",           "color_acento"),
            ("🖼️ Color Fondo (Ventanas)",             "color_fondo"),
            ("🔤 Color Texto (Letras)",               "color_texto"),
        ]

    # Guardar referencias a los canvas de preview
        canvas_refs = {}

        for texto, clave in opciones_color:
            fr_fila = ttk.Frame(tab_custom)
            fr_fila.pack(fill=X, pady=6)

            ttk.Label(fr_fila, text=texto, width=38).pack(side=LEFT)

        # Canvas de preview del color actual
            cv = tk.Canvas(fr_fila, width=36, height=24,
                            highlightthickness=1, highlightbackground="#AAAAAA")
            cv.pack(side=LEFT, padx=5)
            rect_id = cv.create_rectangle(2, 2, 34, 22,
                                        fill=colores_personalizados[clave].get(),
                                        outline="")
            canvas_refs[clave] = (cv, rect_id)

        # Mostrar hex actual
            lbl_hex = ttk.Label(fr_fila,
                                text=colores_personalizados[clave].get(),
                                font=("Consolas", 9), foreground="#555555", width=10)
            lbl_hex.pack(side=LEFT, padx=4)

            def elegir_color(c=clave, v=colores_personalizados[clave],
                            lbl=lbl_hex):
                color = colorchooser.askcolor(
                    initialcolor=v.get(),
                    title=f"Seleccionar {c}",
                    parent=top
                )
                if color[1]:
                    v.set(color[1])
                # Actualizar preview
                    cv_ref, rect_ref = canvas_refs[c]
                    cv_ref.itemconfig(rect_ref, fill=color[1])
                    lbl.config(text=color[1])
                top.lift()
                top.focus_force()

            ttk.Button(fr_fila, text="🎨 Elegir",
                    bootstyle="info-outline",
                    command=elegir_color).pack(side=LEFT)

        ttk.Separator(tab_custom).pack(fill=X, pady=15)

    # Tema Bootstrap base
        fr_boot = ttk.Frame(tab_custom)
        fr_boot.pack(fill=X, pady=5)
        ttk.Label(fr_boot, text="🎨 Tema Base (Bootstrap):",
                font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 10))

        temas_bootstrap = ["flatly","cosmo","litera","minty","pulse","sandstone",
                            "united","yeti","darkly","superhero","solar","cyborg","journal"]
        tema_bootstrap_var = tk.StringVar(value=self.tema_actual["tema_bootstrap"])
        ttk.Combobox(fr_boot, textvariable=tema_bootstrap_var,
                    values=temas_bootstrap, state="readonly",
                    width=15).pack(side=LEFT)

        def guardar_tema_personalizado():
            tema_nuevo = {k: v.get() for k, v in colores_personalizados.items()}
            tema_nuevo["tema_bootstrap"] = tema_bootstrap_var.get()
            GestorTemas.guardar_tema(self.db, tema_nuevo)
            top.destroy()
            self.solicitar_reinicio()

        ttk.Button(tab_custom, text="💾  GUARDAR Y REINICIAR",
                bootstyle="success",
                command=guardar_tema_personalizado).pack(fill=X, pady=20, ipady=6)

    def calcular_datos_kardex(self):
        """
        Calcula los datos para el Anexo C.
        CORRECCIÓN: Incluye explícitamente 'ALTA INICIAL' como entrada
        para que las facturas iniciales aparezcan correctamente en el Excel.
        """
        try:
            mes  = int(self.cb_mes_k.get())
            anio = int(self.ent_anio_k.get())
            partida_sel = self.cb_partida_k.get()
        except:
            messagebox.showerror("Error", "Verifica Mes y Año")
            return None

        inicio_mes = datetime(anio, mes, 1)
        ultimo_dia = calendar.monthrange(anio, mes)[1]
        fin_mes    = datetime(anio, mes, ultimo_dia, 23, 59, 59)

        # Seleccionar materiales (filtrado o todos)
        sql    = "SELECT id, partida, material FROM inventario"
        params = []
        if partida_sel and partida_sel != "TODAS":
            sql += " WHERE partida = ?"
            params.append(partida_sel)
        sql += " ORDER BY partida, material"

        materiales        = self.db.consultar(sql, tuple(params))
        datos_procesados  = []

        for idx, mat in enumerate(materiales, 1):
            mat_nom = mat['material']

            # Traer todo el historial de este material
            historial = self.db.consultar(
                "SELECT fecha_hora, tipo, cantidad, factura "
                "FROM historial WHERE material = ? ORDER BY id ASC",
                (mat_nom,)
            )

            ex_ant        = 0
            entradas_mes  = 0
            salidas_dias  = {d: 0 for d in range(1, 32)}
            facturas_mes  = set()

            for h in historial:

                # ── Parsear fecha ─────────────────────────────────────
                try:
                    try:
                        f_obj = datetime.strptime(h['fecha_hora'], "%d/%m/%Y %H:%M")
                    except:
                        f_obj = datetime.strptime(h['fecha_hora'], "%d/%m/%Y")
                except:
                    continue

                cant = h['cantidad']
                tipo = (h['tipo'] or "").strip().upper()

                # ── Clasificar movimiento ─────────────────────────────
                # Cualquier variante de entrada:
                #   ENTRADA, HISTORICO (+), ALTA, ALTA INICIAL
                es_entrada = (
                    "ENTRADA"  in tipo or
                    "(+)"      in tipo or
                    "ALTA"     in tipo   # captura ALTA e ALTA INICIAL
                )

                # ── Antes del mes → va a Existencia Anterior ─────────
                if f_obj < inicio_mes:
                    if es_entrada:
                        ex_ant += cant
                    else:
                        ex_ant -= cant

                # ── Durante el mes → entradas y salidas por día ───────
                elif inicio_mes <= f_obj <= fin_mes:
                    if es_entrada:
                        entradas_mes += cant

                        # Capturar factura (incluye facturas de ALTA INICIAL)
                        fac = (h['factura'] or "").strip()
                        if fac and fac.upper() != "S/F":
                            facturas_mes.add(fac.upper())

                    else:
                        # Salida → acumular en el día correspondiente
                        dia = f_obj.day
                        salidas_dias[dia] += cant

            # Concatenar facturas del mes
            str_facturas = ", ".join(sorted(facturas_mes)) if facturas_mes else ""

            total_sal = sum(salidas_dias.values())
            ex_act    = (ex_ant + entradas_mes) - total_sal

            row = {
                "NP":          idx,
                "UNIDAD":      "PZA",
                "DESC":        mat_nom,
                "FACTURA":     str_facturas,
                "EX_ANT":      ex_ant,
                "RECIBIDOS":   entradas_mes,
                "SALIDAS_DIAS":salidas_dias,
                "TOTAL_SAL":   total_sal,
                "EX_ACT":      ex_act,
                "PARTIDA":     mat['partida']
            }
            datos_procesados.append(row)

        return datos_procesados, mes, anio, partida_sel
    
    def abrir_reporte_movimientos(self):
        """
        Ventana para generar un Excel con todas las ENTRADAS (incluye
        ALTA INICIAL con su factura) y SALIDAS en un rango de fechas.
        Completamente independiente del Anexo C.
        """
        top = tk.Toplevel(self.root)
        top.title("📋 Reporte de Movimientos por Fecha")
        self.centrar_ventana_emergente(top, 520, 420)
        top.grab_set()

        fr = ttk.Frame(top, padding=25)
        fr.pack(fill=BOTH, expand=True)

        ttk.Label(
            fr,
            text="Reporte de Entradas y Salidas",
            font=("Segoe UI", 14, "bold"),
            bootstyle="primary"
        ).pack(pady=(0, 5))

        ttk.Label(
            fr,
            text="Genera un Excel detallado con todos los movimientos del periodo.",
            font=("Segoe UI", 9),
            foreground="gray"
        ).pack(pady=(0, 18))

        # ── Rango de fechas ───────────────────────────────────────────
        fr_fechas = ttk.LabelFrame(
            fr, text=" 📅 Rango de Fechas ",
            padding=15, bootstyle="info")
        fr_fechas.pack(fill=X, pady=(0, 12))

        def crear_fila_fecha(parent, label_txt, dia_def, mes_def, anio_def):
            """Crea una fila con tres Entry para DD / MM / AAAA"""
            fr_fila = ttk.Frame(parent)
            fr_fila.pack(fill=X, pady=4)

            ttk.Label(fr_fila, text=label_txt, width=20).pack(side=LEFT)

            e_dia = ttk.Entry(
                fr_fila, width=3,
                justify=CENTER, font=("Segoe UI", 10, "bold"))
            e_dia.insert(0, dia_def)
            e_dia.pack(side=LEFT)

            ttk.Label(fr_fila, text=" / ").pack(side=LEFT)

            e_mes = ttk.Entry(
                fr_fila, width=3,
                justify=CENTER, font=("Segoe UI", 10, "bold"))
            e_mes.insert(0, mes_def)
            e_mes.pack(side=LEFT)

            ttk.Label(fr_fila, text=" / ").pack(side=LEFT)

            e_anio = ttk.Entry(
                fr_fila, width=5,
                justify=CENTER, font=("Segoe UI", 10, "bold"))
            e_anio.insert(0, anio_def)
            e_anio.pack(side=LEFT)

            # Tab entre campos
            e_dia.bind( "<Tab>", lambda e: (e_mes.focus_set(),  "break")[1])
            e_mes.bind( "<Tab>", lambda e: (e_anio.focus_set(), "break")[1])

            return e_dia, e_mes, e_anio

        hoy = datetime.now()
        e_dia_i, e_mes_i, e_anio_i = crear_fila_fecha(
            fr_fechas, "Desde (DD/MM/AAAA):",
            "01",
            str(hoy.month).zfill(2),
            str(hoy.year)
        )
        e_dia_f, e_mes_f, e_anio_f = crear_fila_fecha(
            fr_fechas, "Hasta (DD/MM/AAAA):",
            str(hoy.day).zfill(2),
            str(hoy.month).zfill(2),
            str(hoy.year)
        )

        # ── Filtro opcional de partida ─────────────────────────────────
        fr_opt = ttk.LabelFrame(
            fr, text=" Filtro Opcional ",
            padding=12, bootstyle="secondary")
        fr_opt.pack(fill=X, pady=(0, 18))

        fr_part_row = ttk.Frame(fr_opt)
        fr_part_row.pack(fill=X)

        ttk.Label(fr_part_row, text="Partida:").pack(side=LEFT)

        rows_p  = self.db.consultar(
            "SELECT valor FROM catalogos WHERE tipo='PARTIDA' ORDER BY valor ASC")
        lista_p = ["TODAS"] + [r['valor'] for r in rows_p]

        cb_part = ttk.Combobox(
            fr_part_row, values=lista_p,
            state="readonly", width=22)
        cb_part.current(0)
        cb_part.pack(side=LEFT, padx=10)

        # ── Lógica de generación ───────────────────────────────────────
        def generar_excel():
            # Validar fechas
            try:
                fecha_ini = datetime(
                    int(e_anio_i.get()), int(e_mes_i.get()), int(e_dia_i.get()))
                fecha_fin = datetime(
                    int(e_anio_f.get()), int(e_mes_f.get()), int(e_dia_f.get()),
                    23, 59, 59)
            except ValueError as err:
                messagebox.showerror(
                    "Fecha inválida",
                    f"Revisa los valores ingresados.\n\n{err}",
                    parent=top)
                top.lift()
                return

            if fecha_ini > fecha_fin:
                messagebox.showerror(
                    "Error de fechas",
                    "La fecha inicial no puede ser mayor que la final.",
                    parent=top)
                top.lift()
                return

            partida_filtro = cb_part.get()

            # Traer TODO el historial
            filas = self.db.consultar(
                "SELECT fecha_hora, tipo, partida, material, "
                "cantidad, factura, destino, responsable, entrego "
                "FROM historial ORDER BY id ASC"
            )

            entradas = []
            salidas  = []

            for f in filas:
                # Parsear fecha
                try:
                    try:
                        f_obj = datetime.strptime(f['fecha_hora'], "%d/%m/%Y %H:%M")
                    except:
                        f_obj = datetime.strptime(f['fecha_hora'], "%d/%m/%Y")
                except:
                    continue

                # Filtro de rango
                if not (fecha_ini <= f_obj <= fecha_fin):
                    continue

                # Filtro de partida
                if partida_filtro != "TODAS" and f['partida'] != partida_filtro:
                    continue

                tipo = (f['tipo'] or "").strip().upper()
                cant = f['cantidad']

                # Formatear entero si aplica
                cant_fmt = (
                    int(cant)
                    if isinstance(cant, float) and cant == int(cant)
                    else cant
                )

                es_entrada = (
                    "ENTRADA" in tipo or
                    "(+)"     in tipo or
                    "ALTA"    in tipo
                )

                if es_entrada:
                    entradas.append({
                        "FECHA":       f['fecha_hora'],
                        "TIPO":        f['tipo'],
                        "PARTIDA":     f['partida'] or "",
                        "MATERIAL":    f['material'],
                        "CANTIDAD":    cant_fmt,
                        "FACTURA":     (f['factura'] or "S/F").upper(),
                        "RESPONSABLE": f['responsable'] or ""
                    })
                else:
                    salidas.append({
                        "FECHA":       f['fecha_hora'],
                        "TIPO":        f['tipo'],
                        "PARTIDA":     f['partida'] or "",
                        "MATERIAL":    f['material'],
                        "CANTIDAD":    cant_fmt,
                        "DESTINO":     f['destino'] or "",
                        "SOLICITA":    f['responsable'] or "",
                        "ENTREGO":     f['entrego'] or "",
                        "FOLIO":       f['factura'] or ""
                    })

            if not entradas and not salidas:
                messagebox.showwarning(
                    "Sin datos",
                    "No se encontraron movimientos en el periodo seleccionado.",
                    parent=top)
                top.lift()
                return

            # Pedir ruta de guardado
            ruta = filedialog.asksaveasfilename(
                parent=top,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=(
                    f"Movimientos_"
                    f"{fecha_ini.strftime('%d%m%Y')}_"
                    f"{fecha_fin.strftime('%d%m%Y')}.xlsx"
                )
            )
            if not ruta:
                top.lift()
                return

            # ── Construir Excel ───────────────────────────────────────
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, Alignment, PatternFill, Border,
                Side as ExcelSide)

            wb = Workbook()
            ws = wb.active
            ws.title = "Movimientos"

            # Estilos
            thin   = ExcelSide(border_style="thin",   color="CCCCCC")
            medium = ExcelSide(border_style="medium",  color="888888")
            borde  = Border(top=thin, left=thin, right=thin, bottom=thin)
            borde_medio = Border(
                top=medium, left=medium, right=medium, bottom=medium)
            centro = Alignment(horizontal="center", vertical="center")
            izq    = Alignment(
                horizontal="left", vertical="center", wrap_text=True)

            HEX_AZUL  = "1F4E79"
            HEX_VERDE = "1B5E20"
            HEX_ROJO  = "7B1010"
            HEX_GRIS  = "F5F5F5"

            fill_azul  = PatternFill(fill_type="solid", fgColor=HEX_AZUL)
            fill_verde = PatternFill(fill_type="solid", fgColor=HEX_VERDE)
            fill_rojo  = PatternFill(fill_type="solid", fgColor=HEX_ROJO)
            fill_v_sub = PatternFill(fill_type="solid", fgColor="2E7D32")
            fill_r_sub = PatternFill(fill_type="solid", fgColor="C62828")
            fill_ent   = PatternFill(fill_type="solid", fgColor="F1F8E9")
            fill_sal   = PatternFill(fill_type="solid", fgColor="FFF3F3")
            fill_tot_v = PatternFill(fill_type="solid", fgColor="E8F5E9")
            fill_tot_r = PatternFill(fill_type="solid", fgColor="FFEBEE")

            def celda(ws, row, col, valor, fuente=None, alin=None,
                      relleno=None, brd=None):
                c = ws.cell(row=row, column=col, value=valor)
                if fuente:  c.font      = fuente
                if alin:    c.alignment = alin
                if relleno: c.fill      = relleno
                if brd:     c.border    = brd
                return c

            # ── Encabezado institucional ──────────────────────────────
            h1 = self.db.get_config("HEADER_L1") or "INSTITUCIÓN"
            h2 = self.db.get_config("HEADER_L2") or "SUBDIRECCIÓN"
            h4 = self.db.get_config("HEADER_L4") or "DEPARTAMENTO"

            str_ini = fecha_ini.strftime("%d/%m/%Y")
            str_fin = fecha_fin.strftime("%d/%m/%Y")

            for fila_h, texto, fs in [
                (1, h1, 13), (2, h2, 10), (3, h4, 10)
            ]:
                ws.merge_cells(f"A{fila_h}:I{fila_h}")
                celda(ws, fila_h, 1, texto,
                      fuente=Font(bold=True, size=fs, color=HEX_AZUL),
                      alin=centro)

            ws.row_dimensions[1].height = 22
            ws.row_dimensions[4].height = 6   # separador

            # Título principal
            ws.merge_cells("A5:I5")
            celda(ws, 5, 1,
                  f"REPORTE DE MOVIMIENTOS  —  {str_ini}  al  {str_fin}",
                  fuente=Font(bold=True, size=13, color="FFFFFF"),
                  alin=centro, relleno=fill_azul)
            ws.row_dimensions[5].height = 24

            if partida_filtro != "TODAS":
                ws.merge_cells("A6:I6")
                celda(ws, 6, 1,
                      f"Filtrado por Partida: {partida_filtro}",
                      fuente=Font(italic=True, size=10, color="555555"),
                      alin=centro)

            fila_act = 8

            # ── SECCIÓN ENTRADAS ──────────────────────────────────────
            ws.merge_cells(f"A{fila_act}:G{fila_act}")
            celda(ws, fila_act, 1,
                  f"ENTRADAS / ALTAS DE MATERIAL  ({len(entradas)} registros)",
                  fuente=Font(bold=True, size=11, color="FFFFFF"),
                  alin=centro, relleno=fill_verde)
            ws.row_dimensions[fila_act].height = 20
            fila_act += 1

            # Cabecera entradas
            hdrs_ent = [
                "FECHA", "TIPO DE MOVIMIENTO", "PARTIDA",
                "DESCRIPCIÓN DEL MATERIAL", "CANTIDAD",
                "FACTURA / REFERENCIA", "RESPONSABLE"
            ]
            for col, h in enumerate(hdrs_ent, 1):
                celda(ws, fila_act, col, h,
                      fuente=Font(bold=True, size=9, color="FFFFFF"),
                      alin=centro, relleno=fill_v_sub, brd=borde)
            ws.row_dimensions[fila_act].height = 18
            fila_act += 1

            total_ent = 0
            for i, e in enumerate(entradas):
                relleno_e = fill_ent if i % 2 == 0 else None
                datos = [
                    e["FECHA"], e["TIPO"], e["PARTIDA"],
                    e["MATERIAL"], e["CANTIDAD"],
                    e["FACTURA"], e["RESPONSABLE"]
                ]
                for col, val in enumerate(datos, 1):
                    celda(ws, fila_act, col, val,
                          fuente=Font(size=9),
                          alin=izq if col == 4 else centro,
                          relleno=relleno_e, brd=borde)
                ws.row_dimensions[fila_act].height = 16
                total_ent += (
                    e["CANTIDAD"]
                    if isinstance(e["CANTIDAD"], (int, float)) else 0
                )
                fila_act += 1

            # Fila total entradas
            ws.merge_cells(f"A{fila_act}:D{fila_act}")
            celda(ws, fila_act, 1, "TOTAL DE UNIDADES INGRESADAS:",
                  fuente=Font(bold=True, size=10, color=HEX_VERDE),
                  alin=Alignment(horizontal="right", vertical="center"),
                  relleno=fill_tot_v, brd=borde)
            celda(ws, fila_act, 5,
                  int(total_ent) if total_ent == int(total_ent) else total_ent,
                  fuente=Font(bold=True, size=11, color=HEX_VERDE),
                  alin=centro, relleno=fill_tot_v, brd=borde)
            ws.row_dimensions[fila_act].height = 18
            fila_act += 2   # separador entre secciones

            # ── SECCIÓN SALIDAS ───────────────────────────────────────
            ws.merge_cells(f"A{fila_act}:I{fila_act}")
            celda(ws, fila_act, 1,
                  f"SALIDAS DE MATERIAL  ({len(salidas)} registros)",
                  fuente=Font(bold=True, size=11, color="FFFFFF"),
                  alin=centro, relleno=fill_rojo)
            ws.row_dimensions[fila_act].height = 20
            fila_act += 1

            # Cabecera salidas
            hdrs_sal = [
                "FECHA", "TIPO DE MOVIMIENTO", "PARTIDA",
                "DESCRIPCIÓN DEL MATERIAL", "CANTIDAD",
                "ÁREA / DESTINO", "SOLICITA", "ENTREGÓ", "FOLIO / VALE"
            ]
            for col, h in enumerate(hdrs_sal, 1):
                celda(ws, fila_act, col, h,
                      fuente=Font(bold=True, size=9, color="FFFFFF"),
                      alin=centro, relleno=fill_r_sub, brd=borde)
            ws.row_dimensions[fila_act].height = 18
            fila_act += 1

            total_sal = 0
            for i, s in enumerate(salidas):
                relleno_s = fill_sal if i % 2 == 0 else None
                datos = [
                    s["FECHA"], s["TIPO"], s["PARTIDA"],
                    s["MATERIAL"], s["CANTIDAD"],
                    s["DESTINO"], s["SOLICITA"],
                    s["ENTREGO"], s["FOLIO"]
                ]
                for col, val in enumerate(datos, 1):
                    celda(ws, fila_act, col, val,
                          fuente=Font(size=9),
                          alin=izq if col == 4 else centro,
                          relleno=relleno_s, brd=borde)
                ws.row_dimensions[fila_act].height = 16
                total_sal += (
                    s["CANTIDAD"]
                    if isinstance(s["CANTIDAD"], (int, float)) else 0
                )
                fila_act += 1

            # Fila total salidas
            ws.merge_cells(f"A{fila_act}:D{fila_act}")
            celda(ws, fila_act, 1, "TOTAL DE UNIDADES SALIDAS:",
                  fuente=Font(bold=True, size=10, color=HEX_ROJO),
                  alin=Alignment(horizontal="right", vertical="center"),
                  relleno=fill_tot_r, brd=borde)
            celda(ws, fila_act, 5,
                  int(total_sal) if total_sal == int(total_sal) else total_sal,
                  fuente=Font(bold=True, size=11, color=HEX_ROJO),
                  alin=centro, relleno=fill_tot_r, brd=borde)
            ws.row_dimensions[fila_act].height = 18

            # ── Anchos de columna ─────────────────────────────────────
            anchos = {
                "A": 18, "B": 20, "C": 10,
                "D": 42, "E": 11, "F": 24,
                "G": 22, "H": 22, "I": 16
            }
            for col, w in anchos.items():
                ws.column_dimensions[col].width = w

            # ── Guardar y abrir ───────────────────────────────────────
            try:
                wb.save(ruta)
                messagebox.showinfo(
                    "✅ Exportado",
                    f"Reporte generado correctamente.\n\n"
                    f"Entradas: {len(entradas)} registros\n"
                    f"Salidas:  {len(salidas)} registros",
                    parent=top)
                top.lift()
                os.startfile(ruta)
            except PermissionError:
                messagebox.showwarning(
                    "Archivo abierto",
                    "El archivo está abierto en Excel.\n"
                    "Ciérralo e intenta de nuevo.",
                    parent=top)
                top.lift()
            except Exception as err:
                messagebox.showerror("Error al guardar", f"{err}", parent=top)
                top.lift()

        # ── Botones ───────────────────────────────────────────────────
        ttk.Button(
            fr,
            text="💾  Generar y Exportar Excel",
            bootstyle="success",
            command=generar_excel
        ).pack(fill=X, ipady=7, pady=(0, 6))

        ttk.Button(
            fr,
            text="Cerrar",
            bootstyle="secondary-outline",
            command=top.destroy
        ).pack(fill=X)
    
    def generar_vista_anexo_c(self):
        res = self.calcular_datos_kardex()
        if not res: return
        datos, _, _, _ = res
        
        # Limpiar y llenar
        for i in self.tree_kardex.get_children(): self.tree_kardex.delete(i)
        
        for d in datos:
            # Lista de valores para las columnas 1-31
            vals_dias = [d['SALIDAS_DIAS'][dia] if d['SALIDAS_DIAS'][dia] > 0 else "" for dia in range(1, 32)]
            
            valores = [d['NP'], d['UNIDAD'], d['DESC'], d['FACTURA'], d['EX_ANT'], d['RECIBIDOS']] + vals_dias + [d['TOTAL_SAL'], d['EX_ACT']]
            self.tree_kardex.insert("", END, values=valores)
    
    # ------------------------------------------------------------------
  
    # ------------------------------------------------------------------
  


    def exportar_excel_anexo_c(self):
        # Importación segura
        from openpyxl.styles import Font, Alignment, Border, Side as ExcelSide
        
        res = self.calcular_datos_kardex()
        if not res: return
        datos, mes, anio, partida_sel = res
        
        ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not ruta: return

        wb = Workbook()
        ws = wb.active
        
        nombres_meses = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", 
                         "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        nombre_mes = nombres_meses[mes] 
        
        ws.title = f"ANEXO C {nombre_mes}-{anio}"
        
        # ESTILOS
        thin = ExcelSide(border_style="thin", color="000000") 
        borde_todo = Border(top=thin, left=thin, right=thin, bottom=thin)
        centrado = Alignment(horizontal='center', vertical='center', wrap_text=True)
        negrita = Font(bold=True, size=10, name='Arial')
        
        # --- ENCABEZADOS DINÁMICOS (DESDE DB) ---
        ws.merge_cells('A1:AK1'); ws['A1'] = "ANEXO C"
        
        # Recuperar configuración o usar default si está vacío
        h1 = self.db.get_config("HEADER_L1") or "INSTITUCIÓN"
        h2 = self.db.get_config("HEADER_L2") or "SUBDIRECCIÓN"
        h3 = self.db.get_config("HEADER_L3") or "DIRECCIÓN GENERAL"
        h4 = self.db.get_config("HEADER_L4") or "DEPARTAMENTO"
        
        ws.merge_cells('A2:AK2'); ws['A2'] = h1
        ws.merge_cells('A3:AK3'); ws['A3'] = h2
        ws.merge_cells('A4:AK4'); ws['A4'] = h3
        ws.merge_cells('A5:AK5'); ws['A5'] = h4
        
        # LÓGICA DE NOMBRE LARGO
        txt_partida = "TODAS LAS PARTIDAS"
        if partida_sel and partida_sel != "TODAS":
            res_desc = self.db.consultar("SELECT descripcion FROM partidas_desc WHERE codigo=?", (partida_sel,))
            nombre_largo = res_desc[0]['descripcion'] if res_desc else ""
            txt_partida = f"PARTIDA {partida_sel} {nombre_largo}"
            
        ws.merge_cells('A7:AK7')
        ws['A7'] = f"CONTROL DE LA {txt_partida} CORRESPONDIENTE AL MES DE {nombre_mes} {anio}"
        
        for row in ws.iter_rows(min_row=1, max_row=7):
            for cell in row:
                cell.alignment = centrado
                cell.font = negrita

        # --- CABECERA DE TABLA ---
        headers_fijos = [("A", "N.P."), ("B", "UNIDAD"), ("C", "DESCRIPCIÓN"), 
                         ("D", "FACTURA"), ("E", "EXIST.\nANT."), ("F", "EFECTOS\nRECIBIDOS")]
        
        for col, texto in headers_fijos:
            ws.merge_cells(f'{col}9:{col}10')
            cell = ws[f'{col}9']
            cell.value = texto
            cell.alignment = centrado
            cell.font = Font(bold=True, size=8)
            cell.border = borde_todo
            ws[f'{col}10'].border = borde_todo 

        ws.merge_cells('G9:AK9')
        ws['G9'] = "SALIDAS (DÍAS DEL MES)"
        ws['G9'].alignment = centrado; ws['G9'].font = Font(bold=True, size=8); ws['G9'].border = borde_todo
        
        col_idx = 7
        for dia in range(1, 32):
            cell = ws.cell(row=10, column=col_idx, value=dia)
            cell.alignment = centrado; cell.font = Font(size=8); cell.border = borde_todo
            ws.column_dimensions[chr(64+col_idx) if col_idx <= 26 else f"A{chr(64+col_idx-26)}"].width = 3.5
            col_idx += 1

        ws.merge_cells('AL9:AL10'); ws['AL9'] = "TOTAL\nSALIDA"
        ws.merge_cells('AM9:AM10'); ws['AM9'] = "EXIST.\nACT."
        for col in ['AL', 'AM']:
            cell = ws[f'{col}9']
            cell.alignment = centrado; cell.font = Font(bold=True, size=8); cell.border = borde_todo
            ws[f'{col}10'].border = borde_todo

        # --- DATOS ---
        fila_act = 11
        for d in datos:
            ws.cell(row=fila_act, column=1, value=d['NP']).border = borde_todo
            ws.cell(row=fila_act, column=2, value=d['UNIDAD']).border = borde_todo
            ws.cell(row=fila_act, column=3, value=d['DESC']).border = borde_todo
            ws.cell(row=fila_act, column=4, value=d['FACTURA']).border = borde_todo
            ws.cell(row=fila_act, column=5, value=d['EX_ANT']).border = borde_todo
            ws.cell(row=fila_act, column=6, value=d['RECIBIDOS']).border = borde_todo
            
            col_d = 7
            for dia in range(1, 32):
                val = d['SALIDAS_DIAS'][dia]
                ws.cell(row=fila_act, column=col_d, value=val if val > 0 else "").border = borde_todo
                col_d += 1
            
            ws.cell(row=fila_act, column=38, value=d['TOTAL_SAL']).border = borde_todo
            ws.cell(row=fila_act, column=39, value=d['EX_ACT']).border = borde_todo
            fila_act += 1
            
        ws.column_dimensions['C'].width = 40 
        ws.column_dimensions['D'].width = 15 

        # --- PIE DE PÁGINA (FIRMAS) ---
        fila_firmas = fila_act + 4
        firmas = [
            ("ELABORÓ", "B", "E"), 
            ("SUPERVISÓ", "M", "P"), 
            ("Vo. Bo.", "S", "V"), 
            ("CONSTAME", "AA", "AE")
        ]
        
        top_line = ExcelSide(border_style="medium", color="000000")
        for titulo, col_inicio, col_fin in firmas:
            ws.merge_cells(f'{col_inicio}{fila_firmas}:{col_fin}{fila_firmas}')
            cell = ws[f'{col_inicio}{fila_firmas}']
            cell.value = titulo
            cell.alignment = centrado; cell.font = negrita
            for c_idx in range(ws[f'{col_inicio}1'].column, ws[f'{col_fin}1'].column + 1):
                ws.cell(row=fila_firmas, column=c_idx).border = Border(top=top_line)

        try:
            wb.save(ruta)
            messagebox.showinfo("Éxito", "Reporte Anexo C generado correctamente.")
            os.startfile(ruta)
        except PermissionError:
            messagebox.showwarning("Archivo Abierto", 
                                   f"No se pudo guardar el archivo.\n\n"
                                   f"El archivo '{os.path.basename(ruta)}' está abierto.\n"
                                   "Por favor, CIÉRRALO y vuelve a intentar.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el Excel: {e}")

    def centrar_ventana_emergente(self, ventana, ancho, alto):
        """Centra una ventana Toplevel en la pantalla y define su tamaño mínimo"""
        ventana.update_idletasks() # Necesario para cálculos precisos
        
        # Obtener dimensiones de la pantalla
        ws = ventana.winfo_screenwidth()
        hs = ventana.winfo_screenheight()
        
        # Calcular posición X e Y
        x = (ws // 2) - (ancho // 2)
        y = (hs // 2) - (alto // 2)
        
        # Aplicar geometría y tamaño mínimo
        ventana.geometry(f"{ancho}x{alto}+{int(x)}+{int(y)}")
        ventana.minsize(ancho, alto) # Evita que se haga demasiado pequeña y corte botones
        
        # Asegurar que se pueda maximizar (True, True es el defecto, pero confirmamos)
        ventana.resizable(True, True) 
        
        # Ponerla al frente
        ventana.lift()
        ventana.focus_force()


    def solicitar_reinicio(self):
        """Pregunta y ejecuta el reinicio automático (CORREGIDO PARA RUTAS CON ESPACIOS)"""
        respuesta = messagebox.askyesno(
            "Configuración Guardada", 
            "Para que los cambios visuales se apliquen correctamente, es necesario reiniciar el sistema.\n\n"
            "¿Deseas reiniciar AHORA?",
            icon='warning'
        )
        if respuesta:
            # 1. Cerrar ventana actual
            self.root.destroy()
            
            # 2. Preparar el reinicio manejando ESPACIOS en la ruta
            python = sys.executable
            script = sys.argv[0]
            
            # Si el nombre del archivo o la carpeta tiene espacios, le ponemos comillas
            if " " in script:
                script = f'"{script}"'
            
            # 3. Reiniciar
            # os.execl reemplaza el proceso actual por uno nuevo
            try:
                os.execl(python, python, script, *sys.argv[1:])
            except Exception as e:
                # Si falla el reinicio automático, avisamos
                print(f"Error al reiniciar: {e}")
                sys.exit()
        
    # --- EN LA CLASE SistemaInventario (NUEVA FUNCIÓN) ---
    def abrir_gestion_usuarios(self):
        if self.usuario.get('rol') != 'ADMIN':
            messagebox.showerror("Acceso Denegado", "Se requieren permisos de administrador.")
            return

        top = tk.Toplevel(self.root)
        top.title("Gestión de Usuarios y Permisos")
        top.state('zoomed')
        top.minsize(900, 600)
        top.grab_set()
        top.focus_force()

        # ── Variables de control ──────────────────────────────────────
        var_id        = tk.StringVar()
        var_user      = tk.StringVar()
        var_pass      = tk.StringVar()
        var_nombre    = tk.StringVar()
        var_email     = tk.StringVar()
        var_rol       = tk.StringVar(value="OPERADOR")
        var_foto_path = tk.StringVar()

        var_p_crear = tk.IntVar(); var_p_ent  = tk.IntVar(); var_p_sal  = tk.IntVar()
        var_p_edit  = tk.IntVar(); var_p_del  = tk.IntVar()
        var_p_cat   = tk.IntVar(); var_p_hist = tk.IntVar(); var_p_conf = tk.IntVar()

        self.usuario_seleccionado_id = None
        self.password_actual_hash    = ""

        import tkinter.ttk as original_ttk

        main_container = ttk.Frame(top, padding=10)
        main_container.pack(fill=BOTH, expand=True)

        paned = original_ttk.PanedWindow(main_container, orient=HORIZONTAL)
        paned.pack(fill=BOTH, expand=True)

        # ══════════════════════════════════════════════════════════════
        # PANEL IZQUIERDO: LISTA DE USUARIOS
        # ══════════════════════════════════════════════════════════════
        fr_lista = ttk.Labelframe(paned, text=" Usuarios Registrados ", padding=10, bootstyle="info")
        paned.add(fr_lista, weight=1)

        cols_u = ("ID", "USUARIO", "ROL", "NOMBRE")
        tree_users = ttk.Treeview(fr_lista, columns=cols_u, show="headings", selectmode="browse")

        tree_users.heading("ID",      text="ID");      tree_users.column("ID",      width=40,  anchor=CENTER)
        tree_users.heading("USUARIO", text="Usuario"); tree_users.column("USUARIO", width=100)
        tree_users.heading("ROL",     text="Rol");     tree_users.column("ROL",     width=80,  anchor=CENTER)
        tree_users.heading("NOMBRE",  text="Nombre");  tree_users.column("NOMBRE",  width=150)

        sc_u = ttk.Scrollbar(fr_lista, orient=VERTICAL, command=tree_users.yview)
        tree_users.configure(yscrollcommand=sc_u.set)
        tree_users.pack(side=LEFT, fill=BOTH, expand=True)
        sc_u.pack(side=RIGHT, fill=Y)

        ttk.Label(
            fr_lista,
            text="💡 Selecciona un usuario para editarlo",
            font=("Segoe UI", 8), bootstyle="secondary"
        ).pack(side=BOTTOM, fill=X)

        # ══════════════════════════════════════════════════════════════
        # PANEL DERECHO: FORMULARIO CON SCROLL
        # ══════════════════════════════════════════════════════════════
        fr_form = ttk.Labelframe(
            paned, text=" Ficha de Usuario y Permisos ",
            padding=15, bootstyle="primary"
        )
        paned.add(fr_form, weight=3)

        canvas_form    = tk.Canvas(fr_form, highlightthickness=0)
        scrollbar_form = ttk.Scrollbar(fr_form, orient=VERTICAL, command=canvas_form.yview)
        scrollable_frame = ttk.Frame(canvas_form)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas_form.configure(scrollregion=canvas_form.bbox("all"))
        )
        canvas_form.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas_form.configure(yscrollcommand=scrollbar_form.set)

        def _on_mousewheel(event):
            canvas_form.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_scroll(event):
            canvas_form.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_scroll(event):
            canvas_form.unbind_all("<MouseWheel>")

        canvas_form.bind("<Enter>", _bind_scroll)
        canvas_form.bind("<Leave>", _unbind_scroll)

        scrollbar_form.pack(side=RIGHT, fill=Y)
        canvas_form.pack(side=LEFT, fill=BOTH, expand=True)

        tab_order = []

        # ── Sección Foto ──────────────────────────────────────────────
        fr_foto = ttk.Frame(scrollable_frame)
        fr_foto.pack(fill=X, pady=(0, 10))

        self.lbl_preview = ttk.Label(
            fr_foto, text="👤",
            font=("Segoe UI Emoji", 40), anchor=CENTER
        )
        self.lbl_preview.pack(side=LEFT, padx=10)

        def seleccionar_foto():
            ruta = filedialog.askopenfilename(
                parent=top,
                filetypes=[("Imágenes", "*.png;*.jpg;*.jpeg")]
            )
            if ruta:
                var_foto_path.set(ruta)
                try:
                    img = Image.open(ruta).resize((70, 70), Image.LANCZOS)
                    self.tk_foto_temp = ImageTk.PhotoImage(img, master=top)
                    self.lbl_preview.configure(image=self.tk_foto_temp, text="")
                except:
                    pass
            top.lift()
            top.focus_force()

        ttk.Button(
            fr_foto, text="📂 Cambiar Foto",
            bootstyle="secondary-outline",
            command=seleccionar_foto
        ).pack(side=LEFT, padx=10)

        # ── Campos de texto ───────────────────────────────────────────
        fr_campos = ttk.Frame(scrollable_frame)
        fr_campos.pack(fill=X)
        fr_campos.columnconfigure(1, weight=1)

        # Fila 0: Usuario
        ttk.Label(fr_campos, text="Usuario (Login):").grid(row=0, column=0, sticky=W, pady=5)
        e_user = ttk.Entry(fr_campos, textvariable=var_user, font=("Segoe UI", 10, "bold"))
        e_user.grid(row=0, column=1, sticky=EW, pady=5, padx=5)
        tab_order.append(e_user)

        # Fila 1: Contraseña
        ttk.Label(fr_campos, text="Contraseña:").grid(row=1, column=0, sticky=W, pady=5)
        fr_pass = ttk.Frame(fr_campos)
        fr_pass.grid(row=1, column=1, sticky=EW, pady=5, padx=5)

        e_pass = ttk.Entry(fr_pass, textvariable=var_pass, show="*")
        e_pass.pack(side=LEFT, fill=X, expand=True)
        tab_order.append(e_pass)

        ver_pass = tk.BooleanVar()
        ttk.Checkbutton(
            fr_pass, text="👁️", variable=ver_pass, bootstyle="toolbutton",
            command=lambda: e_pass.config(show="" if ver_pass.get() else "*")
        ).pack(side=LEFT)

        ttk.Label(
            fr_campos,
            text="📋 Mínimo 8 caracteres · Una MAYÚSCULA · una minúscula · un número · solo alfanumérico",
            font=("Segoe UI", 7, "italic"), foreground="#888888"
        ).grid(row=2, column=1, sticky=W, padx=5)

        # Fila 3: Nombre
        ttk.Label(fr_campos, text="Nombre Completo:").grid(row=3, column=0, sticky=W, pady=5)
        e_nombre = ttk.Entry(fr_campos, textvariable=var_nombre)
        e_nombre.grid(row=3, column=1, sticky=EW, pady=5, padx=5)
        tab_order.append(e_nombre)

        # Fila 4: Email
        ttk.Label(fr_campos, text="Correo Electrónico:").grid(row=4, column=0, sticky=W, pady=5)
        e_email = ttk.Entry(fr_campos, textvariable=var_email)
        e_email.grid(row=4, column=1, sticky=EW, pady=5, padx=5)
        tab_order.append(e_email)

        # Fila 5: Rol
        ttk.Label(fr_campos, text="Rol (Etiqueta):").grid(row=5, column=0, sticky=W, pady=5)
        cbox_rol = ttk.Combobox(
            fr_campos, textvariable=var_rol,
            values=["OPERADOR", "ADMIN", "SOLO LECTURA"],
            state="readonly"
        )
        cbox_rol.grid(row=5, column=1, sticky=EW, pady=5, padx=5)
        tab_order.append(cbox_rol)

        # Tab order manual
        def focus_next(event):
            try:
                idx = tab_order.index(event.widget)
                tab_order[(idx + 1) % len(tab_order)].focus_set()
            except ValueError:
                pass
            return "break"

        def focus_prev(event):
            try:
                idx = tab_order.index(event.widget)
                tab_order[(idx - 1) % len(tab_order)].focus_set()
            except ValueError:
                pass
            return "break"

        for widget in tab_order:
            widget.bind("<Tab>",       focus_next)
            widget.bind("<Shift-Tab>", focus_prev)

        # ── Sección Permisos ──────────────────────────────────────────
        fr_permisos = ttk.LabelFrame(
            scrollable_frame, text=" Configuración de Accesos ",
            padding=10, bootstyle="warning"
        )
        fr_permisos.pack(fill=X, pady=15)

        col1 = ttk.Frame(fr_permisos); col1.pack(side=LEFT, fill=Y, expand=True, padx=5)
        col2 = ttk.Frame(fr_permisos); col2.pack(side=LEFT, fill=Y, expand=True, padx=5)

        ttk.Label(col1, text="Operativos",    font=("Segoe UI", 8, "bold"), foreground="gray").pack(anchor=W)
        ttk.Checkbutton(col1, text="Crear Materiales",    variable=var_p_crear, bootstyle="round-toggle").pack(anchor=W, pady=2)
        ttk.Checkbutton(col1, text="Registrar ENTRADAS",  variable=var_p_ent,   bootstyle="round-toggle").pack(anchor=W, pady=2)
        ttk.Checkbutton(col1, text="Registrar SALIDAS",   variable=var_p_sal,   bootstyle="round-toggle").pack(anchor=W, pady=2)
        ttk.Checkbutton(col1, text="Editar Maestros",     variable=var_p_edit,  bootstyle="round-toggle").pack(anchor=W, pady=2)

        ttk.Label(col2, text="Módulos Admin", font=("Segoe UI", 8, "bold"), foreground="gray").pack(anchor=W)
        ttk.Checkbutton(col2, text="Gestión Catálogos",   variable=var_p_cat,   bootstyle="round-toggle").pack(anchor=W, pady=2)
        ttk.Checkbutton(col2, text="Modificar Histórico", variable=var_p_hist,  bootstyle="round-toggle").pack(anchor=W, pady=2)
        ttk.Checkbutton(col2, text="Ajustes (Logos)",     variable=var_p_conf,  bootstyle="round-toggle").pack(anchor=W, pady=2)

        def al_cambiar_rol(event):
            rol = var_rol.get()
            if rol == "ADMIN":
                for v in [var_p_crear, var_p_ent, var_p_sal, var_p_edit,
                          var_p_del, var_p_cat, var_p_hist, var_p_conf]:
                    v.set(1)
            elif rol == "OPERADOR":
                var_p_crear.set(1); var_p_ent.set(1); var_p_sal.set(1)
                var_p_edit.set(0);  var_p_del.set(0)
                var_p_cat.set(0);   var_p_hist.set(0); var_p_conf.set(0)
            elif rol == "SOLO LECTURA":
                for v in [var_p_crear, var_p_ent, var_p_sal, var_p_edit,
                          var_p_del, var_p_cat, var_p_hist, var_p_conf]:
                    v.set(0)

        cbox_rol.bind("<<ComboboxSelected>>", al_cambiar_rol)

        # ══════════════════════════════════════════════════════════════
        # FUNCIONES CRUD
        # ══════════════════════════════════════════════════════════════
        def limpiar_form():
            self.usuario_seleccionado_id = None
            self.password_actual_hash    = ""
            var_user.set(""); var_pass.set("")
            var_nombre.set(""); var_email.set("")
            var_foto_path.set("")
            var_rol.set("OPERADOR")
            self.lbl_preview.configure(image="", text="👤")
            tree_users.selection_remove(tree_users.selection())
            al_cambiar_rol(None)
            btn_guardar.configure(text="💾 CREAR NUEVO", bootstyle="success")
            e_user.focus_set()

        def cargar_usuarios_en_lista():
            for item in tree_users.get_children():
                tree_users.delete(item)
            filas = self.db.consultar(
                "SELECT id, usuario, rol, nombre_completo FROM usuarios ORDER BY usuario ASC"
            )
            for f in filas:
                tree_users.insert("", END, values=(f['id'], f['usuario'], f['rol'], f['nombre_completo']))

        def llenar_formulario(event):
            sel = tree_users.selection()
            if not sel:
                return
            id_u = tree_users.item(sel[0])['values'][0]
            self.usuario_seleccionado_id = id_u

            datos = self.db.consultar("SELECT * FROM usuarios WHERE id=?", (id_u,))
            if datos:
                u = dict(datos[0])
                var_user.set(u['usuario'])
                self.password_actual_hash = u['password']
                var_pass.set("")
                var_nombre.set(u.get('nombre_completo', ''))
                var_email.set(u.get('email', ''))
                var_rol.set(u['rol'])
                var_foto_path.set(u.get('foto_path', ''))

                import json
                try:
                    p = json.loads(u.get('permisos', '{}'))
                    var_p_crear.set(p.get('crear',     0))
                    var_p_ent.set(  p.get('entrada',   0))
                    var_p_sal.set(  p.get('salida',    0))
                    var_p_edit.set( p.get('editar',    0))
                    var_p_del.set(  p.get('eliminar',  0))
                    var_p_cat.set(  p.get('catalogos', 0))
                    var_p_hist.set( p.get('historico', 0))
                    var_p_conf.set( p.get('ajustes',   0))
                except:
                    al_cambiar_rol(None)

                foto = u.get('foto_path', '')
                if foto and os.path.exists(foto):
                    try:
                        img = Image.open(foto).resize((70, 70), Image.LANCZOS)
                        self.tk_foto_temp = ImageTk.PhotoImage(img, master=top)
                        self.lbl_preview.configure(image=self.tk_foto_temp, text="")
                    except:
                        pass
                else:
                    self.lbl_preview.configure(image="", text="👤")

                btn_guardar.configure(text="💾 ACTUALIZAR", bootstyle="warning")

        tree_users.bind("<<TreeviewSelect>>", llenar_formulario)

        def guardar():
            import re, json

            # ── CAMBIO: se elimina .upper() para respetar el texto tal como se escribe ──
            u = var_user.get().strip()
            p = var_pass.get().strip()
            r = var_rol.get()
            n = var_nombre.get().strip().upper() or u.upper()
            e = var_email.get().strip()
            f = var_foto_path.get()

            if not u:
                messagebox.showwarning(
                    "Campo requerido",
                    "El nombre de usuario es obligatorio.",
                    parent=top
                )
                top.lift(); top.focus_force()
                return

            es_nuevo    = (self.usuario_seleccionado_id is None)
            cambio_pass = bool(p)

            if es_nuevo and not p:
                messagebox.showwarning(
                    "Contraseña requerida",
                    "⚠️  Debes asignar una contraseña al nuevo usuario.\n\n"
                    "📋  REQUISITOS OBLIGATORIOS:\n"
                    "   • Mínimo 8 caracteres\n"
                    "   • Al menos UNA MAYÚSCULA  (A-Z)\n"
                    "   • Al menos una minúscula  (a-z)\n"
                    "   • Al menos UN número      (0-9)\n"
                    "   • Solo alfanumérico (sin símbolos)\n\n"
                    "Ejemplo válido:  Almacen2024",
                    parent=top
                )
                top.lift(); top.focus_force()
                e_pass.focus_set()
                return

            if cambio_pass:
                errores = []
                if len(p) < 8:
                    errores.append(f"  • Mínimo 8 caracteres (tienes {len(p)})")
                if not re.search(r'[A-Z]', p):
                    errores.append("  • Falta al menos una MAYÚSCULA  (A-Z)")
                if not re.search(r'[a-z]', p):
                    errores.append("  • Falta al menos una minúscula  (a-z)")
                if not re.search(r'[0-9]', p):
                    errores.append("  • Falta al menos un número      (0-9)")
                if not re.match(r'^[a-zA-Z0-9]+$', p):
                    errores.append("  • Solo letras y números (sin espacios ni símbolos)")

                if errores:
                    messagebox.showerror(
                        "❌  Contraseña no válida",
                        "La contraseña NO cumple los estándares:\n\n"
                        + "\n".join(errores) +
                        "\n\n──────────────────────────────────\n"
                        "📋  REQUISITOS OBLIGATORIOS:\n"
                        "   ✅  Mínimo 8 caracteres\n"
                        "   ✅  Al menos una MAYÚSCULA  (A-Z)\n"
                        "   ✅  Al menos una minúscula  (a-z)\n"
                        "   ✅  Al menos un número      (0-9)\n"
                        "   ✅  Solo letras y números   (sin símbolos)\n\n"
                        "Ejemplo válido:  Almacen2024",
                        parent=top
                    )
                    top.lift(); top.focus_force()
                    e_pass.focus_set()
                    return

                salt    = secrets.token_hex(16)
                h_nuevo = hashlib.pbkdf2_hmac(
                    'sha256', p.encode(), salt.encode(), 100000
                ).hex()
                p_fin = f"{salt}:{h_nuevo}"
            else:
                p_fin = self.password_actual_hash

            permisos = json.dumps({
                "crear":     var_p_crear.get(),
                "entrada":   var_p_ent.get(),
                "salida":    var_p_sal.get(),
                "editar":    var_p_edit.get(),
                "eliminar":  var_p_del.get(),
                "catalogos": var_p_cat.get(),
                "historico": var_p_hist.get(),
                "ajustes":   var_p_conf.get()
            })

            try:
                if self.usuario_seleccionado_id:
                    self.db.ejecutar(
                        "UPDATE usuarios SET usuario=?, password=?, rol=?, "
                        "nombre_completo=?, email=?, foto_path=?, permisos=? WHERE id=?",
                        (u, p_fin, r, n, e, f, permisos, self.usuario_seleccionado_id)
                    )
                    accion = "actualizado"
                else:
                    self.db.ejecutar(
                        "INSERT INTO usuarios "
                        "(usuario, password, rol, nombre_completo, email, foto_path, permisos) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (u, p_fin, r, n, e, f, permisos)
                    )
                    accion = "creado"

                limpiar_form()
                cargar_usuarios_en_lista()
                messagebox.showinfo(
                    "✅  Guardado",
                    f"Usuario '{u}' {accion} correctamente.",
                    parent=top
                )
                top.lift(); top.focus_force()

            except sqlite3.IntegrityError:
                messagebox.showerror(
                    "Usuario duplicado",
                    f"Ya existe un usuario con el nombre '{u}'.\n"
                    "Elige un nombre diferente.",
                    parent=top
                )
                top.lift(); top.focus_force()
            except Exception as ex:
                messagebox.showerror(
                    "Error al guardar",
                    f"No se pudo guardar el usuario:\n{ex}",
                    parent=top
                )
                top.lift(); top.focus_force()

        def eliminar():
            if self.usuario_seleccionado_id:
                if messagebox.askyesno("Borrar", "¿Eliminar usuario?", parent=top):
                    self.db.ejecutar(
                        "DELETE FROM usuarios WHERE id=?",
                        (self.usuario_seleccionado_id,)
                    )
                    limpiar_form()
                    cargar_usuarios_en_lista()
                    top.lift(); top.focus_force()

        # ── Botonera inferior ─────────────────────────────────────────
        fr_btns = ttk.Frame(fr_form, padding=(0, 10))
        fr_btns.pack(side=BOTTOM, fill=X)

        ttk.Button(
            fr_btns, text="🧹 Limpiar",
            bootstyle="secondary-outline",
            command=limpiar_form
        ).pack(side=LEFT, expand=True, fill=X, padx=2)

        btn_guardar = ttk.Button(
            fr_btns, text="💾 CREAR NUEVO",
            bootstyle="success", command=guardar
        )
        btn_guardar.pack(side=LEFT, expand=True, fill=X, padx=2)

        ttk.Button(
            fr_btns, text="🗑️ Eliminar",
            bootstyle="danger",
            command=eliminar
        ).pack(side=LEFT, expand=True, fill=X, padx=2)

        # Carga inicial
        cargar_usuarios_en_lista()
        e_user.focus_set()


    def registrar_accion(self, tipo, partida, material, cantidad, destino, detalles=""):
        """Registra auditoría en el historial (quién hizo qué)."""
        try:
            fecha       = datetime.now().strftime("%d/%m/%Y %H:%M")
            usuario_act = self.usuario.get('usuario', 'SISTEMA') if hasattr(self, 'usuario') else 'SISTEMA'

            self.db.ejecutar(
                "INSERT INTO historial "
                "(fecha_hora, tipo, partida, material, cantidad, "
                "destino, responsable, entrego, factura, usuario_sistema) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (fecha, tipo, partida, material, cantidad,
                 destino, usuario_act, "AUDITORIA", detalles, usuario_act)
            )
        except Exception as e:
            print(f"Error log: {e}")

    # ------------------------------------------------------------------
    # FUNCIONES DE EDICIÓN Y BORRADO DE MATERIALES
    # ------------------------------------------------------------------
    def editar_material_seleccionado(self):
        """
        Permite corregir datos. 
        Sincroniza AUTOMÁTICAMENTE el historial si cambias el nombre o la factura.
        """
        sel = self.tree_inv.selection()
        if not sel: return
        
        # Obtener datos actuales del renglón seleccionado
        item = self.tree_inv.item(sel[0])
        valores = item['values'] 
        id_mat = valores[0]
        # Recuperamos datos frescos de la BD para no fallar
        db_data = self.db.consultar("SELECT * FROM inventario WHERE id=?", (id_mat,))
        if not db_data: return
        
        data = db_data[0]
        partida_actual = data['partida']
        nombre_actual = data['material']
        factura_actual = data['factura']

        # Ventana de Edición
        top = tk.Toplevel(self.root)
        top.title("Editar y Sincronizar")
        self.centrar_ventana_emergente(top, 450, 400)

        ttk.Label(top, text="✏️ Editar Material", font=("Segoe UI", 12, "bold"), bootstyle="primary").pack(pady=10)

        # Campos
        ttk.Label(top, text="Partida:").pack(anchor=W, padx=20)
        e_partida = ttk.Entry(top); e_partida.pack(fill=X, padx=20)
        e_partida.insert(0, partida_actual)

        ttk.Label(top, text="Descripción (Nombre):").pack(anchor=W, padx=20)
        e_mat = ttk.Entry(top); e_mat.pack(fill=X, padx=20)
        e_mat.insert(0, nombre_actual)
        
        ttk.Label(top, text="Factura (Referencia):").pack(anchor=W, padx=20)
        e_fac = ttk.Entry(top); e_fac.pack(fill=X, padx=20)
        e_fac.insert(0, factura_actual)

        def guardar_cambios():
            p_new = e_partida.get().strip()
            m_new = e_mat.get().strip().upper()
            f_new = e_fac.get().strip().upper()

            if not p_new or not m_new:
                messagebox.showwarning("Error", "El nombre y partida son obligatorios.")
                return

            if messagebox.askyesno("Confirmar", "¿Guardar cambios?\nSe actualizará el inventario y el historial."):
                try:
                    # 1. ACTUALIZAR TABLA INVENTARIO
                    self.db.ejecutar("""
                        UPDATE inventario SET partida=?, material=?, factura=? WHERE id=?
                    """, (p_new, m_new, f_new, id_mat))
                    
                    # 2. SINCRONIZAR HISTORIAL (MAGIA AQUÍ)
                    
                    # A) Si cambió el NOMBRE, actualizamos todo el historial viejo
                    if m_new != nombre_actual: 
                        self.db.ejecutar("""
                            UPDATE historial SET material=? WHERE material=?
                        """, (m_new, nombre_actual))
                        print(f"Historial renombrado: {nombre_actual} -> {m_new}")

                    # B) Si cambió la FACTURA, preguntamos si actualizar entradas viejas
                    if f_new != factura_actual:
                        if messagebox.askyesno("Actualizar Facturas", 
                                               "Has cambiado la factura.\n"
                                               "¿Quieres aplicar esta factura a las ENTRADAS pasadas en el historial?"):
                            self.db.ejecutar("""
                                UPDATE historial SET factura=? 
                                WHERE material=? AND (tipo LIKE '%ENTRADA%' OR tipo LIKE '%ALTA%')
                            """, (f_new, m_new))

                    # 3. LOG DE AUDITORÍA
                    self.registrar_accion("EDICION", p_new, m_new, 0, "SISTEMA", "Actualización de datos maestros")
                    
                    messagebox.showinfo("Éxito", "Material y Historial sincronizados correctamente.")
                    top.destroy()
                    
                    # Recargar todo
                    self.cargar_tabla_inventario()
                    self.cargar_tabla_historial()
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Fallo al guardar: {e}")

        ttk.Button(top, text="💾 GUARDAR Y SINCRONIZAR", bootstyle="success", command=guardar_cambios).pack(fill=X, padx=20, pady=20)
    def eliminar_material_seleccionado(self):
        """SOLO ADMIN: Elimina un material completo"""
        if self.usuario.get('rol') != 'ADMIN':
            messagebox.showerror("Acceso Denegado", "Solo el Administrador puede eliminar materiales.")
            return

        sel = self.tree_inv.selection()
        if not sel: return
        
        item = self.tree_inv.item(sel[0])
        id_mat = item['values'][0]
        nombre_mat = item['values'][2]

        if messagebox.askyesno("⚠️ ELIMINAR MATERIAL", 
                               f"¿Estás seguro de ELIMINAR PERMANENTEMENTE:\n\n'{nombre_mat}'?\n\n"
                               "Se perderá el stock actual. Esta acción es irreversible.", icon='warning'):
            try:
                self.db.ejecutar("DELETE FROM inventario WHERE id=?", (id_mat,))
                self.registrar_accion("ELIMINADO", "N/A", nombre_mat, 0, "PAPELERA", "Admin eliminó material")
                messagebox.showinfo("Listo", "Material eliminado.")
                self.cargar_tabla_inventario()
            except Exception as e:
                messagebox.showerror("Error", f"Fallo: {e}")

    def revertir_historial_admin(self):
        """
        Permite al admin borrar un registro del historial y opcionalmente
        REVERTIR el efecto que tuvo en el stock. (SIN DEJAR LOG DE AUDITORÍA)
        """
        sel = self.tree_hist.selection()
        if not sel: return
        
        item = self.tree_hist.item(sel[0])
        vals = item['values']
        
        # OBTENER DATOS DEL RENGLÓN SELECCIONADO
        # Indices: 0=ID, 2=TIPO, 3=PARTIDA, 4=MATERIAL, 5=CANTIDAD
        id_hist = vals[0]
        tipo = vals[2]
        partida = vals[3]
        material = vals[4]
        
        try: 
            cantidad = float(vals[5])
        except: 
            cantidad = 0
        
        # 1. CONFIRMACIÓN DE BORRADO
        if messagebox.askyesno("Eliminar Historial", 
                               f"¿Eliminar registro de '{tipo}' (ID: {id_hist})?\n\n"
                               "Esta acción será permanente."):
            
            revertir = False
            # Solo preguntamos si queremos afectar stock si fue un movimiento real
            if "ENTRADA" in tipo or "SALIDA" in tipo or "HISTORICO" in tipo:
                revertir = messagebox.askyesno("Revertir Stock", 
                                               f"Este registro movió {cantidad} piezas.\n\n"
                                               "¿Deseas REVERTIR ese movimiento en el inventario actual?\n"
                                               "(Si dices SÍ, el stock se ajustará automáticamente).")

            try:
                if revertir:
                    # BUSCAR EL MATERIAL EN EL INVENTARIO ACTUAL
                    res = self.db.consultar("SELECT stock FROM inventario WHERE partida=? AND material=?", (partida, material))
                    if res:
                        stock_actual = res[0]['stock']
                        nuevo_stock = stock_actual
                        
                        # LÓGICA INVERSA (DESHACER EL CAMBIO)
                        if "ENTRADA" in tipo or "HISTORICO (+)" in tipo:
                            # Si metimos cosas y borramos el registro -> RESTAMOS (quitamos lo que metimos por error)
                            nuevo_stock = stock_actual - cantidad
                        elif "SALIDA" in tipo or "HISTORICO (-)" in tipo:
                            # Si sacamos cosas y borramos el registro -> DEVOLVEMOS (regresamos lo que sacamos)
                            nuevo_stock = stock_actual + cantidad
                        
                        # Actualizar en Base de Datos
                        self.db.ejecutar("UPDATE inventario SET stock=? WHERE partida=? AND material=?", 
                                         (nuevo_stock, partida, material))
                        print(f"Stock revertido. Nuevo stock: {nuevo_stock}")
                    else:
                        messagebox.showwarning("Aviso", "El material ya no existe en el inventario, solo se borrará el historial.")

                # 2. ELIMINAR EL RENGLÓN DEL HISTORIAL
                self.db.ejecutar("DELETE FROM historial WHERE id=?", (id_hist,))
                
                # --- AQUÍ QUITAMOS LA LÍNEA DE AUDITORÍA "LOG_BORRADO" ---
                # self.registrar_accion(...)  <-- ELIMINADO

                messagebox.showinfo("Listo", "Registro eliminado correctamente.")
                
                # ACTUALIZAR TABLAS
                self.cargar_tabla_historial()
                self.cargar_tabla_inventario()

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo completar: {e}")
    
    # ------------------------------------------------------------------
    #  NUEVA SECCIÓN: MÓDULO DE CONSUMO Y ESTADÍSTICAS (BI)
    # ------------------------------------------------------------------
    # ------------------------------------------------------------------
    #  MÓDULO DE CONSUMO (CORREGIDO: GRÁFICA + LISTA COMPLETA)
    # ------------------------------------------------------------------
    def setup_tab_consumo(self):
        """Panel de Consumo con fechas exactas día/mes/año"""

        # ── PANEL SUPERIOR: FILTROS ──────────────────────────────────
        fr_filtros = ttk.LabelFrame(
            self.tab_consumo,
            text=" 🔍 Configuración del Reporte ",
            padding=12, bootstyle="primary")
        fr_filtros.pack(fill=X, padx=5, pady=(5, 0))

        # Fila 1 — Tipo de reporte
        fr_tipo = ttk.Frame(fr_filtros)
        fr_tipo.pack(fill=X, pady=(0, 8))

        ttk.Label(
            fr_tipo,
            text="Tipo de Reporte:",
            font=("Segoe UI", 10, "bold")
        ).pack(side=LEFT, padx=(0, 10))

        self.var_tipo_reporte = tk.StringVar(value="PERIODO")

        ttk.Radiobutton(
            fr_tipo,
            text="📅  Consumos del Periodo",
            variable=self.var_tipo_reporte,
            value="PERIODO",
            bootstyle="primary",
            command=self._actualizar_filtros_visibles
        ).pack(side=LEFT, padx=8)

        ttk.Radiobutton(
            fr_tipo,
            text="📂  Consumos del Periodo por Partida",
            variable=self.var_tipo_reporte,
            value="PERIODO_PARTIDA",
            bootstyle="primary",
            command=self._actualizar_filtros_visibles
        ).pack(side=LEFT, padx=8)

        ttk.Radiobutton(
            fr_tipo,
            text="📊  Consumo General por Partida",
            variable=self.var_tipo_reporte,
            value="GENERAL_PARTIDA",
            bootstyle="primary",
            command=self._actualizar_filtros_visibles
        ).pack(side=LEFT, padx=8)

        ttk.Separator(fr_filtros, orient=HORIZONTAL).pack(fill=X, pady=6)

        # Fila 2 — Filtros dinámicos
        self.fr_filtros_dinamicos = ttk.Frame(fr_filtros)
        self.fr_filtros_dinamicos.pack(fill=X)

        # ── Grupo A: Fechas exactas con día ───────────────────────────
        self.fr_grupo_fechas = ttk.Frame(self.fr_filtros_dinamicos)

        ttk.Label(
            self.fr_grupo_fechas,
            text="📅 Desde (DD/MM/AAAA):",
            font=("Segoe UI", 9, "bold")
        ).pack(side=LEFT)

        # Campo día inicio
        self.ent_dia_ini = ttk.Entry(
            self.fr_grupo_fechas, width=3, justify=CENTER,
            font=("Segoe UI", 10, "bold"))
        self.ent_dia_ini.insert(0, "01")
        self.ent_dia_ini.pack(side=LEFT, padx=(6, 1))

        ttk.Label(self.fr_grupo_fechas, text="/").pack(side=LEFT)

        # Campo mes inicio
        self.ent_mes_ini = ttk.Entry(
            self.fr_grupo_fechas, width=3, justify=CENTER,
            font=("Segoe UI", 10, "bold"))
        self.ent_mes_ini.insert(0, "01")
        self.ent_mes_ini.pack(side=LEFT, padx=1)

        ttk.Label(self.fr_grupo_fechas, text="/").pack(side=LEFT)

        # Campo año inicio
        self.ent_anio_ini = ttk.Entry(
            self.fr_grupo_fechas, width=5, justify=CENTER,
            font=("Segoe UI", 10, "bold"))
        self.ent_anio_ini.insert(0, str(datetime.now().year))
        self.ent_anio_ini.pack(side=LEFT, padx=(1, 15))

        # ── TAB automático entre campos de fecha ──────────────────────
        self.ent_dia_ini.bind("<Tab>",
            lambda e: (self.ent_mes_ini.focus_set(), "break")[1])
        self.ent_mes_ini.bind("<Tab>",
            lambda e: (self.ent_anio_ini.focus_set(), "break")[1])

        # Separador visual
        ttk.Label(
            self.fr_grupo_fechas,
            text="➜  Hasta (DD/MM/AAAA):",
            font=("Segoe UI", 9, "bold")
        ).pack(side=LEFT, padx=(0, 6))

        # Campo día fin
        self.ent_dia_fin = ttk.Entry(
            self.fr_grupo_fechas, width=3, justify=CENTER,
            font=("Segoe UI", 10, "bold"))
        self.ent_dia_fin.insert(0, str(datetime.now().day).zfill(2))
        self.ent_dia_fin.pack(side=LEFT, padx=(0, 1))

        ttk.Label(self.fr_grupo_fechas, text="/").pack(side=LEFT)

        # Campo mes fin
        self.ent_mes_fin = ttk.Entry(
            self.fr_grupo_fechas, width=3, justify=CENTER,
            font=("Segoe UI", 10, "bold"))
        self.ent_mes_fin.insert(0, str(datetime.now().month).zfill(2))
        self.ent_mes_fin.pack(side=LEFT, padx=1)

        ttk.Label(self.fr_grupo_fechas, text="/").pack(side=LEFT)

        # Campo año fin
        self.ent_anio_fin = ttk.Entry(
            self.fr_grupo_fechas, width=5, justify=CENTER,
            font=("Segoe UI", 10, "bold"))
        self.ent_anio_fin.insert(0, str(datetime.now().year))
        self.ent_anio_fin.pack(side=LEFT, padx=(1, 0))

        # TAB entre campos fin
        self.ent_dia_fin.bind("<Tab>",
            lambda e: (self.ent_mes_fin.focus_set(), "break")[1])
        self.ent_mes_fin.bind("<Tab>",
            lambda e: (self.ent_anio_fin.focus_set(), "break")[1])

        # ── Grupo B: Selector de Partida ─────────────────────────────
        self.fr_grupo_partida = ttk.Frame(self.fr_filtros_dinamicos)

        ttk.Label(
            self.fr_grupo_partida,
            text="📂 Partida:",
            font=("Segoe UI", 9, "bold")
        ).pack(side=LEFT)

        self.cb_partida_consumo = ttk.Combobox(
            self.fr_grupo_partida,
            state="readonly", width=20)
        self.cb_partida_consumo.pack(side=LEFT, padx=(6, 0))

        # Cargar partidas
        rows_p = self.db.consultar(
            "SELECT valor FROM catalogos WHERE tipo='PARTIDA' ORDER BY valor ASC")
        lista_p = [r['valor'] for r in rows_p]
        self.cb_partida_consumo['values'] = lista_p
        if lista_p:
            self.cb_partida_consumo.current(0)

        # ── Botón Generar + Exportar ──────────────────────────────────
        fr_btn = ttk.Frame(fr_filtros)
        fr_btn.pack(fill=X, pady=(10, 0))

        ttk.Button(
            fr_btn,
            text="🔄  Generar Reporte",
            bootstyle="primary",
            command=self.generar_grafica_consumo
        ).pack(side=LEFT, ipady=4, padx=(0, 8))

        ttk.Button(
            fr_btn,
            text="💾  Exportar a Excel",
            bootstyle="success-outline",
            command=self.exportar_excel_consumo
        ).pack(side=LEFT, ipady=4, padx=(0, 15))

        self.lbl_resumen_total = ttk.Label(
            fr_btn, text="",
            font=("Segoe UI", 11, "bold"),
            foreground="#2980b9")
        self.lbl_resumen_total.pack(side=LEFT)

        # ── PANEL CENTRAL: GRÁFICA ────────────────────────────────────
        self.fr_grafica_container = ttk.Frame(
            self.tab_consumo,
            padding=5, relief="solid", borderwidth=1)
        self.fr_grafica_container.pack(
            fill=BOTH, expand=True, padx=5, pady=5)

        ttk.Label(
            self.fr_grafica_container,
            text="📊  Configura los filtros y haz clic en 'Generar Reporte'",
            foreground="gray",
            font=("Segoe UI", 12)
        ).place(relx=0.5, rely=0.5, anchor=CENTER)

        # ── PANEL INFERIOR: TABLA DETALLE ─────────────────────────────
        fr_tabla = ttk.LabelFrame(
            self.tab_consumo,
            text=" 📋 Detalle de Consumo (Ranking Completo) ",
            padding=8, bootstyle="secondary")
        fr_tabla.pack(fill=X, padx=5, pady=(0, 5))

        cols = ("RANK", "PARTIDA", "MATERIAL", "CANTIDAD", "PORCENTAJE")
        self.tree_consumo = ttk.Treeview(
            fr_tabla, columns=cols,
            show="headings", height=7, bootstyle="info")

        self.tree_consumo.heading("RANK",       text="N°")
        self.tree_consumo.column( "RANK",       width=40,  anchor=CENTER)
        self.tree_consumo.heading("PARTIDA",    text="PARTIDA")
        self.tree_consumo.column( "PARTIDA",    width=80,  anchor=CENTER)
        self.tree_consumo.heading("MATERIAL",   text="MATERIAL / PRODUCTO")
        self.tree_consumo.column( "MATERIAL",   width=380)
        self.tree_consumo.heading("CANTIDAD",   text="CONSUMO (Pzas)")
        self.tree_consumo.column( "CANTIDAD",   width=110, anchor=CENTER)
        self.tree_consumo.heading("PORCENTAJE", text="% TOTAL")
        self.tree_consumo.column( "PORCENTAJE", width=90,  anchor=CENTER)

        sc_tabla = ttk.Scrollbar(
            fr_tabla, orient=VERTICAL,
            command=self.tree_consumo.yview)
        self.tree_consumo.configure(yscrollcommand=sc_tabla.set)
        self.tree_consumo.pack(side=LEFT, fill=BOTH, expand=True)
        sc_tabla.pack(side=RIGHT, fill=Y)

        # Mostrar filtros iniciales
        self._actualizar_filtros_visibles()


    def generar_grafica_consumo(self):
        """
        Gráfica de consumo optimizada para 200+ materiales.
        - Tabla muestra todos con scroll virtual
        - Gráfica muestra Top N configurable (10, 15, 20)
        - Consulta SQL directa sin procesar todo en Python
        """
        import matplotlib.animation as animation

        # ── Limpiar área ──────────────────────────────────────────────
        for widget in self.fr_grafica_container.winfo_children():
            widget.destroy()
        for i in self.tree_consumo.get_children():
            self.tree_consumo.delete(i)

        if hasattr(self, 'mi_animacion') and self.mi_animacion:
            try:
                self.mi_animacion.event_source.stop()
            except:
                pass

        tipo = self.var_tipo_reporte.get()

        # ── Leer y validar fechas ─────────────────────────────────────
        def leer_fecha(ent_dia, ent_mes, ent_anio, nombre):
            try:
                dia  = int(ent_dia.get().strip())
                mes  = int(ent_mes.get().strip())
                anio = int(ent_anio.get().strip())
                if not (1 <= dia <= 31): raise ValueError(f"Día inválido en {nombre}")
                if not (1 <= mes <= 12): raise ValueError(f"Mes inválido en {nombre}")
                if anio < 2000:          raise ValueError(f"Año inválido en {nombre}")
                return datetime(anio, mes, dia)
            except ValueError as e:
                messagebox.showerror(
                    "❌ Fecha inválida",
                    f"Error en la fecha {nombre}:\n{e}\n\n"
                    "Formato correcto: DD / MM / AAAA",
                    parent=self.root)
                return None

        # ── Construir consulta SQL directa (más eficiente que Python) ─
        try:
            condiciones_sql = []
            params_sql      = []

            if tipo in ("PERIODO", "PERIODO_PARTIDA"):
                fecha_ini = leer_fecha(
                    self.ent_dia_ini, self.ent_mes_ini,
                    self.ent_anio_ini, "DESDE")
                if fecha_ini is None: return

                fecha_fin = leer_fecha(
                    self.ent_dia_fin, self.ent_mes_fin,
                    self.ent_anio_fin, "HASTA")
                if fecha_fin is None: return

                fecha_fin = fecha_fin.replace(hour=23, minute=59, second=59)

                if fecha_ini > fecha_fin:
                    messagebox.showerror(
                        "❌ Error de fechas",
                        "La fecha DESDE no puede ser mayor que HASTA.",
                        parent=self.root)
                    return

                str_ini = fecha_ini.strftime("%d/%m/%Y")
                str_fin = fecha_fin.strftime("%d/%m/%Y")

                if tipo == "PERIODO_PARTIDA":
                    partida_filtro = self.cb_partida_consumo.get()
                    titulo_grafica = (
                        f"Consumo Partida {partida_filtro}\n"
                        f"{str_ini}  →  {str_fin}")
                else:
                    partida_filtro = None
                    titulo_grafica = f"Consumo del {str_ini}  →  {str_fin}"

            elif tipo == "GENERAL_PARTIDA":
                partida_filtro = self.cb_partida_consumo.get()
                titulo_grafica = f"Consumo General  —  Partida {partida_filtro}"
                fecha_ini      = None
                fecha_fin      = None

        except Exception as e:
            messagebox.showerror("Error", f"Error procesando filtros:\n{e}")
            return

        # ── Obtener datos agrupados directo desde SQL ─────────────────
        # Mucho más rápido que iterar en Python con 200+ materiales
        try:
            sql_base = """
                SELECT
                    h.material,
                    COALESCE(i.partida, 'S/P') AS partida,
                    SUM(h.cantidad)            AS total
                FROM historial h
                LEFT JOIN inventario i ON h.material = i.material
                WHERE (
                    h.tipo LIKE '%SALIDA%'
                    OR h.tipo LIKE '%HISTORICO (-)%'
                )
            """
            params = []

            if fecha_ini and fecha_fin:
                # Filtrar por rango de fechas usando substr de SQLite
                sql_base += """
                    AND (
                        CAST(substr(h.fecha_hora, 7, 4) AS INTEGER) * 10000
                        + CAST(substr(h.fecha_hora, 4, 2) AS INTEGER) * 100
                        + CAST(substr(h.fecha_hora, 1, 2) AS INTEGER)
                    ) BETWEEN ? AND ?
                """
                fecha_ini_int = (fecha_ini.year * 10000
                                 + fecha_ini.month * 100
                                 + fecha_ini.day)
                fecha_fin_int = (fecha_fin.year * 10000
                                 + fecha_fin.month * 100
                                 + fecha_fin.day)
                params.extend([fecha_ini_int, fecha_fin_int])

            if partida_filtro:
                sql_base += " AND i.partida = ?"
                params.append(partida_filtro)

            sql_base += " GROUP BY h.material, i.partida ORDER BY total DESC"

            datos_sql = self.db.consultar(sql_base, tuple(params))

        except Exception as e:
            messagebox.showerror("Error SQL", f"{e}")
            return

        # ── Sin datos ─────────────────────────────────────────────────
        if not datos_sql:
            ttk.Label(
                self.fr_grafica_container,
                text="⚠️  Sin movimientos para los filtros seleccionados.",
                font=("Segoe UI", 14), foreground="#e74c3c"
            ).place(relx=0.5, rely=0.5, anchor=CENTER)
            self.lbl_resumen_total.config(text="Total: 0 pzas")
            return

        total = sum(float(d['total']) for d in datos_sql)

        self.lbl_resumen_total.config(
            text=f"Total: {int(total):,} pzas  |  {len(datos_sql)} materiales")

        # ── Llenar tabla detalle (todos los materiales, paginado por Treeview) ─
        # Treeview con scroll ya maneja miles de filas sin problema
        for idx, d in enumerate(datos_sql, 1):
            cant = float(d['total'])
            pct  = (cant / total) * 100
            self.tree_consumo.insert(
                "", END,
                values=(
                    f"{idx}°",
                    d['partida'],
                    d['material'],
                    f"{int(cant):,}" if cant == int(cant) else f"{cant:,.1f}",
                    f"{pct:.1f}%"
                )
            )

        # ── Preparar datos para la gráfica ────────────────────────────
        # Selector de Top N dinámico
        TOP_N = getattr(self, '_top_n_grafica', 15)

        nombres_g    = [d['material'] for d in datos_sql[:TOP_N]]
        cantidades_g = [float(d['total']) for d in datos_sql[:TOP_N]]

        # Invertir para que el mayor quede arriba
        nombres_g.reverse()
        cantidades_g.reverse()
        n = len(nombres_g)

        # ── Colores del tema ──────────────────────────────────────────
        c_prim  = self.tema_actual.get("color_primario", "#3498db")
        c_acent = self.tema_actual.get("color_acento",   "#e74c3c")
        c_fondo = self.tema_actual.get("color_fondo",    "#FFFFFF")
        c_texto = self.tema_actual.get("color_texto",    "#000000")

        def hex_a_rgb(h):
            h = h.lstrip("#")
            return tuple(int(h[i:i+2], 16) / 255 for i in (0, 2, 4))

        rgb_p = hex_a_rgb(c_prim)
        rgb_a = hex_a_rgb(c_acent)
        colores_barras = [
            tuple(
                rgb_p[k] + (rgb_a[k] - rgb_p[k]) * (i / max(n - 1, 1))
                for k in range(3))
            for i in range(n)
        ]

        # ── Tamaño dinámico de figura según número de barras ──────────
        alto_fig  = max(4.5, n * 0.45)   # mínimo 4.5, crece con los items
        fig = Figure(figsize=(7, alto_fig), dpi=100)
        fig.patch.set_facecolor(c_fondo)

        ax = fig.add_subplot(111)
        ax.set_facecolor(c_fondo)

        # Tamaño de fuente dinámico en etiquetas del eje Y
        fs_etiquetas = max(6, 9 - (n // 8))

        barras = ax.barh(
            nombres_g, [0] * n,
            color=colores_barras,
            height=0.65, alpha=0.92
        )

        # Truncar nombres largos en el eje Y para que no se monten
        etiquetas_cortas = [
            (nom[:35] + "…") if len(nom) > 35 else nom
            for nom in nombres_g
        ]
        ax.set_yticks(range(n))
        ax.set_yticklabels(etiquetas_cortas, fontsize=fs_etiquetas)

        # Título con cantidad total de materiales
        titulo_completo = (
            f"{titulo_grafica}\n"
            f"(Top {TOP_N} de {len(datos_sql)} materiales)"
            if len(datos_sql) > TOP_N
            else titulo_grafica
        )

        ax.set_title(
            titulo_completo,
            fontsize=10, fontweight='bold',
            color=c_prim, pad=10)

        ax.set_xlabel(
            "Cantidad consumida (Pzas)",
            fontsize=9, color=c_texto)

        ax.tick_params(colors=c_texto, labelsize=fs_etiquetas)

        for sp in ax.spines.values():
            sp.set_visible(False)

        ax.xaxis.grid(True, color="#DDDDDD", linestyle="--", linewidth=0.7)
        ax.set_axisbelow(True)

        max_val = max(cantidades_g) if cantidades_g else 10
        ax.set_xlim(0, max_val * 1.22)

        # Etiquetas al final de barra (número formateado)
        textos_val = []
        for bar, val in zip(barras, cantidades_g):
            txt = ax.text(
                0,
                bar.get_y() + bar.get_height() / 2,
                f" {int(val):,} pzas",
                va='center', ha='left',
                fontsize=max(6, fs_etiquetas),
                color=c_texto, fontweight='bold')
            textos_val.append((txt, val))

        fig.tight_layout(pad=1.5)

        # ── Animación ─────────────────────────────────────────────────
        FRAMES = 30   # menos frames = más fluido con muchas barras

        def animar(frame):
            progreso = frame / FRAMES
            p = 1 - (1 - progreso) ** 3
            for bar, (txt, target) in zip(barras, textos_val):
                w = target * p
                bar.set_width(w)
                txt.set_x(w)
            return list(barras) + [t for t, _ in textos_val]

        # ── Hover tooltip ─────────────────────────────────────────────
        annot = ax.annotate(
            "", xy=(0, 0), xytext=(8, 8),
            textcoords="offset points",
            bbox=dict(boxstyle="round,pad=0.4",
                      fc="white", ec=c_prim, alpha=0.95),
            fontsize=9, fontweight='bold', color=c_prim)
        annot.set_visible(False)

        def on_hover(event):
            if event.inaxes != ax:
                annot.set_visible(False)
                fig.canvas.draw_idle()
                return
            for bar, (_, val), nom in zip(barras, textos_val, nombres_g):
                cont, _ = bar.contains(event)
                if cont:
                    pct_h = (val / total) * 100
                    annot.xy = (
                        bar.get_width(),
                        bar.get_y() + bar.get_height() / 2)
                    annot.set_text(
                        f"{nom[:40]}\n"
                        f"{int(val):,} pzas  ({pct_h:.1f}%)")
                    annot.set_visible(True)
                    fig.canvas.draw_idle()
                    return
            annot.set_visible(False)
            fig.canvas.draw_idle()

        # ── Canvas con scroll vertical para gráficas muy largas ───────
        fr_canvas_scroll = ttk.Frame(self.fr_grafica_container)
        fr_canvas_scroll.pack(fill=BOTH, expand=True)

        canvas_scroll = tk.Canvas(
            fr_canvas_scroll,
            highlightthickness=0,
            bg=c_fondo)
        sc_graf = ttk.Scrollbar(
            fr_canvas_scroll, orient=VERTICAL,
            command=canvas_scroll.yview)
        canvas_scroll.configure(yscrollcommand=sc_graf.set)

        # Solo mostrar scrollbar si hay muchas barras
        if n > 12:
            sc_graf.pack(side=RIGHT, fill=Y)
        canvas_scroll.pack(side=LEFT, fill=BOTH, expand=True)

        # ── Selector Top N ────────────────────────────────────────────
        fr_top_n = ttk.Frame(self.fr_grafica_container)
        fr_top_n.pack(fill=X, padx=5, pady=(0, 3))

        ttk.Label(
            fr_top_n,
            text=f"Mostrando Top {TOP_N} en gráfica  |  Total: {len(datos_sql)} materiales",
            font=("Segoe UI", 8), foreground="gray"
        ).pack(side=LEFT)

        if len(datos_sql) > 10:
            ttk.Label(fr_top_n, text="   Ver Top:", font=("Segoe UI", 8)).pack(side=LEFT)
            for n_val in [10, 15, 20, 30]:
                if n_val <= len(datos_sql):
                    ttk.Button(
                        fr_top_n,
                        text=str(n_val),
                        bootstyle="link",
                        width=3,
                        command=lambda v=n_val: self._cambiar_top_n(v)
                    ).pack(side=LEFT)

        # ── Dibujar matplotlib en tkinter ─────────────────────────────
        canvas_mpl = FigureCanvasTkAgg(fig, master=canvas_scroll)
        canvas_mpl.draw()

        widget_mpl = canvas_mpl.get_tk_widget()
        canvas_scroll.create_window((0, 0), window=widget_mpl, anchor="nw")

        def configurar_scroll(event):
            canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all"))

        widget_mpl.bind("<Configure>", configurar_scroll)

        # Scroll con rueda del ratón
        def _scroll_grafica(event):
            canvas_scroll.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas_scroll.bind("<Enter>",
            lambda e: canvas_scroll.bind_all("<MouseWheel>", _scroll_grafica))
        canvas_scroll.bind("<Leave>",
            lambda e: canvas_scroll.unbind_all("<MouseWheel>"))

        canvas_mpl.mpl_connect("motion_notify_event", on_hover)

        self.mi_animacion = animation.FuncAnimation(
            fig, animar,
            frames=FRAMES + 1,
            interval=20,
            blit=False,
            repeat=False)
        
    def _cambiar_top_n(self, n):
        """Cambia el Top N de la gráfica y la regenera"""
        self._top_n_grafica = n
        self.generar_grafica_consumo()

        
    def exportar_excel_consumo(self):
        """
        Exporta el reporte de consumo a Excel con tabla y gráfica.
        Compatible con los nuevos campos de fecha DD/MM/AAAA.
        """
        from openpyxl import Workbook
        from openpyxl.styles import (Font, Alignment, PatternFill,
                                      Border, Side as ExcelSide)
        from openpyxl.chart import BarChart, Reference

        # ── Verificar que hay datos ───────────────────────────────────
        filas_tabla = self.tree_consumo.get_children()
        if not filas_tabla:
            messagebox.showwarning(
                "Sin datos",
                "Primero genera el reporte antes de exportar.",
                parent=self.root)
            return

        # ── Pedir ruta ────────────────────────────────────────────────
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Reporte_Consumo.xlsx",
            title="Guardar Reporte de Consumo")
        if not ruta:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Consumo"

            # ── Estilos ───────────────────────────────────────────────
            c_prim   = self.tema_actual.get("color_primario", "#1F4E79")
            hex_prim = c_prim.lstrip("#")
            hex_gris = "F2F2F2"
            hex_blnc = "FFFFFF"

            thin  = ExcelSide(border_style="thin", color="CCCCCC")
            borde = Border(
                top=thin, left=thin,
                right=thin, bottom=thin)

            fuente_titulo = Font(bold=True, size=14,
                                  color=hex_prim, name="Segoe UI")
            fuente_subtit = Font(bold=True, size=10,
                                  color="444444", name="Segoe UI")
            fuente_header = Font(bold=True, size=10,
                                  color=hex_blnc, name="Segoe UI")
            fuente_normal = Font(size=10, name="Segoe UI")
            fuente_total  = Font(bold=True, size=10,
                                  color=hex_prim, name="Segoe UI")

            fill_header = PatternFill(
                fill_type="solid", fgColor=hex_prim)
            fill_gris   = PatternFill(
                fill_type="solid", fgColor=hex_gris)
            fill_total  = PatternFill(
                fill_type="solid", fgColor="EBF3FB")

            centro = Alignment(
                horizontal="center", vertical="center")
            izq    = Alignment(
                horizontal="left", vertical="center",
                wrap_text=True)

            # ── Encabezados institucionales ───────────────────────────
            h1 = self.db.get_config("HEADER_L1") or "INSTITUCIÓN"
            h2 = self.db.get_config("HEADER_L2") or "SUBDIRECCIÓN"
            h4 = self.db.get_config("HEADER_L4") or "DEPARTAMENTO"

            ws.merge_cells("A1:F1")
            ws["A1"]           = h1
            ws["A1"].font      = fuente_titulo
            ws["A1"].alignment = centro

            ws.merge_cells("A2:F2")
            ws["A2"]           = h2
            ws["A2"].font      = fuente_subtit
            ws["A2"].alignment = centro

            ws.merge_cells("A3:F3")
            ws["A3"]           = h4
            ws["A3"].font      = fuente_subtit
            ws["A3"].alignment = centro

            # ── Título del reporte ────────────────────────────────────
            tipo = self.var_tipo_reporte.get()

            titulos_tipo = {
                "PERIODO":         "REPORTE DE CONSUMO POR PERIODO",
                "PERIODO_PARTIDA": "REPORTE DE CONSUMO POR PERIODO Y PARTIDA",
                "GENERAL_PARTIDA": "REPORTE DE CONSUMO GENERAL POR PARTIDA"
            }
            titulo_rep = titulos_tipo.get(tipo, "REPORTE DE CONSUMO")

            # ── Subtítulo con parámetros ──────────────────────────────
            # Leer desde los nuevos campos DD/MM/AAAA
            if tipo in ("PERIODO", "PERIODO_PARTIDA"):
                try:
                    dia_i  = self.ent_dia_ini.get().strip().zfill(2)
                    mes_i  = self.ent_mes_ini.get().strip().zfill(2)
                    anio_i = self.ent_anio_ini.get().strip()
                    dia_f  = self.ent_dia_fin.get().strip().zfill(2)
                    mes_f  = self.ent_mes_fin.get().strip().zfill(2)
                    anio_f = self.ent_anio_fin.get().strip()

                    str_ini   = f"{dia_i}/{mes_i}/{anio_i}"
                    str_fin   = f"{dia_f}/{mes_f}/{anio_f}"
                    subtitulo = f"Periodo: {str_ini}  —  {str_fin}"

                    if tipo == "PERIODO_PARTIDA":
                        subtitulo += (
                            f"  |  Partida: "
                            f"{self.cb_partida_consumo.get()}")
                except:
                    subtitulo = "Periodo no disponible"
            else:
                subtitulo = (
                    f"Partida: {self.cb_partida_consumo.get()}"
                    f"  |  Todo el historial")

            ws.merge_cells("A5:F5")
            ws["A5"] = titulo_rep
            ws["A5"].font = Font(
                bold=True, size=13,
                color=hex_prim, name="Segoe UI")
            ws["A5"].alignment = centro

            ws.merge_cells("A6:F6")
            ws["A6"] = subtitulo
            ws["A6"].font = Font(
                size=10, italic=True,
                color="555555", name="Segoe UI")
            ws["A6"].alignment = centro

            # ── Encabezado de tabla ───────────────────────────────────
            fila_hdr = 8
            headers  = ["N°", "PARTIDA", "MATERIAL / PRODUCTO",
                         "CONSUMO (Pzas)", "% DEL TOTAL", "ACUMULADO"]

            for col, texto in enumerate(headers, 1):
                cell           = ws.cell(
                    row=fila_hdr, column=col, value=texto)
                cell.font      = fuente_header
                cell.fill      = fill_header
                cell.alignment = centro
                cell.border    = borde

            # ── Datos desde la tabla visual ───────────────────────────
            fila_dat   = fila_hdr + 1
            total_pzas = 0
            acumulado  = 0
            datos_graf = []

            for i, item_id in enumerate(filas_tabla):
                vals = self.tree_consumo.item(item_id)['values']
                rank    = vals[0]
                partida = vals[1]
                mat     = vals[2]
                cant    = float(str(vals[3]).replace(",", ""))
                pct     = vals[4]

                total_pzas += cant
                acumulado  += cant
                datos_graf.append((mat[:35], cant))

                ws.cell(row=fila_dat, column=1,
                        value=rank).alignment = centro
                ws.cell(row=fila_dat, column=2,
                        value=partida).alignment = centro
                ws.cell(row=fila_dat, column=3,
                        value=mat).alignment = izq
                ws.cell(row=fila_dat, column=4,
                        value=cant).alignment = centro
                ws.cell(row=fila_dat, column=5,
                        value=pct).alignment = centro
                ws.cell(row=fila_dat, column=6,
                        value=acumulado).alignment = centro

                for col in range(1, 7):
                    cell        = ws.cell(row=fila_dat, column=col)
                    cell.font   = fuente_normal
                    cell.border = borde
                    if i % 2 == 0:
                        cell.fill = fill_gris

                fila_dat += 1

            # ── Fila TOTAL ────────────────────────────────────────────
            ws.merge_cells(f"A{fila_dat}:C{fila_dat}")
            ws[f"A{fila_dat}"]           = "TOTAL GENERAL"
            ws[f"A{fila_dat}"].font      = fuente_total
            ws[f"A{fila_dat}"].alignment = centro
            ws[f"A{fila_dat}"].fill      = fill_total

            ws[f"D{fila_dat}"]           = total_pzas
            ws[f"D{fila_dat}"].font      = fuente_total
            ws[f"D{fila_dat}"].alignment = centro
            ws[f"D{fila_dat}"].fill      = fill_total

            ws[f"E{fila_dat}"]           = "100%"
            ws[f"E{fila_dat}"].font      = fuente_total
            ws[f"E{fila_dat}"].alignment = centro
            ws[f"E{fila_dat}"].fill      = fill_total

            for col in range(1, 7):
                ws.cell(row=fila_dat, column=col).border = borde

            # ── Anchos de columna ─────────────────────────────────────
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 45
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 14

            for fila in range(fila_hdr, fila_dat + 1):
                ws.row_dimensions[fila].height = 20
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[5].height = 28

            # ── GRÁFICA DE BARRAS (openpyxl) ──────────────────────────
            ws_graf = wb.create_sheet(title="Datos Gráfica")
            ws_graf["A1"] = "Material"
            ws_graf["B1"] = "Consumo"

            LIMITE_GRAF = 15
            for i, (nom, val) in enumerate(
                    datos_graf[:LIMITE_GRAF], 2):
                ws_graf.cell(row=i, column=1, value=nom)
                ws_graf.cell(row=i, column=2, value=val)

            n_datos = min(len(datos_graf), LIMITE_GRAF)

            chart             = BarChart()
            chart.type        = "bar"
            chart.grouping    = "clustered"
            chart.title       = titulo_rep
            chart.y_axis.title = "Material"
            chart.x_axis.title = "Cantidad (Pzas)"
            chart.style       = 10
            chart.width       = 25
            chart.height      = max(10, n_datos * 0.9)

            data_ref = Reference(
                ws_graf,
                min_col=2, min_row=1,
                max_col=2, max_row=n_datos + 1)
            cats_ref = Reference(
                ws_graf,
                min_col=1, min_row=2,
                max_row=n_datos + 1)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.series[0].graphicalProperties.solidFill = hex_prim

            ws.add_chart(chart, f"A{fila_dat + 3}")

            # ── Guardar ───────────────────────────────────────────────
            wb.save(ruta)
            messagebox.showinfo(
                "✅  Exportado",
                f"Reporte guardado correctamente:\n{ruta}",
                parent=self.root)
            os.startfile(ruta)

        except PermissionError:
            messagebox.showwarning(
                "Archivo abierto",
                "No se pudo guardar.\n"
                "El archivo está abierto en Excel.\n"
                "Ciérralo e intenta de nuevo.",
                parent=self.root)
        except Exception as e:
            messagebox.showerror(
                "Error al exportar",
                f"No se pudo generar el archivo:\n{e}",
                parent=self.root)
            
    def tiene_permiso(self, accion):
        """
        Verifica si el usuario actual tiene permiso para una acción específica.
        Acciones: 'crear', 'entrada', 'salida', 'editar', 'eliminar', 'ajustes'
        """
        # 1. Si es ADMIN por Rol, tiene acceso a todo siempre
        if self.usuario.get('rol') == 'ADMIN':
            return True
            
        # 2. Leer permisos JSON del usuario
        try:
            permisos_str = self.usuario.get('permisos', '{}')
            if not permisos_str: return False
            
            import json
            permisos_dict = json.loads(permisos_str)
            
            # Retorna True si la clave existe y es 1 (True)
            return permisos_dict.get(accion, 0) == 1
        except:
            return False

class SelectorDB:
    ARCHIVO_CONFIG = "config_conexion.json"

    @staticmethod
    def obtener_ruta_db(root_padre):
        """
        Intenta obtener la ruta. Si falla, abre ventana usando 'root_padre' como base.
        """
        # 1. Intentar leer config
        if os.path.exists(SelectorDB.ARCHIVO_CONFIG):
            try:
                with open(SelectorDB.ARCHIVO_CONFIG, 'r') as f:
                    data = json.load(f)
                    ruta = data.get("ruta_db")
                    if ruta and os.path.exists(ruta):
                        return ruta
            except: pass

        # 2. Si falla, pedir al usuario usando la raíz existente
        return SelectorDB.abrir_ventana_seleccion(root_padre)
    
    

    @staticmethod
    def abrir_ventana_seleccion(root_padre):
        """Muestra el selector como una ventana hija (Toplevel)"""
        
        selector = tk.Toplevel(root_padre)
        selector.title("Configuración Inicial")
        
        # SE ELIMINÓ 'selector.transient(root_padre)' para que tenga su 
        # propia presencia en la barra de tareas de Windows sin depender del cuadro blanco.
        
        selector.grab_set()
        
        w, h = 500, 350
        ws = root_padre.winfo_screenwidth()
        hs = root_padre.winfo_screenheight()
        selector.geometry(f'{w}x{h}+{int((ws/2)-(w/2))}+{int((hs/2)-(h/2))}')
        
        resultado = [""] 

        ttk.Label(selector, text="BIENVENIDO AL SISTEMA", 
                 font=("Arial", 16, "bold"), bootstyle="primary").pack(pady=(30, 10))
        
        ttk.Label(selector, text="Para comenzar, necesitamos una Base de Datos.\n¿Qué deseas hacer?", 
                 justify=CENTER).pack(pady=10)

        def guardar_y_cerrar(ruta):
            try:
                with open(SelectorDB.ARCHIVO_CONFIG, 'w') as f:
                    json.dump({"ruta_db": ruta}, f)
                resultado[0] = ruta
                selector.destroy() 
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar config: {e}")

        def crear_nueva():
            ruta = filedialog.asksaveasfilename(
                parent=selector, 
                title="Crear Nueva Base de Datos",
                defaultextension=".db",
                filetypes=[("Archivos SQLite", "*.db")],
                initialfile="inventario_unindetec.db"
            )
            if ruta: guardar_y_cerrar(ruta)

        def abrir_existente():
            ruta = filedialog.askopenfilename(
                parent=selector,
                title="Seleccionar Base de Datos Existente",
                filetypes=[("Archivos SQLite", "*.db")]
            )
            if ruta: guardar_y_cerrar(ruta)
            
        def al_cerrar():
            if not resultado[0]:
                if messagebox.askyesno("Salir", "¿Deseas salir del sistema?"):
                    root_padre.destroy() 
                    sys.exit()
            else:
                selector.destroy()

        selector.protocol("WM_DELETE_WINDOW", al_cerrar)

        fr_btns = ttk.Frame(selector, padding=20)
        fr_btns.pack(fill=BOTH, expand=True)

        ttk.Button(fr_btns, text="📂 BUSCAR ARCHIVO EXISTENTE", bootstyle="info", 
                   command=abrir_existente).pack(fill=X, pady=5, ipady=8)
        
        ttk.Label(fr_btns, text="- O -", bootstyle="secondary").pack(pady=5)
        
        ttk.Button(fr_btns, text="✨ CREAR BASE DE DATOS NUEVA", bootstyle="success", 
                   command=crear_nueva).pack(fill=X, pady=5, ipady=8)

        root_padre.wait_window(selector)
        
        return resultado[0]
# FUNCIONES GLOBALES (FUERA DE LAS CLASES)
# ==========================================

# ==========================================
# 1. PANTALLA DE CARGA "MODO CIBERSEGURIDAD"


def mostrar_splash_epico(db_manager, callback):
    """
    Splash Screen Profesional: Estilo Minimalista/Corporativo (Blanco)
    CORREGIDO: Sin el error de letterspacing
    """
    splash = tk.Toplevel()
    splash.overrideredirect(True)
    
    # --- CONFIGURACIÓN DE COLORES Y ESTILO ---
    COLOR_FONDO = "#FFFFFF"       # Blanco Puro
    COLOR_TITULO = "#2c3e50"      # Gris Azulado Oscuro
    COLOR_SUBTITULO = "#7f8c8d"   # Gris Medio
    COLOR_ACCENTO = "#3498db"     # Azul Corporativo
    COLOR_BARRA_FONDO = "#ecf0f1" # Gris muy claro
    
    splash.configure(bg=COLOR_FONDO)
    
    # --- DIMENSIONES Y CENTRADO ---
    ancho = 600
    alto = 350
    ws = splash.winfo_screenwidth()
    hs = splash.winfo_screenheight()
    x = (ws/2) - (ancho/2)
    y = (hs/2) - (alto/2)
    splash.geometry(f'{ancho}x{alto}+{int(x)}+{int(y)}')
    
    # --- OBTENER DATOS DE LA BD ---
    tit = db_manager.get_config("TITULO_APP")
    if not tit: tit = "SOTFWARE DE USO LIBRE"
    
    sub = db_manager.get_config("SUBTITULO_APP")
    if not sub: sub = "CONTROL DE PARTIDAS"
    
    logo_path = db_manager.get_config("LOGO_APP")
    
    # --- INTERFAZ GRÁFICA ---
    
    # 1. Contenedor Central
    fr_centro = tk.Frame(splash, bg=COLOR_FONDO)
    fr_centro.pack(expand=True, fill=BOTH, padx=40)
    
    # 2. Logo (Si existe)
    if logo_path and os.path.exists(logo_path):
        try:
            img = Image.open(logo_path)
            img = img.resize((80, 80), Image.LANCZOS)
            img_tk = ImageTk.PhotoImage(img)
            lbl_img = tk.Label(fr_centro, image=img_tk, bg=COLOR_FONDO)
            lbl_img.image = img_tk # Referencia vital
            lbl_img.pack(pady=(30, 15))
        except: pass

    # 3. Títulos
    lbl_titulo = tk.Label(fr_centro, text=tit, font=("Segoe UI", 24, "bold"), 
                          bg=COLOR_FONDO, fg=COLOR_TITULO)
    lbl_titulo.pack(pady=(0, 5))
    
    # TRUCO: Para simular el espaciado elegante, usamos espacios entre letras
    sub_espaciado = " ".join(list(sub.upper()))
    
    lbl_sub = tk.Label(fr_centro, text=sub_espaciado, font=("Segoe UI", 9, "bold"), 
                       bg=COLOR_FONDO, fg=COLOR_SUBTITULO)
    lbl_sub.pack(pady=(0, 40))

    # 4. Barra de Progreso (Canvas)
    canvas_ancho = 500
    canvas_alto = 4
    canvas = tk.Canvas(fr_centro, width=canvas_ancho, height=canvas_alto, 
                       bg=COLOR_BARRA_FONDO, highlightthickness=0)
    canvas.pack(pady=(0, 10))
    
    # Crear el rectángulo de progreso
    barra = canvas.create_rectangle(0, 0, 0, canvas_alto, fill=COLOR_ACCENTO, width=0)
    
    # 5. Texto de estado
    lbl_estado = tk.Label(fr_centro, text="Iniciando...", 
                          font=("Segoe UI", 9), bg=COLOR_FONDO, fg=COLOR_SUBTITULO)
    lbl_estado.pack()

    # --- ANIMACIÓN ---
    mensajes = [
        "Cargando configuración...",
        "Verificando base de datos...",
        "Preparando interfaz...",
        "Listo..."
    ]
    
    progreso_actual = 0
    paso = 2 # Velocidad de avance de la barra
    
    def animar():
        nonlocal progreso_actual
        if progreso_actual < canvas_ancho:
            progreso_actual += paso
            canvas.coords(barra, 0, 0, progreso_actual, canvas_alto)
            
            # Cambiar texto
            porcentaje = (progreso_actual / canvas_ancho) * 100
            idx = int((porcentaje / 100) * (len(mensajes) - 1))
            lbl_estado.config(text=mensajes[idx])
            
            splash.after(15, animar) # 15ms por frame
        else:
            splash.after(400, lambda: [splash.destroy(), callback()])

    splash.after(200, animar)

def iniciar_sistema(root_login, db_manager, usuario_data):
    """Cierra login y abre sistema principal"""
    root_login.withdraw()  # Ocultar en lugar de destruir
    
    root_main = ttk.Window(themename="flatly")
    app = SistemaInventario(root_main, db_manager, usuario_data)
    
    # Cuando se cierre el sistema principal, cerrar todo
    def al_cerrar():
        root_main.destroy()
        root_login.destroy()
    
    root_main.protocol("WM_DELETE_WINDOW", al_cerrar)
    root_main.mainloop()


if __name__ == "__main__":
    print("--- INICIANDO SISTEMA ---")
    
    # 1. Crear ventana principal oculta (Para que los menús emergentes tengan un padre)
    app = tb.Window(themename="flatly")
    app.withdraw() # MANTENER OCULTA, evita el cuadro blanco
    app.title("Sistema de Inventario")

    # 2. PREPARACIÓN Y VALIDACIÓN DE BD
    archivo_config = "config_conexion.json"  
    ruta_db = None
    
    # A) Leer archivo de configuración de conexión
    if os.path.exists(archivo_config):
        try:
            with open(archivo_config, 'r') as f:
                data = json.load(f)
                posible_ruta = data.get("ruta_db")
                if posible_ruta and os.path.exists(posible_ruta):
                    ruta_db = posible_ruta
        except: pass

    # B) Si no hay archivo config o la ruta no existe (NAS desconectado)
    if not ruta_db or not os.path.exists(ruta_db):
        if 'SelectorDB' in globals():
            # NO usamos app.deiconify() aquí, para que no salga el cuadro blanco
            ruta_seleccionada = SelectorDB.abrir_ventana_seleccion(app)
            if ruta_seleccionada:
                ruta_db = ruta_seleccionada
            else:
                print("Operación cancelada por el usuario.")
                sys.exit()
        else:
            ruta_db = "inventario_unindetec.db"

    # 3. INSTANCIAR GESTOR
    db_temp = GestorBaseDatos(ruta_db)
    
    # 4. LEER CONFIGURACIÓN VISUAL DESDE LA BD
    tema_guardado = db_temp.get_config("TEMA_BOOTSTRAP") or "flatly"
    fuente_guardada = db_temp.get_config("FUENTE_SISTEMA") or "Segoe UI"
    tamano_fuente = 10 

    app.style.theme_use(tema_guardado)
    # Aplicar estilos completos del tema
    tema_completo = GestorTemas.get_tema_actual(db_temp)
    GestorTemas.aplicar_estilos_completos(app, tema_completo)
    
    # 5. APLICAR FUENTE
    estilo = ttk.Style()
    estilo.configure(".", font=(fuente_guardada, tamano_fuente))
    estilo.configure("Treeview.Heading", font=(fuente_guardada, tamano_fuente, "bold"))
    estilo.configure("Treeview", font=(fuente_guardada, tamano_fuente))
    
    w, h = 500, 400
    ws, hs = app.winfo_screenwidth(), app.winfo_screenheight()
    app.geometry(f"{w}x{h}+{int((ws/2)-(w/2))}+{int((hs/2)-(h/2))}")

    # 6. FUNCIONES DE FLUJO
    def iniciar_app_principal(usuario_data):
        """Cierra Login y MUESTRA el Sistema Principal"""
        for widget in app.winfo_children(): widget.destroy()
        app.withdraw() 
        
        sistema = SistemaInventario(app, db_temp, usuario_data)
        
        try:
            app.state('zoomed') 
        except:
            app.attributes('-zoomed', True)

        app.update_idletasks()
        app.deiconify()

    def mostrar_login():
        """Muestra la pantalla de Login"""
        for widget in app.winfo_children(): widget.destroy()
        try:
            LoginWindow(app, db_temp, iniciar_app_principal)
        except Exception as e:
            messagebox.showerror("Error", f"Error iniciando Login: {e}")

    mostrar_login()
    app.mainloop()